import difflib
import io
from datetime import date, datetime, timedelta, timezone

import openpyxl
from flask import (
    Blueprint,
    Response,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import extract, func, or_
from sqlalchemy.orm import joinedload

from models import db
from models.contact import Email, LocalStatusEnum
from models.district_model import District
from models.event import Event, EventFormat, EventStatus, EventTeacher, EventType
from models.google_sheet import GoogleSheet
from models.organization import Organization, VolunteerOrganization
from models.reports import VirtualSessionDistrictCache, VirtualSessionReportCache
from models.school_model import School
from models.teacher import Teacher
from models.teacher_progress import TeacherProgress
from models.volunteer import EventParticipation, Volunteer
from routes.decorators import district_scoped_required
from routes.utils import admin_required

# Import blueprint from virtual routes
from routes.virtual.routes import virtual_bp

# --- Helper Functions for School Year ---


def get_current_school_year() -> str:
    """Determines the current school year string (e.g., '2024-2025')."""
    today = date.today()
    if today.month >= 6:  # School year starts in June
        start_year = today.year
    else:
        start_year = today.year - 1
    return f"{start_year}-{start_year + 1}"


def get_school_year_dates(school_year: str) -> tuple[datetime, datetime]:
    """Calculates the start and end dates for a given school year string."""
    try:
        start_year = int(school_year.split("-")[0])
        end_year = start_year + 1
        date_from = datetime(start_year, 6, 1, 0, 0, 0)
        date_to = datetime(end_year, 5, 31, 23, 59, 59)
        return date_from, date_to
    except (ValueError, IndexError):
        current_sy = get_current_school_year()
        return get_school_year_dates(current_sy)


def generate_school_year_options(start_cal_year=2018, end_cal_year=None) -> list[str]:
    """Generates a list of school year strings for dropdowns."""
    if end_cal_year is None:
        end_cal_year = date.today().year + 1

    school_years = []
    for year in range(end_cal_year, start_cal_year - 1, -1):
        school_years.append(f"{year}-{year + 1}")
    return school_years


def get_current_virtual_year() -> str:
    """Determines the current virtual session year string (August 1st to July 31st)."""
    today = date.today()
    if today.month > 7 or (today.month == 7 and today.day == 31):  # After July 31st
        start_year = today.year
    else:
        start_year = today.year - 1
    return f"{start_year}-{start_year + 1}"


def get_virtual_year_dates(virtual_year: str) -> tuple[datetime, datetime]:
    """Calculates the start and end dates for a given virtual session year string (8/1 to 7/31)."""
    try:
        start_year = int(virtual_year.split("-")[0])
        end_year = start_year + 1
        date_from = datetime(start_year, 8, 1, 0, 0, 0)  # August 1st start
        date_to = datetime(end_year, 7, 31, 23, 59, 59)  # July 31st end
        return date_from, date_to
    except (ValueError, IndexError):
        current_vy = get_current_virtual_year()
        return get_virtual_year_dates(current_vy)


# --- End Helper Functions ---

# --- Cache Management Functions ---


def is_cache_valid(cache_record, max_age_hours=24):
    """
    Check if a cache record is still valid based on age.

    Args:
        cache_record: The cache record to check
        max_age_hours: Maximum age in hours before cache is considered stale

    Returns:
        bool: True if cache is valid, False otherwise
    """
    if not cache_record:
        return False

    from datetime import datetime, timedelta, timezone

    # Convert to timezone-aware datetime if needed
    if cache_record.last_updated.tzinfo is None:
        last_updated = cache_record.last_updated.replace(tzinfo=timezone.utc)
    else:
        last_updated = cache_record.last_updated

    now = datetime.now(timezone.utc)
    max_age = timedelta(hours=max_age_hours)

    return (now - last_updated) < max_age


def get_virtual_session_cache(virtual_year, date_from=None, date_to=None):
    """
    Get cached virtual session report data.

    Args:
        virtual_year: The virtual year to get cache for
        date_from: Optional start date filter
        date_to: Optional end date filter

    Returns:
        VirtualSessionReportCache or None if not found/expired
    """
    cache_query = VirtualSessionReportCache.query.filter_by(
        virtual_year=virtual_year,
        date_from=date_from.date() if date_from else None,
        date_to=date_to.date() if date_to else None,
    )

    cache_record = cache_query.first()

    if is_cache_valid(cache_record):
        return cache_record

    return None


def save_virtual_session_cache(
    virtual_year,
    date_from,
    date_to,
    session_data,
    district_summaries,
    overall_summary,
    filter_options,
):
    """
    Save virtual session report data to cache.

    Args:
        virtual_year: The virtual year
        date_from: Start date filter
        date_to: End date filter
        session_data: Session data to cache
        district_summaries: District summary data
        overall_summary: Overall summary data
        filter_options: Filter options data
    """
    try:
        # Check if cache record exists
        cache_record = VirtualSessionReportCache.query.filter_by(
            virtual_year=virtual_year,
            date_from=date_from.date() if date_from else None,
            date_to=date_to.date() if date_to else None,
        ).first()

        if cache_record:
            # Update existing record
            cache_record.session_data = session_data
            cache_record.district_summaries = district_summaries
            cache_record.overall_summary = overall_summary
            cache_record.filter_options = filter_options
            cache_record.last_updated = datetime.now(timezone.utc)
        else:
            # Create new record
            cache_record = VirtualSessionReportCache(
                virtual_year=virtual_year,
                date_from=date_from.date() if date_from else None,
                date_to=date_to.date() if date_to else None,
                session_data=session_data,
                district_summaries=district_summaries,
                overall_summary=overall_summary,
                filter_options=filter_options,
            )
            db.session.add(cache_record)

        db.session.commit()
        print(f"Virtual session cache saved for {virtual_year}")

    except Exception as e:
        db.session.rollback()
        print(f"Error saving virtual session cache: {str(e)}")


def get_virtual_session_district_cache(
    district_name, virtual_year, date_from=None, date_to=None
):
    """
    Get cached district virtual session report data.

    Args:
        district_name: The district name
        virtual_year: The virtual year to get cache for
        date_from: Optional start date filter
        date_to: Optional end date filter

    Returns:
        VirtualSessionDistrictCache or None if not found/expired
    """
    cache_query = VirtualSessionDistrictCache.query.filter_by(
        district_name=district_name,
        virtual_year=virtual_year,
        date_from=date_from.date() if date_from else None,
        date_to=date_to.date() if date_to else None,
    )

    cache_record = cache_query.first()

    if is_cache_valid(cache_record):
        return cache_record

    return None


def save_virtual_session_district_cache(
    district_name,
    virtual_year,
    date_from,
    date_to,
    session_data,
    monthly_stats,
    school_breakdown,
    teacher_breakdown,
    summary_stats,
):
    """
    Save district virtual session report data to cache.

    Args:
        district_name: The district name
        virtual_year: The virtual year
        date_from: Start date filter
        date_to: End date filter
        session_data: Session data to cache
        monthly_stats: Monthly statistics
        school_breakdown: School breakdown data
        teacher_breakdown: Teacher breakdown data
        summary_stats: Summary statistics
    """
    try:
        # Check if cache record exists
        cache_record = VirtualSessionDistrictCache.query.filter_by(
            district_name=district_name,
            virtual_year=virtual_year,
            date_from=date_from.date() if date_from else None,
            date_to=date_to.date() if date_to else None,
        ).first()

        if cache_record:
            # Update existing record
            cache_record.session_data = session_data
            cache_record.monthly_stats = monthly_stats
            cache_record.school_breakdown = school_breakdown
            cache_record.teacher_breakdown = teacher_breakdown
            cache_record.summary_stats = summary_stats
            cache_record.last_updated = datetime.now(timezone.utc)
        else:
            # Create new record
            cache_record = VirtualSessionDistrictCache(
                district_name=district_name,
                virtual_year=virtual_year,
                date_from=date_from.date() if date_from else None,
                date_to=date_to.date() if date_to else None,
                session_data=session_data,
                monthly_stats=monthly_stats,
                school_breakdown=school_breakdown,
                teacher_breakdown=teacher_breakdown,
                summary_stats=summary_stats,
            )
            db.session.add(cache_record)

        db.session.commit()
        print(
            f"Virtual session district cache saved for {district_name} {virtual_year}"
        )

    except Exception as e:
        db.session.rollback()
        print(f"Error saving virtual session district cache: {str(e)}")


def invalidate_virtual_session_caches(virtual_year=None):
    """
    Invalidate virtual session caches for a specific year or all years.

    Args:
        virtual_year: Specific year to invalidate, or None for all years
    """
    try:
        if virtual_year:
            # Invalidate specific year
            VirtualSessionReportCache.query.filter_by(
                virtual_year=virtual_year
            ).delete()
            VirtualSessionDistrictCache.query.filter_by(
                virtual_year=virtual_year
            ).delete()
        else:
            # Invalidate all caches
            VirtualSessionReportCache.query.delete()
            VirtualSessionDistrictCache.query.delete()

        db.session.commit()
        print(f"Virtual session caches invalidated for {virtual_year or 'all years'}")

    except Exception as e:
        db.session.rollback()
        print(f"Error invalidating virtual session caches: {str(e)}")


# --- End Cache Management Functions ---

# --- Data Processing Helper Functions ---


def apply_runtime_filters(session_data, filters):
    """
    Apply runtime filters to session data.

    Args:
        session_data: List of session records
        filters: Dictionary of filter criteria

    Returns:
        Filtered session data
    """
    filtered_data = []

    for session in session_data:
        # Apply career cluster filter
        if filters.get("career_cluster"):
            if (
                not session.get("topic_theme")
                or filters["career_cluster"].lower()
                not in session["topic_theme"].lower()
            ):
                continue

        # Apply school filter
        if filters.get("school"):
            if (
                not session.get("school_name")
                or filters["school"].lower() not in session["school_name"].lower()
            ):
                continue

        # Apply district filter
        if filters.get("district"):
            if session.get("district") != filters["district"]:
                continue

        # Apply status filter (case-insensitive with mapping)
        if filters.get("status"):
            session_status = session.get("status", "").strip().lower()
            filter_status = filters["status"].strip().lower()

            # Map common status variations
            status_mapping = {
                "completed": [
                    "completed",
                    "successfully completed",
                    "white lable completed",
                ],
                "requested": [
                    "requested",
                    "teacher requested",
                ],
                "simulcast": ["simulcast"],
                "no show": [
                    "no show",
                    "teacher no-show",
                    "local professional no-show",
                    "pathful professional no-show",
                    "local professional no show",
                ],
                "cancelled": [
                    "cancelled",
                    "teacher cancelation",
                    "local professional cancellation",
                    "pathful professional cancellation",
                    "inclement weather cancellation",
                    "technical difficulties",
                ],
                "draft": [
                    "draft",
                    "moved to in-person session",
                    "unfilled",
                    "registered",
                    "count",
                    "white lable unfilled",
                ],
            }

            # Check if the filter status matches any of the mapped statuses
            filter_matched = False
            for mapped_status, variations in status_mapping.items():
                if filter_status == mapped_status and session_status in variations:
                    filter_matched = True
                    break

            if not filter_matched and session_status != filter_status:
                continue

        # Apply text search filter (searches teacher name, session title, and presenter)
        if filters.get("search"):
            search_term = filters["search"].strip().lower()
            if search_term:
                # Get field values, handling None and empty strings
                teacher_name = str(session.get("teacher_name") or "").lower()
                session_title = str(session.get("session_title") or "").lower()
                presenter = str(session.get("presenter") or "").lower()

                # Check if search term matches any of the fields (case-insensitive substring match)
                if (
                    search_term not in teacher_name
                    and search_term not in session_title
                    and search_term not in presenter
                ):
                    continue

        filtered_data.append(session)

    return filtered_data


def calculate_summaries_from_sessions(session_data, show_all_districts=False):
    """
    Calculate district summaries and overall summary from session data.
    Only counts sessions with status "Completed" or "Simulcast".

    Args:
        session_data: List of session records
        show_all_districts: If True, show all districts. If False, only show main districts.

    Returns:
        Tuple of (district_summaries, overall_summary)
    """
    # Include all districts that have data
    # Note: We don't filter districts here anymore - show all districts with data

    print(
        f"DEBUG: calculate_summaries_from_sessions called with {len(session_data)} sessions"
    )
    print(f"DEBUG: show_all_districts = {show_all_districts}")

    district_summaries = {}
    overall_stats = {
        "teacher_count": set(),  # Changed to set for unique counting
        "student_count": 0,
        "session_count": set(),
        "experience_count": 0,  # Changed back to integer for counting every row
        "organization_count": set(),
        "professional_count": set(),
        "professional_of_color_count": set(),
        "local_professional_count": set(),
        "local_sessions": set(),
        "poc_sessions": set(),
        "school_count": set(),
    }

    # First pass: initialize all districts that have any sessions
    districts_found = set()
    for session in session_data:
        if session["district"]:
            districts_found.add(session["district"])
            if session["district"] not in district_summaries:
                district_summaries[session["district"]] = {
                    "teachers": set(),
                    "schools": set(),
                    "sessions": set(),
                    "total_students": 0,
                    "total_experiences": 0,
                    "organizations": set(),
                    "professionals": set(),
                    "professionals_of_color": set(),
                    "local_professionals": set(),
                }

    print(f"DEBUG: Districts found in session_data: {sorted(list(districts_found))}")
    print(
        f"DEBUG: district_summaries initialized with keys: {sorted(list(district_summaries.keys()))}"
    )

    # Second pass: process only completed sessions for counting
    completed_sessions = 0
    for session in session_data:
        # Only count sessions with completed status (case-insensitive)
        session_status = session.get("status", "").strip().lower()
        if session_status not in ["completed", "simulcast", "successfully completed"]:
            continue

        completed_sessions += 1
        if session["district"]:
            district_summary = district_summaries[session["district"]]

            # Count unique teachers (only from completed sessions)
            if session["teacher_name"]:
                district_summary["teachers"].add(session["teacher_name"])
                overall_stats["teacher_count"].add(session["teacher_name"])

            # Count unique schools
            if session["school_name"]:
                district_summary["schools"].add(session["school_name"])
                overall_stats["school_count"].add(session["school_name"])

            # Count unique sessions (prefer id when available)
            if session.get("session_title"):
                district_summary["sessions"].add(session["session_title"])
                overall_stats["session_count"].add(session["session_title"])
            if session.get("event_id"):
                # Attach id sets lazily to avoid key errors if older cache present
                if "session_ids" not in district_summary:
                    district_summary["session_ids"] = set()
                district_summary["session_ids"].add(session["event_id"])
                overall_stats.setdefault("session_ids", set()).add(session["event_id"])

            # Count experiences (every row counts as an experience)
            district_summary["total_experiences"] += 1
            overall_stats["experience_count"] += 1

            # Student count will be calculated after we have unique teacher count

            # Process presenters and check for People of Color; also mark session as Local/POC
            if session["presenter_data"]:
                for presenter_data in session["presenter_data"]:
                    presenter_name = presenter_data.get("name", "")
                    presenter_id = presenter_data.get("id")
                    if presenter_name:
                        # Count unique professionals by ID if available, otherwise by name
                        if presenter_id:
                            district_summary["professionals"].add(presenter_id)
                            overall_stats["professional_count"].add(presenter_id)
                        else:
                            district_summary["professionals"].add(presenter_name)
                            overall_stats["professional_count"].add(presenter_name)

                        # Count only primary organization from this presenter
                        organization_name = presenter_data.get("organization_name")
                        if organization_name:
                            district_summary["organizations"].add(organization_name)
                            overall_stats["organization_count"].add(organization_name)

                        # Check if this presenter is marked as People of Color
                        if presenter_data.get("is_people_of_color", False):
                            if presenter_id:
                                district_summary["professionals_of_color"].add(
                                    presenter_id
                                )
                                overall_stats["professional_of_color_count"].add(
                                    presenter_id
                                )
                            else:
                                district_summary["professionals_of_color"].add(
                                    presenter_name
                                )
                                overall_stats["professional_of_color_count"].add(
                                    presenter_name
                                )
                        # Count local professionals (based on derived is_local flag)
                        if presenter_data.get("is_local"):
                            if presenter_id:
                                district_summary["local_professionals"].add(
                                    presenter_id
                                )
                                overall_stats["local_professional_count"].add(
                                    presenter_id
                                )
                            else:
                                district_summary["local_professionals"].add(
                                    presenter_name
                                )
                                overall_stats["local_professional_count"].add(
                                    presenter_name
                                )
                        # Count local professionals (based on volunteer.local_status)
                        try:
                            # We don't have local flag in presenter_data today; derive from volunteer.id later if needed
                            # For now, assume presenter_data includes a boolean 'is_local' when set by import
                            is_local = presenter_data.get("is_local")
                            if is_local:
                                if presenter_id:
                                    overall_stats["local_professional_count"].add(
                                        presenter_id
                                    )
                                else:
                                    overall_stats["local_professional_count"].add(
                                        presenter_name
                                    )
                        except Exception:
                            pass

                # Session-level flags (local / POC sessions)
                # Prefer event_id for consistency with session counting logic
                sid_flag = session.get("event_id") or session.get("session_title")
                if sid_flag:
                    if any(p.get("is_local") for p in session["presenter_data"]):
                        overall_stats["local_sessions"].add(sid_flag)
                        # mirror per-district; initialize lazily
                        if "local_sessions" not in district_summary:
                            district_summary["local_sessions"] = set()
                        district_summary["local_sessions"].add(sid_flag)
                    if any(
                        p.get("is_people_of_color") for p in session["presenter_data"]
                    ):
                        overall_stats["poc_sessions"].add(sid_flag)
                        if "poc_sessions" not in district_summary:
                            district_summary["poc_sessions"] = set()
                        district_summary["poc_sessions"].add(sid_flag)
            elif session["presenter"]:
                # Fallback to old presenter format
                presenters = [p.strip() for p in session["presenter"].split(",")]
                for presenter in presenters:
                    if presenter:
                        district_summary["professionals"].add(presenter)
                        overall_stats["professional_count"].add(presenter)
                        # For old format, we can't determine POC status, so don't count them

    # Convert sets to counts and filter allowed districts
    for district_name, summary in district_summaries.items():
        summary["teacher_count"] = len(summary["teachers"])
        summary["school_count"] = len(summary["schools"])
        # Prefer event-id based counting if present
        summary["session_count"] = (
            len(summary.get("session_ids", set()))
            if summary.get("session_ids")
            else len(summary["sessions"])
        )
        summary["organization_count"] = len(summary["organizations"])
        summary["professional_count"] = len(summary["professionals"])
        summary["professional_of_color_count"] = len(summary["professionals_of_color"])
        summary["local_professional_count"] = len(summary["local_professionals"])
        # Session-level flags
        summary["local_session_count"] = len(summary.get("local_sessions", set()))
        summary["poc_session_count"] = len(summary.get("poc_sessions", set()))
        if summary["session_count"]:
            summary["local_session_percent"] = round(
                100 * summary["local_session_count"] / summary["session_count"]
            )
            summary["poc_session_percent"] = round(
                100 * summary["poc_session_count"] / summary["session_count"]
            )
        else:
            summary["local_session_percent"] = 0
            summary["poc_session_percent"] = 0

        # Calculate student count as unique teachers × 25
        summary["total_students"] = summary["teacher_count"] * 25

        del summary["teachers"]
        del summary["schools"]
        del summary["sessions"]
        if "session_ids" in summary:
            del summary["session_ids"]
        del summary["organizations"]
        del summary["professionals"]
        del summary["professionals_of_color"]
        del summary["local_professionals"]
        if "local_sessions" in summary:
            del summary["local_sessions"]
        if "poc_sessions" in summary:
            del summary["poc_sessions"]

    # Filter to only show main districts by default (unless admin requests all)
    if not show_all_districts:
        main_districts = {
            "Hickman Mills School District",
            "Grandview School District",
            "Kansas City Kansas Public Schools",
        }
        # Only show main districts when show_all_districts=False
        district_summaries = {
            k: v for k, v in district_summaries.items() if k in main_districts
        }

    # Calculate overall student count as unique teachers × 25
    unique_teacher_count = len(overall_stats["teacher_count"])
    overall_student_count = unique_teacher_count * 25

    # Prepare overall summary stats
    # Use session_ids (event_id) if available for session counting, otherwise fall back to session_count
    overall_session_count = (
        len(overall_stats.get("session_ids", set()))
        if overall_stats.get("session_ids")
        else len(overall_stats["session_count"])
    )

    overall_summary = {
        "teacher_count": unique_teacher_count,
        "student_count": overall_student_count,
        "session_count": overall_session_count,
        "experience_count": overall_stats["experience_count"],  # Use integer value
        "organization_count": len(overall_stats["organization_count"]),
        "professional_count": len(overall_stats["professional_count"]),
        "professional_of_color_count": len(
            overall_stats["professional_of_color_count"]
        ),
        "local_professional_count": len(overall_stats["local_professional_count"]),
        "school_count": len(overall_stats["school_count"]),
        "local_session_count": len(overall_stats["local_sessions"]),
        "poc_session_count": len(overall_stats["poc_sessions"]),
    }

    # Percentages overall - use session_ids for denominator if available
    denom = (
        len(overall_stats.get("session_ids", set()))
        if overall_stats.get("session_ids")
        else (overall_summary["session_count"] or 1)
    )
    if denom == 0:
        denom = 1
    overall_summary["local_session_percent"] = (
        round(100 * overall_summary["local_session_count"] / denom)
        if overall_summary["local_session_count"]
        else 0
    )
    overall_summary["poc_session_percent"] = (
        round(100 * overall_summary["poc_session_count"] / denom)
        if overall_summary["poc_session_count"]
        else 0
    )

    print(
        f"DEBUG: calculate_summaries_from_sessions returning {len(district_summaries)} districts: {sorted(list(district_summaries.keys()))}"
    )
    return district_summaries, overall_summary


def apply_sorting_and_pagination(session_data, request_args, current_filters):
    """
    Apply sorting and pagination to session data.

    Args:
        session_data: List of session records
        request_args: Request arguments for sorting/pagination
        current_filters: Current filter settings

    Returns:
        Dictionary with paginated_data and pagination info
    """
    # Apply sorting
    sort_column = request_args.get("sort", "date")
    sort_order = request_args.get("order", "desc")

    # Define sortable columns mapping
    sortable_columns = {
        "status": "status",
        "date": "date",
        "time": "time",
        "session_type": "session_type",
        "teacher_name": "teacher_name",
        "school_name": "school_name",
        "school_level": "school_level",
        "district": "district",
        "session_title": "session_title",
        "presenter": "presenter",
        "presenter_organization": "presenter_organization",
        "topic_theme": "topic_theme",
    }

    # Apply sorting if valid column
    if sort_column in sortable_columns:
        reverse_order = sort_order == "desc"

        # Handle special sorting for date/time columns
        if sort_column == "date":

            def date_sort_key(session):
                try:
                    if session["date"]:
                        date_parts = session["date"].split("/")
                        if len(date_parts) == 3:
                            month = int(date_parts[0])
                            day = int(date_parts[1])
                            year = int(date_parts[2])
                            if year < 50:
                                year += 2000
                            else:
                                year += 1900
                            return datetime(year, month, day)
                        elif len(date_parts) == 2:
                            month = int(date_parts[0])
                            day = int(date_parts[1])
                            virtual_year_start = int(
                                current_filters["year"].split("-")[0]
                            )
                            if month >= 7:
                                year = virtual_year_start
                            else:
                                year = virtual_year_start + 1
                            return datetime(year, month, day)
                    return datetime.min
                except (ValueError, IndexError):
                    return datetime.min

            session_data.sort(key=date_sort_key, reverse=reverse_order)

        elif sort_column == "time":

            def time_sort_key(session):
                try:
                    if session["time"]:
                        return datetime.strptime(session["time"], "%I:%M %p").time()
                    return datetime.min.time()
                except ValueError:
                    return datetime.min.time()

            session_data.sort(key=time_sort_key, reverse=reverse_order)

        else:

            def string_sort_key(session):
                value = session.get(sortable_columns[sort_column], "") or ""
                return str(value).lower()

            session_data.sort(key=string_sort_key, reverse=reverse_order)

    # Add sorting info to current_filters for template
    current_filters["sort"] = sort_column
    current_filters["order"] = sort_order

    # Apply pagination
    try:
        page = int(request_args.get("page", 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    try:
        per_page = int(request_args.get("per_page", 25))
        if per_page < 1:
            per_page = 25
    except ValueError:
        per_page = 25

    total_records = len(session_data)
    total_pages = (total_records + per_page - 1) // per_page
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_data = session_data[start_idx:end_idx]

    pagination = {
        "current_page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "total_records": total_records,
    }

    return {"paginated_data": paginated_data, "pagination": pagination}


def get_google_sheet_url(virtual_year):
    """
    Get Google Sheet URL for a virtual year.

    Args:
        virtual_year: The virtual year to get sheet for

    Returns:
        Google Sheet URL or None
    """
    google_sheet = GoogleSheet.query.filter_by(
        academic_year=virtual_year, purpose="virtual_sessions"
    ).first()
    if google_sheet and google_sheet.decrypted_sheet_id:
        return (
            f"https://docs.google.com/spreadsheets/d/{google_sheet.decrypted_sheet_id}"
        )
    return None


def _get_primary_org_name_for_volunteer(volunteer):
    """
    Get the primary organization name for a volunteer.
    Prefer explicit primary org from VolunteerOrganization, fall back to organization_name field.
    Uses the same pattern as routes/reports/volunteers_by_event.py:_get_primary_org_name
    """
    # Prefer explicit primary org from association table
    if getattr(volunteer, "volunteer_organizations", None):
        try:
            primary = next(
                (vo for vo in volunteer.volunteer_organizations if vo.is_primary), None
            )
            if primary and primary.organization and primary.organization.name:
                return primary.organization.name
            # Fall back to first org if no primary
            first = (
                volunteer.volunteer_organizations[0].organization
                if volunteer.volunteer_organizations
                else None
            )
            if first and first.name:
                return first.name
        except (StopIteration, AttributeError, IndexError):
            pass
    # Avoid showing raw Salesforce IDs stored in organization_name
    org_name = getattr(volunteer, "organization_name", None)
    if org_name and not (len(str(org_name)) == 18 and str(org_name).isalnum()):
        return org_name
    return None


def compute_virtual_session_data(virtual_year, date_from, date_to, filters):
    """
    Compute virtual session data from database.

    Args:
        virtual_year: The virtual year
        date_from: Start date
        date_to: End date
        filters: Filter criteria

    Returns:
        Tuple of (session_data, district_summaries, overall_summary, filter_options)
    """
    # Base query for virtual session events
    from sqlalchemy.orm import selectinload

    from models import eagerload_event_bundle
    from models.organization import VolunteerOrganization

    base_query = (
        eagerload_event_bundle(Event.query)
        .options(
            selectinload(Event.volunteers)
            .selectinload(Volunteer.volunteer_organizations)
            .selectinload(VolunteerOrganization.organization)
        )
        .filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )
    )

    # Apply database-level filters
    if filters.get("career_cluster"):
        base_query = base_query.filter(
            Event.series.ilike(f'%{filters["career_cluster"]}%')
        )

    if filters.get("status"):
        base_query = base_query.filter(Event.status == filters["status"])

    # Get all events
    events = base_query.order_by(Event.start_date.desc()).all()

    # First pass: collect all districts, schools, career clusters, and statuses from raw events
    all_districts = set()
    all_schools = set()
    all_career_clusters = set()
    all_statuses = set()

    for event in events:
        # Collect districts from events
        if event.districts:
            all_districts.add(event.districts[0].name)
        elif event.district_partner:
            all_districts.add(event.district_partner)

        # Collect career clusters and statuses
        if event.series:
            all_career_clusters.add(event.series)
        if event.status:
            all_statuses.add(event.status.value)

        # Collect schools and their districts from teacher registrations
        for teacher_reg in event.teacher_registrations:
            teacher = teacher_reg.teacher
            if teacher and teacher.school_id:
                school = School.query.get(teacher.school_id)
                if school:
                    all_schools.add(school.name)
                    # Also add the school's district if available
                    if hasattr(school, "district") and school.district:
                        all_districts.add(school.district.name)

    # Build session data
    session_data = []

    for event in events:
        # Get all teacher registrations for this event
        teacher_registrations = event.teacher_registrations

        if teacher_registrations:
            # Create a row for each teacher registration
            for teacher_reg in teacher_registrations:
                teacher = teacher_reg.teacher

                # Get school information
                school = None
                school_name = ""
                school_level = ""

                if teacher:
                    if hasattr(teacher, "school_obj") and teacher.school_obj:
                        school = teacher.school_obj
                        school_name = school.name
                        school_level = getattr(school, "level", "")
                    elif teacher.school_id:
                        school = School.query.get(teacher.school_id)
                        if school:
                            school_name = school.name
                            school_level = getattr(school, "level", "")

                # Determine district from multiple sources
                # Prioritize teacher's school district over event's district
                district_name = None
                if school and hasattr(school, "district") and school.district:
                    # Teacher's school district takes priority
                    district_name = school.district.name
                elif event.districts:
                    district_name = event.districts[0].name
                elif event.district_partner:
                    district_name = event.district_partner
                else:
                    district_name = "Unknown District"

                # Apply district filter if specified
                if filters.get("district") and district_name != filters["district"]:
                    continue

                # Apply school filter if specified
                if (
                    filters.get("school")
                    and school_name
                    and filters["school"].lower() not in school_name.lower()
                ):
                    continue

                session_data.append(
                    {
                        "event_id": event.id,
                        "source_host": getattr(event, "session_host", "") or "",
                        # Use the event's status for session-level filtering (completed/simulcast),
                        # not the individual teacher registration status
                        "status": (
                            event.status.value if getattr(event, "status", None) else ""
                        )
                        or "registered",
                        "date": (
                            event.start_date.strftime("%m/%d/%y")
                            if event.start_date
                            else ""
                        ),
                        "time": (
                            event.start_date.strftime("%I:%M %p")
                            if event.start_date
                            else ""
                        ),
                        "session_type": event.additional_information or "",
                        "teacher_name": (
                            f"{teacher.first_name} {teacher.last_name}"
                            if teacher
                            else ""
                        ),
                        "teacher_id": teacher.id if teacher else None,
                        "school_name": school_name,
                        "school_level": school_level,
                        "district": district_name,
                        "session_title": event.title,
                        "presenter": (
                            ", ".join([v.full_name for v in event.volunteers])
                            if event.volunteers
                            else ""
                        ),
                        "presenter_organization": (
                            ", ".join(
                                [
                                    _get_primary_org_name_for_volunteer(v)
                                    or "Independent"
                                    for v in event.volunteers
                                ]
                            )
                            if event.volunteers
                            else ""
                        ),
                        "presenter_data": (
                            [
                                {
                                    "id": v.id,
                                    "name": v.full_name,
                                    "is_people_of_color": v.is_people_of_color,
                                    "organization_name": v.organization_name,
                                    "organizations": (
                                        [org.name for org in v.organizations]
                                        if v.organizations
                                        else []
                                    ),
                                    "is_local": getattr(v, "local_status", None)
                                    == LocalStatusEnum.local,
                                }
                                for v in event.volunteers
                            ]
                            if event.volunteers
                            else []
                        ),
                        "topic_theme": event.series or "",
                        "session_link": event.registration_link or "",
                        "session_id": event.session_id or "",
                        "participant_count": event.participant_count or 0,
                        "duration": event.duration or 0,
                        "is_simulcast": teacher_reg.is_simulcast,
                    }
                )

        else:
            # Event with no teacher registrations - show the event itself
            district_name = None
            if event.districts:
                district_name = event.districts[0].name
            elif event.district_partner:
                district_name = event.district_partner
            else:
                district_name = "Unknown District"

            # Apply district filter if specified
            if filters.get("district") and district_name != filters["district"]:
                continue

            session_data.append(
                {
                    "event_id": event.id,
                    "source_host": getattr(event, "session_host", "") or "",
                    "status": event.status.value if event.status else "",
                    "date": (
                        event.start_date.strftime("%m/%d/%y")
                        if event.start_date
                        else ""
                    ),
                    "time": (
                        event.start_date.strftime("%I:%M %p")
                        if event.start_date
                        else ""
                    ),
                    "session_type": event.additional_information or "",
                    "teacher_name": "",
                    "teacher_id": None,
                    "school_name": "",
                    "school_level": "",
                    "district": district_name,
                    "session_title": event.title,
                    "presenter": (
                        ", ".join([v.full_name for v in event.volunteers])
                        if event.volunteers
                        else ""
                    ),
                    "presenter_organization": (
                        ", ".join(
                            [
                                _get_primary_org_name_for_volunteer(v) or "Independent"
                                for v in event.volunteers
                            ]
                        )
                        if event.volunteers
                        else ""
                    ),
                    "presenter_data": (
                        [
                            {
                                "id": v.id,
                                "name": v.full_name,
                                "is_people_of_color": v.is_people_of_color,
                                "organization_name": v.organization_name,
                                "organizations": (
                                    [org.name for org in v.organizations]
                                    if v.organizations
                                    else []
                                ),
                                "is_local": getattr(v, "local_status", None)
                                == LocalStatusEnum.local,
                            }
                            for v in event.volunteers
                        ]
                        if event.volunteers
                        else []
                    ),
                    "topic_theme": event.series or "",
                    "session_link": event.registration_link or "",
                    "session_id": event.session_id or "",
                    "participant_count": event.participant_count or 0,
                    "duration": event.duration or 0,
                    "is_simulcast": False,
                }
            )

    # Calculate summaries
    show_all_districts = filters.get("show_all_districts", False)
    district_summaries, overall_summary = calculate_summaries_from_sessions(
        session_data, show_all_districts
    )

    # Ensure all districts with data are included in summaries
    # Only create empty summaries for districts that appear in all_districts but have no completed sessions
    # This allows showing all districts that exist, even if they have zero completed sessions
    for district_name in all_districts:
        if district_name not in district_summaries:
            # Only create empty summary if we want to show all districts
            # Otherwise, skip districts with no completed sessions
            district_summaries[district_name] = {
                "teacher_count": 0,
                "total_students": 0,
                "session_count": 0,
                "total_experiences": 0,
                "organization_count": 0,
                "professional_count": 0,
                "professional_of_color_count": 0,
                "local_professional_count": 0,
                "school_count": 0,
                "local_session_count": 0,
                "poc_session_count": 0,
                "local_session_percent": 0,
                "poc_session_percent": 0,
            }

    # Show all districts that have data, but prioritize the main districts
    main_districts = {
        "Hickman Mills School District",
        "Grandview School District",
        "Kansas City Kansas Public Schools",
    }
    # Include all districts that have data, but put main districts first
    all_districts_list = sorted(list(all_districts))
    main_districts_list = [d for d in all_districts_list if d in main_districts]
    other_districts_list = [d for d in all_districts_list if d not in main_districts]
    filtered_districts = main_districts_list + other_districts_list

    # Prepare filter options
    virtual_year_options = generate_school_year_options()
    filter_options = {
        "school_years": virtual_year_options,
        "career_clusters": sorted(list(all_career_clusters)),
        "schools": sorted(list(all_schools)),
        "districts": sorted(filtered_districts),
        "statuses": sorted(list(all_statuses)),
    }

    return session_data, district_summaries, overall_summary, filter_options


def _district_name_matches(target_district_name, compare_district_name):
    """
    Check if a district name matches the target district, including aliases.

    Args:
        target_district_name: The target district name (from URL/filter)
        compare_district_name: The district name to compare against

    Returns:
        bool: True if the districts match (including aliases)
    """
    if not target_district_name or not compare_district_name:
        return False

    # Normalize for comparison (case-insensitive, strip whitespace)
    target_normalized = target_district_name.strip().lower()
    compare_normalized = compare_district_name.strip().lower()

    # Exact match (case-insensitive)
    if target_normalized == compare_normalized:
        return True

    # Check against DISTRICT_MAPPING aliases
    from routes.reports.common import DISTRICT_MAPPING

    # Find the target district in the mapping
    target_mapping = None
    for mapping in DISTRICT_MAPPING.values():
        if mapping["name"].strip().lower() == target_normalized:
            target_mapping = mapping
            break

    if target_mapping:
        # Check if compare_district_name matches any alias
        aliases = target_mapping.get("aliases", [])
        for alias in aliases:
            if alias.strip().lower() == compare_normalized:
                return True

        # Also check reverse: if compare_district_name is the primary name and target is an alias
        if (
            compare_district_name.strip().lower()
            == target_mapping["name"].strip().lower()
        ):
            return True

    # Reverse check: maybe compare_district_name is in the mapping and target is an alias
    for mapping in DISTRICT_MAPPING.values():
        if mapping["name"].strip().lower() == compare_normalized:
            aliases = mapping.get("aliases", [])
            for alias in aliases:
                if alias.strip().lower() == target_normalized:
                    return True
            break

    return False


def compute_virtual_session_district_data(
    district_name, virtual_year, date_from, date_to
):
    """
    Compute district-specific virtual session data from database.

    Args:
        district_name: Name of the district
        virtual_year: The virtual year
        date_from: Start date
        date_to: End date

    Returns:
        Tuple of (session_data, monthly_stats, school_breakdown, teacher_breakdown, summary_stats)
    """
    # Base query for virtual session events
    from sqlalchemy.orm import selectinload

    from models import eagerload_event_bundle
    from models.organization import VolunteerOrganization

    base_query = (
        eagerload_event_bundle(Event.query)
        .options(
            selectinload(Event.volunteers)
            .selectinload(Volunteer.volunteer_organizations)
            .selectinload(VolunteerOrganization.organization)
        )
        .filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )
    )

    events = base_query.order_by(Event.start_date.desc()).all()
    session_dict = {}

    for event in events:
        # Check if any teacher in this event belongs to the target district
        event_has_target_district_teacher = False
        target_district_teachers = set()
        target_district_schools = set()

        for teacher_reg in event.teacher_registrations:
            teacher = teacher_reg.teacher
            if teacher:
                # Determine district using the same logic as main function
                teacher_district = None

                # Check teacher's school district first
                if teacher.school_id:
                    school = School.query.get(teacher.school_id)
                    if school and school.district:
                        teacher_district = school.district.name

                # If no school district, check event's districts
                if not teacher_district and event.districts:
                    teacher_district = event.districts[0].name

                # If still no district, check event's district_partner
                if not teacher_district and event.district_partner:
                    teacher_district = event.district_partner

                # If teacher belongs to target district, include this event
                # Use helper function to handle aliases (e.g., "KCPS (MO)" vs "Kansas City Public Schools (MO)")
                if _district_name_matches(district_name, teacher_district):
                    event_has_target_district_teacher = True
                    target_district_teachers.add(
                        f"{teacher.first_name} {teacher.last_name}"
                    )
                    if teacher.school_id:
                        school = School.query.get(teacher.school_id)
                        if school:
                            target_district_schools.add(school.name)

        # If no teachers from target district, skip this event
        if not event_has_target_district_teacher:
            continue

        # Calculate participant count for this district only
        # Count only teachers from the target district and multiply by 25
        district_teacher_count = len(target_district_teachers)
        district_participant_count = district_teacher_count * 25

        # Create aggregated session record using only target district teachers and schools
        session_dict[event.id] = {
            "event_id": event.id,  # This should always be the correct event ID
            "status": event.status.value if event.status else "",
            "date": event.start_date.strftime("%m/%d/%y") if event.start_date else "",
            "time": event.start_date.strftime("%I:%M %p") if event.start_date else "",
            "session_type": event.additional_information or "",
            "teachers": (
                sorted(target_district_teachers) if target_district_teachers else []
            ),
            "schools": (
                sorted(target_district_schools) if target_district_schools else []
            ),
            "district": district_name,
            "session_title": event.title,
            "presenter": (
                ", ".join([v.full_name for v in event.volunteers])
                if event.volunteers
                else ""
            ),
            "presenter_data": (
                [
                    {
                        "id": v.id,
                        "name": v.full_name,
                        "is_people_of_color": v.is_people_of_color,
                        "organization_name": v.organization_name,
                        "organizations": (
                            [org.name for org in v.organizations]
                            if v.organizations
                            else []
                        ),
                        # Derive locality from volunteer.local_status
                        "is_local": getattr(v, "local_status", None)
                        == LocalStatusEnum.local,
                    }
                    for v in event.volunteers
                ]
                if event.volunteers
                else []
            ),
            "topic_theme": event.series or "",
            "session_link": event.registration_link or "",
            "participant_count": district_participant_count,
            "duration": event.duration or 0,
            "is_simulcast": (
                any([tr.is_simulcast for tr in event.teacher_registrations])
                if event.teacher_registrations
                else False
            ),
        }

    # Convert to list and sort by date descending
    session_data = list(session_dict.values())
    session_data.sort(key=lambda s: s["date"], reverse=True)

    # Calculate summary statistics
    total_teachers = set()
    total_students = 0
    total_unique_sessions = set()
    total_experiences = 0
    total_organizations = set()
    total_professionals = set()
    total_professionals_of_color = set()
    total_local_professionals = set()
    total_schools = set()
    school_breakdown = {}
    teacher_breakdown = {}

    # We need to track teachers by their actual database objects to get IDs
    # Re-query events with teacher data to get proper teacher IDs
    teacher_sessions = {}  # teacher_id -> session_count
    teacher_details = {}  # teacher_id -> {name, school, id}

    for event in events:
        # Check if any teacher in this event belongs to the target district
        event_has_target_district_teacher = False

        for teacher_reg in event.teacher_registrations:
            teacher = teacher_reg.teacher
            if teacher:
                # Determine district using the same logic as main function
                teacher_district = None

                # Check teacher's school district first
                if teacher.school_id:
                    school = School.query.get(teacher.school_id)
                    if school and school.district:
                        teacher_district = school.district.name

                # If no school district, check event's districts
                if not teacher_district and event.districts:
                    teacher_district = event.districts[0].name

                # If still no district, check event's district_partner
                if not teacher_district and event.district_partner:
                    teacher_district = event.district_partner

                # If teacher belongs to target district, include this event
                # Use helper function to handle aliases (e.g., "KCPS (MO)" vs "Kansas City Public Schools (MO)")
                if _district_name_matches(district_name, teacher_district):
                    event_has_target_district_teacher = True
                    break

        # If no teachers from target district, skip this event
        if not event_has_target_district_teacher:
            continue

        # Only count completed sessions for teacher breakdown and totals
        # Treat "moved to in-person session" as successful/completed
        should_skip = False
        _orig_status = (getattr(event, "original_status_string", "") or "").lower()
        moved_to_in_person = "moved to in-person" in _orig_status
        # Don't skip events with no-show status - we need to count those for teacher breakdown
        if (
            event.status
            and event.status.value not in ["Completed", "Simulcast"]
            and not moved_to_in_person
        ):
            should_skip = True

        if should_skip:
            continue

        # Process teacher registrations to get proper IDs
        for teacher_reg in event.teacher_registrations:
            # Only count teachers who actually attended OR were part of events
            # moved to in-person (counted as successful)
            tr_status_norm = (getattr(teacher_reg, "status", "") or "").lower()
            is_no_show = (
                "no-show" in tr_status_norm
                or "no show" in tr_status_norm
                or "did not attend" in tr_status_norm
            )
            is_cancel = "cancel" in tr_status_norm or "withdraw" in tr_status_norm
            if not moved_to_in_person:
                # For regular virtual events, require confirmed attendance and not a no-show/cancel
                if teacher_reg.attendance_confirmed_at is None:
                    continue
                if is_no_show or is_cancel:
                    continue

            teacher = teacher_reg.teacher
            if teacher:
                # Determine district using the same logic as main function
                teacher_district = None

                # Check teacher's school district first
                if teacher.school_id:
                    school = School.query.get(teacher.school_id)
                    if school and school.district:
                        teacher_district = school.district.name

                # If no school district, check event's districts
                if not teacher_district and event.districts:
                    teacher_district = event.districts[0].name

                # If still no district, check event's district_partner
                if not teacher_district and event.district_partner:
                    teacher_district = event.district_partner

                # Skip teachers not from the target district
                # Use helper function to handle aliases (e.g., "KCPS (MO)" vs "Kansas City Public Schools (MO)")
                if not _district_name_matches(district_name, teacher_district):
                    continue

                teacher_id = teacher.id
                teacher_name = f"{teacher.first_name} {teacher.last_name}"

                # Get school info
                school_name = "N/A"
                if hasattr(teacher, "school_obj") and teacher.school_obj:
                    school_name = teacher.school_obj.name
                elif teacher.school_id:
                    school_obj = School.query.get(teacher.school_id)
                    if school_obj:
                        school_name = school_obj.name

                # Track teacher sessions
                if teacher_id not in teacher_sessions:
                    teacher_sessions[teacher_id] = 0
                    teacher_details[teacher_id] = {
                        "id": teacher_id,
                        "name": teacher_name,
                        "school": school_name,
                    }
                teacher_sessions[teacher_id] += 1
                total_teachers.add(teacher_name)

        # Count local professionals for this included, completed event
        for v in event.volunteers or []:
            try:
                from models.contact import LocalStatusEnum as _LSE

                if getattr(v, "local_status", None) == _LSE.local:
                    total_local_professionals.add(
                        v.id or f"{v.first_name} {v.last_name}"
                    )
            except Exception:
                # Fallback string check if enum import not available at runtime
                try:
                    if str(getattr(v, "local_status", "")).lower().endswith("local"):
                        total_local_professionals.add(
                            v.id or f"{v.first_name} {v.last_name}"
                        )
                except Exception:
                    pass

    # Build teacher breakdown from the collected data
    for teacher_id, session_count in teacher_sessions.items():
        teacher_info = teacher_details[teacher_id]
        teacher_breakdown[teacher_id] = {
            "id": teacher_info["id"],
            "name": teacher_info["name"],
            "school": teacher_info["school"],
            "sessions": session_count,
        }

    # Calculate breakdowns - only for completed sessions
    for session in session_data:
        # Only count completed sessions for breakdowns
        session_status = session.get("status", "").strip()
        if session_status not in ["Completed", "Simulcast"]:
            continue

        # Schools - only count schools from the target district
        for school_name in session["schools"]:
            # Check if this school belongs to the target district
            school = School.query.filter_by(name=school_name).first()
            if (
                school
                and school.district
                and _district_name_matches(district_name, school.district.name)
            ):
                total_schools.add(school_name)
                # School breakdown
                if school_name not in school_breakdown:
                    school_breakdown[school_name] = {"name": school_name, "sessions": 0}
                school_breakdown[school_name]["sessions"] += 1

        # Sessions
        if session["session_title"]:
            total_unique_sessions.add(session["session_title"])

        # Students
        if session.get("participant_count", 0) > 0:
            total_students += session["participant_count"]
        else:
            total_students += 25
        total_experiences += 1

        # Presenters/Organizations
        if session["presenter_data"]:
            for presenter_data in session["presenter_data"]:
                presenter_name = presenter_data.get("name", "")
                if presenter_name:
                    total_professionals.add(presenter_name)

                    # Check if this presenter is marked as People of Color
                    if presenter_data.get("is_people_of_color", False):
                        total_professionals_of_color.add(presenter_name)
        elif session["presenter"]:
            # Fallback to old presenter format
            presenters = [p.strip() for p in session["presenter"].split(",")]
            for presenter in presenters:
                if presenter:
                    total_professionals.add(presenter)
                    # For old format, we can't determine POC status, so don't count them

    # Convert breakdowns to sorted lists
    school_breakdown_list = sorted(
        school_breakdown.values(), key=lambda x: x["sessions"], reverse=True
    )
    teacher_breakdown_list = sorted(
        teacher_breakdown.values(), key=lambda x: x["sessions"], reverse=True
    )

    # Calculate summary statistics - only for completed sessions with confirmed attendance
    # Use the same teacher counting logic as the teacher breakdown
    total_teachers_completed = set()
    total_unique_sessions_completed = set()
    total_experiences_completed = 0
    total_organizations_completed = set()
    total_professionals_completed = set()
    total_professionals_of_color_completed = set()

    # Use the same teacher counting logic as the teacher breakdown
    for teacher_id, teacher_info in teacher_details.items():
        if (
            teacher_sessions[teacher_id] > 0
        ):  # Only count teachers who actually attended
            total_teachers_completed.add(teacher_info["name"])

    # Count sessions and other stats for completed sessions
    completed_sessions = [
        s
        for s in session_data
        if s.get("status", "").strip().lower()
        in ["completed", "simulcast", "successfully completed"]
    ]

    for session in completed_sessions:

        # Count unique sessions for completed sessions
        if session["session_title"]:
            total_unique_sessions_completed.add(session["session_title"])

        # Count presenters/organizations for completed sessions
        if session["presenter_data"]:
            for presenter_data in session["presenter_data"]:
                presenter_name = presenter_data.get("name", "")
                if presenter_name:
                    total_professionals_completed.add(presenter_name)

                    # Check if this presenter is marked as People of Color
                    if presenter_data.get("is_people_of_color", False):
                        total_professionals_of_color_completed.add(presenter_name)

                    # Count organizations - only count the main/current organization
                    if presenter_data.get("organization_name"):
                        org_name = presenter_data["organization_name"]
                        if org_name:
                            total_organizations_completed.add(org_name)
                    elif (
                        presenter_data.get("organizations")
                        and presenter_data["organizations"]
                    ):
                        # Fallback to first organization if no main organization is set
                        org_name = presenter_data["organizations"][0]
                        if org_name:
                            total_organizations_completed.add(org_name)
        elif session["presenter"]:
            # Fallback to old presenter format
            presenters = [p.strip() for p in session["presenter"].split(",")]
            for presenter in presenters:
                if presenter:
                    total_professionals_completed.add(presenter)

    # Count experiences as total teacher sessions (each teacher attending counts as an experience)
    total_experiences_completed = sum(teacher_sessions.values())

    # Calculate monthly statistics - only for completed sessions
    monthly_stats = {}
    for session in session_data:
        # Only count completed sessions for monthly stats
        session_status = session.get("status", "").strip().lower()
        if session_status not in ["completed", "simulcast", "successfully completed"]:
            continue

        # Parse month from date
        try:
            date_obj = datetime.strptime(session["date"], "%m/%d/%y")
            month_key = date_obj.strftime("%Y-%m")
            month_name = date_obj.strftime("%B %Y")
        except Exception:
            continue

        if month_key not in monthly_stats:
            monthly_stats[month_key] = {
                "month_name": month_name,
                "total_sessions": 0,
                "total_registered": 0,
                "total_attended": 0,
                "total_duration": 0,
                "avg_attendance_rate": 0.0,
                "avg_duration": 0.0,
                "schools": set(),
                "educators": set(),
                "career_clusters": set(),
                "events": [],
            }

        stats = monthly_stats[month_key]
        stats["total_sessions"] += 1
        stats["total_registered"] += session.get("participant_count", 0)
        stats["total_attended"] += session.get("participant_count", 0)
        stats["total_duration"] += session.get("duration", 0)

        for school_name in session["schools"]:
            stats["schools"].add(school_name)
        for teacher_name in session["teachers"]:
            stats["educators"].add(teacher_name)
        if session["topic_theme"]:
            stats["career_clusters"].add(session["topic_theme"])

        stats["events"].append(
            {
                "title": session["session_title"],
                "date": session["date"],
                "time": session["time"],
                "duration": session["duration"],
                "registered": session.get("participant_count", 0),
                "attended": session.get("participant_count", 0),
                "schools": (
                    ", ".join(session["schools"]) if session["schools"] else "N/A"
                ),
                "educators": (
                    ", ".join(session["teachers"]) if session["teachers"] else "N/A"
                ),
                "career_cluster": session["topic_theme"],
                "event_id": session["event_id"],
                "session_link": session.get("session_link", ""),
                "presenter": session["presenter"],
                "presenter_data": session.get("presenter_data", []),
            }
        )

    # Finalize monthly stats
    for stats in monthly_stats.values():
        if stats["total_registered"] > 0:
            stats["avg_attendance_rate"] = (
                stats["total_attended"] / stats["total_registered"]
            ) * 100
        if stats["total_sessions"] > 0:
            stats["avg_duration"] = stats["total_duration"] / stats["total_sessions"]
        stats["school_count"] = len(stats["schools"])
        stats["educator_count"] = len(stats["educators"])
        stats["career_cluster_count"] = len(stats["career_clusters"])
        del stats["schools"]
        del stats["educators"]
        del stats["career_clusters"]

    sorted_monthly_stats = dict(sorted(monthly_stats.items()))

    # Prepare summary statistics - use completed sessions only for summary stats
    # Calculate estimated students as unique teachers * 25
    estimated_students = len(total_teachers_completed) * 25

    # Compute local professional count across completed sessions
    local_professionals = set()
    for stats in monthly_stats.values():
        for evt in stats["events"]:
            for p in evt.get("presenter_data", []) or []:
                try:
                    # Prefer id when available
                    pid = p.get("id") or p.get("name")
                    if p.get("is_local"):
                        local_professionals.add(pid)
                except Exception:
                    pass

    # Add district-level session coverage fields based on session_data
    # Build sets of unique session ids and flags
    _session_ids = set()
    _local_sessions = set()
    _poc_sessions = set()
    for s in session_data:
        sid = s.get("event_id") or s.get("session_title")
        if not sid:
            continue
        # Only count completed/simulcast at district detail level
        st = (s.get("status") or "").strip().lower()
        if st not in ("completed", "simulcast", "successfully completed"):
            continue
        _session_ids.add(sid)
        try:
            if any(p.get("is_local") for p in s.get("presenter_data", []) or []):
                _local_sessions.add(sid)
            if any(
                p.get("is_people_of_color") for p in s.get("presenter_data", []) or []
            ):
                _poc_sessions.add(sid)
        except Exception:
            pass

    summary_stats = {
        "total_teachers": len(total_teachers_completed),
        "total_students": estimated_students,
        "total_unique_sessions": len(total_unique_sessions_completed),
        "total_experiences": total_experiences_completed,
        "total_organizations": len(total_organizations_completed),
        "total_professionals": len(total_professionals_completed),
        "total_professionals_of_color": len(total_professionals_of_color_completed),
        "total_local_professionals": len(total_local_professionals),
        "total_schools": len(total_schools),
        # District-level coverage
        "local_session_count": len(_local_sessions),
        "poc_session_count": len(_poc_sessions),
    }

    # Add percents derived from unique session ids
    denom = len(_session_ids) or 1
    summary_stats["local_session_percent"] = (
        round(100 * summary_stats["local_session_count"] / denom)
        if summary_stats["local_session_count"]
        else 0
    )
    summary_stats["poc_session_percent"] = (
        round(100 * summary_stats["poc_session_count"] / denom)
        if summary_stats["poc_session_count"]
        else 0
    )

    return (
        session_data,
        sorted_monthly_stats,
        school_breakdown_list,
        teacher_breakdown_list,
        summary_stats,
    )


# --- End Data Processing Helper Functions ---


def generate_teacher_progress_excel(
    teacher_progress_data, district_name, virtual_year, date_from, date_to
):
    """
    Generate Excel file with teacher progress data including summary and detailed sheets.

    Args:
        teacher_progress_data: Dictionary with school progress data
        district_name: Name of the district
        virtual_year: Virtual year
        date_from: Start date
        date_to: End date

    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Create Summary Sheet
    summary_ws = wb.create_sheet("Summary")

    # Summary headers
    summary_headers = [
        "District",
        "Virtual Year",
        "Date Range",
        "Total Schools",
        "Total Teachers",
        "Goals Achieved",
        "Goals Achieved %",
        "In Progress",
        "In Progress %",
        "Not Started",
        "Not Started %",
    ]

    # Calculate totals
    total_schools = len(teacher_progress_data)
    total_teachers = sum(
        school_data["total_teachers"] for school_data in teacher_progress_data.values()
    )
    total_achieved = sum(
        school_data["goals_achieved"] for school_data in teacher_progress_data.values()
    )
    total_in_progress = sum(
        school_data["goals_in_progress"]
        for school_data in teacher_progress_data.values()
    )
    total_not_started = sum(
        school_data["goals_not_started"]
        for school_data in teacher_progress_data.values()
    )

    # Write summary headers
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write summary data
    summary_data = [
        district_name,
        virtual_year,
        f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}",
        total_schools,
        total_teachers,
        total_achieved,
        (
            f"{(total_achieved / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
        total_in_progress,
        (
            f"{(total_in_progress / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
        total_not_started,
        (
            f"{(total_not_started / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
    ]

    for col, value in enumerate(summary_data, 1):
        cell = summary_ws.cell(row=2, column=col, value=value)
        cell.border = border
        if col in [4, 5, 6, 8, 10]:  # Numeric columns
            cell.alignment = center_alignment

    # Auto-adjust column widths for summary
    for col in range(1, len(summary_headers) + 1):
        summary_ws.column_dimensions[get_column_letter(col)].width = 15

    # Create School Summary Sheet
    school_summary_ws = wb.create_sheet("School Summary")

    # School summary headers
    school_headers = [
        "School Name",
        "Total Teachers",
        "Goals Achieved",
        "Goals Achieved %",
        "In Progress",
        "In Progress %",
        "Not Started",
        "Not Started %",
    ]

    # Write school summary headers
    for col, header in enumerate(school_headers, 1):
        cell = school_summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write school summary data
    row = 2
    for school_name, school_data in teacher_progress_data.items():
        school_data_row = [
            school_name,
            school_data["total_teachers"],
            school_data["goals_achieved"],
            (
                f"{(school_data['goals_achieved'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
            school_data["goals_in_progress"],
            (
                f"{(school_data['goals_in_progress'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
            school_data["goals_not_started"],
            (
                f"{(school_data['goals_not_started'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
        ]

        for col, value in enumerate(school_data_row, 1):
            cell = school_summary_ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [2, 3, 4, 5, 6, 7, 8]:  # Numeric columns
                cell.alignment = center_alignment

        row += 1

    # Auto-adjust column widths for school summary
    for col in range(1, len(school_headers) + 1):
        school_summary_ws.column_dimensions[get_column_letter(col)].width = 18

    # Create Detailed Teacher Sheet
    teacher_detail_ws = wb.create_sheet("Teacher Details")

    # Teacher detail headers
    teacher_headers = [
        "School",
        "Teacher Name",
        "Email",
        "Grade",
        "Target Sessions",
        "Completed Sessions",
        "Planned Sessions",
        "Progress %",
        "Goal Status",
    ]

    # Write teacher detail headers
    for col, header in enumerate(teacher_headers, 1):
        cell = teacher_detail_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write teacher detail data
    row = 2
    for school_name, school_data in teacher_progress_data.items():
        for teacher in school_data["teachers"]:
            teacher_row = [
                school_name,
                teacher["name"],
                teacher["email"],
                teacher["grade"],
                teacher["target_sessions"],
                teacher["completed_sessions"],
                teacher["planned_sessions"],
                f"{teacher['progress_percentage']:.1f}%",
                teacher["goal_status_text"],
            ]

            for col, value in enumerate(teacher_row, 1):
                cell = teacher_detail_ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 6, 7, 8]:  # Numeric columns
                    cell.alignment = center_alignment

            row += 1

    # Auto-adjust column widths for teacher details
    column_widths = [20, 25, 30, 8, 12, 15, 15, 12, 15]
    for col, width in enumerate(column_widths, 1):
        teacher_detail_ws.column_dimensions[get_column_letter(col)].width = width

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


def load_usage_routes():
    def _group_sessions_for_table(session_rows):
        """
        Group sessions by logical session identifier (title + date + time).
        Combines multiple teachers/presenters for the same session into one row.
        This is a display-only transform and should not be used for summary calculations.

        Works for both APP-created and spreadsheet-imported sessions.
        """
        grouped_sessions = {}
        ungrouped = []

        for row in session_rows or []:
            # Create grouping key from title, date, and time
            title = (row.get("session_title") or "").strip().lower()
            date = (row.get("date") or "").strip()
            time = (row.get("time") or "").strip()

            # If we don't have the required fields, pass through ungrouped
            if not title or not date or not time:
                ungrouped.append(row)
                continue

            # Create a unique key for this session
            session_key = (title, date, time)

            bucket = grouped_sessions.get(session_key)
            if not bucket:
                # Create new bucket with this row's data
                bucket = dict(row)
                bucket["_teacher_names"] = set()
                bucket["_school_names"] = set()
                bucket["_presenter_names"] = set()
                bucket["_presenter_orgs"] = set()
                # Grouped rows should not link to a single teacher
                bucket["teacher_id"] = None
                grouped_sessions[session_key] = bucket

            # Collect teachers and schools
            if row.get("teacher_name"):
                bucket["_teacher_names"].add(row["teacher_name"])
            if row.get("school_name"):
                bucket["_school_names"].add(row["school_name"])

            # Collect presenters
            if row.get("presenter"):
                # Presenter field may contain comma-separated names
                presenters = [
                    p.strip() for p in row["presenter"].split(",") if p.strip()
                ]
                bucket["_presenter_names"].update(presenters)

            # Collect presenter organizations
            if row.get("presenter_organization"):
                # presenter_organization is already comma-separated
                orgs = [
                    o.strip()
                    for o in row["presenter_organization"].split(",")
                    if o.strip()
                ]
                bucket["_presenter_orgs"].update(orgs)

            # Also collect from presenter_data if available
            if row.get("presenter_data"):
                for presenter in row["presenter_data"]:
                    if presenter.get("organization_name"):
                        bucket["_presenter_orgs"].add(presenter["organization_name"])
                    elif presenter.get("organizations"):
                        # If organizations is a list
                        if isinstance(presenter["organizations"], list):
                            bucket["_presenter_orgs"].update(presenter["organizations"])

        # Convert grouped sessions to list format
        grouped = []
        for session_key, bucket in grouped_sessions.items():
            teachers = sorted(bucket.pop("_teacher_names", set()))
            schools = sorted(bucket.pop("_school_names", set()))
            presenters = sorted(bucket.pop("_presenter_names", set()))
            presenter_orgs = sorted(bucket.pop("_presenter_orgs", set()))

            bucket["teacher_name"] = ", ".join(teachers)
            bucket["school_name"] = "; ".join(schools)

            # Update presenter field
            if presenters:
                bucket["presenter"] = ", ".join(presenters)
            else:
                bucket["presenter"] = ""

            # Update presenter organization
            if presenter_orgs:
                bucket["presenter_organization"] = ", ".join(presenter_orgs)
            elif not bucket.get("presenter_organization"):
                bucket["presenter_organization"] = ""

            grouped.append(bucket)

        # Keep original ordering: newest first by date/time as strings is imperfect but matches existing behavior.
        return grouped + ungrouped

    @virtual_bp.route("/usage")
    @login_required
    def virtual_usage():
        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = (
            get_current_virtual_year()
        )  # Changed from get_current_school_year()
        selected_virtual_year = request.args.get(
            "year", default_virtual_year
        )  # Changed variable name

        # Check if refresh is requested
        refresh_requested = request.args.get("refresh", "0") == "1"

        # Calculate default date range based on virtual session year
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )  # Changed from get_school_year_dates()

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        # Check if admin wants to see all districts
        show_all_districts = request.args.get("show_all_districts", "0") == "1"

        current_filters = {
            "year": selected_virtual_year,  # Updated variable name
            "date_from": date_from,
            "date_to": date_to,
            "career_cluster": request.args.get("career_cluster"),
            "school": request.args.get("school"),
            "district": request.args.get("district"),
            "status": request.args.get("status"),
            "search": request.args.get(
                "search"
            ),  # Text search across teacher, title, presenter
            "show_all_districts": show_all_districts,
            "group_manual": request.args.get("group_manual", "0") == "1",
        }

        # Check if we're using default full year date range (for caching)
        is_full_year = date_from == default_date_from and date_to == default_date_to

        # If refresh is requested, invalidate cache for this virtual year
        if refresh_requested and is_full_year:
            print(f"DEBUG: Refresh requested for {selected_virtual_year}")
            invalidate_virtual_session_caches(selected_virtual_year)
            print(
                f"Cache invalidated for virtual session report {selected_virtual_year}"
            )

        # Try to get cached data for full year queries (but not if refresh was requested)
        cached_data = None
        is_cached = False
        last_refreshed = None

        if is_full_year and not refresh_requested:
            cached_data = get_virtual_session_cache(
                selected_virtual_year, date_from, date_to
            )
            if cached_data:
                is_cached = True
                last_refreshed = cached_data.last_updated
                print(
                    f"Using cached data for virtual session report {selected_virtual_year}"
                )
                # Use cached data
                session_data = cached_data.session_data
                district_summaries = cached_data.district_summaries
                overall_summary = cached_data.overall_summary
                filter_options = cached_data.filter_options

                # Filter district summaries to main districts if show_all_districts is False
                if not show_all_districts and district_summaries:
                    main_districts = {
                        "Hickman Mills School District",
                        "Grandview School District",
                        "Kansas City Kansas Public Schools",
                    }
                    district_summaries = {
                        k: v
                        for k, v in district_summaries.items()
                        if k in main_districts
                    }

                # Check if cached data has the new fields; if any are missing, recompute summaries
                if session_data and len(session_data) > 0:
                    sample_session = session_data[0]
                    needs_recompute = False
                    if (
                        "teacher_id" not in sample_session
                        or "presenter_data" not in sample_session
                    ):
                        needs_recompute = True
                    if not needs_recompute and "source_host" not in sample_session:
                        needs_recompute = True
                    # Ensure newer Local/POC session coverage fields exist
                    if not needs_recompute and (
                        overall_summary is None
                        or "local_professional_count" not in overall_summary
                        or "local_session_count" not in overall_summary
                        or "poc_session_count" not in overall_summary
                    ):
                        needs_recompute = True
                    # Also ensure district summaries contain the new session coverage fields
                    if (
                        not needs_recompute
                        and isinstance(district_summaries, dict)
                        and district_summaries
                    ):
                        try:
                            any_district = next(iter(district_summaries.values()))
                            if (
                                "local_session_count" not in any_district
                                or "poc_session_count" not in any_district
                            ):
                                needs_recompute = True
                        except Exception:
                            pass
                    if needs_recompute:
                        (
                            session_data,
                            district_summaries,
                            overall_summary,
                            filter_options,
                        ) = compute_virtual_session_data(
                            selected_virtual_year, date_from, date_to, current_filters
                        )
                        is_cached = False

                # Apply runtime filters if any
                if any(
                    [
                        current_filters["career_cluster"],
                        current_filters["school"],
                        current_filters["district"],
                        current_filters["status"],
                        current_filters.get("search"),  # Include search filter
                    ]
                ):
                    # Store original unfiltered data for district summaries
                    unfiltered_session_data = session_data.copy()
                    session_data = apply_runtime_filters(session_data, current_filters)
                    # Recalculate summaries using same method as individual district pages
                    # Get list of all districts that appear in the unfiltered data
                    all_districts_in_data = set()
                    for session in unfiltered_session_data:
                        if session.get("district"):
                            all_districts_in_data.add(session["district"])

                    # Filter to only show main districts by default (unless admin requests all)
                    main_districts = {
                        "Hickman Mills School District",
                        "Grandview School District",
                        "Kansas City Kansas Public Schools",
                    }

                    # Determine which districts to show in the breakdown
                    if show_all_districts:
                        districts_to_show = all_districts_in_data
                    else:
                        # Only show main districts when show_all_districts=False
                        districts_to_show = {
                            d for d in all_districts_in_data if d in main_districts
                        }

                    # Calculate district summaries using the same method as individual district pages
                    district_summaries = {}
                    for district_name in districts_to_show:
                        try:
                            _, _, _, _, summary_stats = (
                                compute_virtual_session_district_data(
                                    district_name,
                                    selected_virtual_year,
                                    date_from,
                                    date_to,
                                )
                            )
                            district_summaries[district_name] = {
                                "teacher_count": summary_stats.get("total_teachers", 0),
                                "total_students": summary_stats.get(
                                    "total_students", 0
                                ),
                                "session_count": summary_stats.get(
                                    "total_unique_sessions", 0
                                ),
                                "total_experiences": summary_stats.get(
                                    "total_experiences", 0
                                ),
                                "organization_count": summary_stats.get(
                                    "total_organizations", 0
                                ),
                                "professional_count": summary_stats.get(
                                    "total_professionals", 0
                                ),
                                "professional_of_color_count": summary_stats.get(
                                    "total_professionals_of_color", 0
                                ),
                                "local_professional_count": summary_stats.get(
                                    "total_local_professionals", 0
                                ),
                                "school_count": summary_stats.get("total_schools", 0),
                                "local_session_count": summary_stats.get(
                                    "local_session_count", 0
                                ),
                                "poc_session_count": summary_stats.get(
                                    "poc_session_count", 0
                                ),
                                "local_session_percent": summary_stats.get(
                                    "local_session_percent", 0
                                ),
                                "poc_session_percent": summary_stats.get(
                                    "poc_session_percent", 0
                                ),
                            }
                        except Exception as e:
                            print(
                                f"DEBUG: Error calculating stats for {district_name}: {str(e)}"
                            )
                            district_summaries[district_name] = {
                                "teacher_count": 0,
                                "total_students": 0,
                                "session_count": 0,
                                "total_experiences": 0,
                                "organization_count": 0,
                                "professional_count": 0,
                                "professional_of_color_count": 0,
                                "local_professional_count": 0,
                                "school_count": 0,
                                "local_session_count": 0,
                                "poc_session_count": 0,
                                "local_session_percent": 0,
                                "poc_session_percent": 0,
                            }

                    # Recalculate overall summary from unfiltered session data
                    _, overall_summary = calculate_summaries_from_sessions(
                        unfiltered_session_data,
                        current_filters.get("show_all_districts", False),
                    )

                    # Filter district summaries based on user scope
                    if (
                        current_user.scope_type == "district"
                        and current_user.allowed_districts
                    ):
                        import json

                        try:
                            allowed_districts = (
                                json.loads(current_user.allowed_districts)
                                if isinstance(current_user.allowed_districts, str)
                                else current_user.allowed_districts
                            )

                            # Create filtered district summaries
                            filtered_district_summaries = {}
                            for district_name, summary in district_summaries.items():
                                if district_name in allowed_districts:
                                    filtered_district_summaries[district_name] = summary

                            # Replace the original district_summaries with the filtered version
                            district_summaries = filtered_district_summaries
                        except (json.JSONDecodeError, TypeError):
                            # If parsing fails, show no districts
                            district_summaries = {}

                # Apply sorting and pagination as before
                # Only group if explicitly requested (group_manual=1)
                # When group_manual=0 or not set, show individual lines (one per teacher registration)
                if current_filters.get("group_manual"):
                    session_data = _group_sessions_for_table(session_data)
                session_data = apply_sorting_and_pagination(
                    session_data, request.args, current_filters
                )

                return render_template(
                    "virtual/virtual_usage.html",
                    session_data=session_data["paginated_data"],
                    filter_options=filter_options,
                    current_filters=current_filters,
                    pagination=session_data["pagination"],
                    google_sheet_url=get_google_sheet_url(selected_virtual_year),
                    district_summaries=district_summaries,
                    overall_summary=overall_summary,
                    last_refreshed=last_refreshed,
                    is_cached=is_cached,
                )

        # If no cache or not full year, compute data fresh
        print(
            f"Computing fresh data for virtual session report {selected_virtual_year}"
        )

        # Get unfiltered data first to ensure we have all districts for summaries
        print("DEBUG: Getting unfiltered data for district summaries...")
        session_data_unfiltered, _, _, filter_options = compute_virtual_session_data(
            selected_virtual_year, date_from, date_to, {}
        )

        # Get list of all districts that appear in the data
        all_districts_in_data = set()
        for session in session_data_unfiltered:
            if session.get("district"):
                all_districts_in_data.add(session["district"])

        print(
            f"DEBUG: Districts found in session_data: {sorted(list(all_districts_in_data))}"
        )

        # Filter to only show main districts by default (unless admin requests all)
        main_districts = {
            "Hickman Mills School District",
            "Grandview School District",
            "Kansas City Kansas Public Schools",
        }

        # Determine which districts to show in the breakdown
        if show_all_districts:
            districts_to_show = all_districts_in_data
        else:
            # Only show main districts when show_all_districts=False
            districts_to_show = {
                d for d in all_districts_in_data if d in main_districts
            }

        print(
            f"DEBUG: Districts to show in breakdown: {sorted(list(districts_to_show))}"
        )

        # Calculate district summaries using the same method as individual district pages
        # This ensures consistency between breakdown cards and individual district pages
        all_district_summaries = {}
        for district_name in districts_to_show:
            try:
                # Use the same calculation method as individual district page
                _, _, _, _, summary_stats = compute_virtual_session_district_data(
                    district_name, selected_virtual_year, date_from, date_to
                )
                # Convert to the format expected by the breakdown template
                all_district_summaries[district_name] = {
                    "teacher_count": summary_stats.get("total_teachers", 0),
                    "total_students": summary_stats.get("total_students", 0),
                    "session_count": summary_stats.get("total_unique_sessions", 0),
                    "total_experiences": summary_stats.get("total_experiences", 0),
                    "organization_count": summary_stats.get("total_organizations", 0),
                    "professional_count": summary_stats.get("total_professionals", 0),
                    "professional_of_color_count": summary_stats.get(
                        "total_professionals_of_color", 0
                    ),
                    "local_professional_count": summary_stats.get(
                        "total_local_professionals", 0
                    ),
                    "school_count": summary_stats.get("total_schools", 0),
                    "local_session_count": summary_stats.get("local_session_count", 0),
                    "poc_session_count": summary_stats.get("poc_session_count", 0),
                    "local_session_percent": summary_stats.get(
                        "local_session_percent", 0
                    ),
                    "poc_session_percent": summary_stats.get("poc_session_percent", 0),
                }
            except Exception as e:
                print(f"DEBUG: Error calculating stats for {district_name}: {str(e)}")
                # Create empty summary if calculation fails
                all_district_summaries[district_name] = {
                    "teacher_count": 0,
                    "total_students": 0,
                    "session_count": 0,
                    "total_experiences": 0,
                    "organization_count": 0,
                    "professional_count": 0,
                    "professional_of_color_count": 0,
                    "local_professional_count": 0,
                    "school_count": 0,
                    "local_session_count": 0,
                    "poc_session_count": 0,
                    "local_session_percent": 0,
                    "poc_session_percent": 0,
                }

        print(
            f"DEBUG: Calculated district_summaries keys: {list(all_district_summaries.keys()) if all_district_summaries else 'None'}"
        )

        # Now get filtered data for the session table
        session_data, _, overall_summary, _ = compute_virtual_session_data(
            selected_virtual_year, date_from, date_to, current_filters
        )

        # Apply runtime filters (including search) if any
        if any(
            [
                current_filters["career_cluster"],
                current_filters["school"],
                current_filters["district"],
                current_filters["status"],
                current_filters.get("search"),  # Include search filter
            ]
        ):
            session_data = apply_runtime_filters(session_data, current_filters)

        # Use the calculated district summaries
        district_summaries = all_district_summaries

        print(
            f"DEBUG: Final district_summaries keys: {list(district_summaries.keys()) if district_summaries else 'None'}"
        )

        # Filter district summaries based on user scope
        if current_user.scope_type == "district" and current_user.allowed_districts:
            import json

            try:
                allowed_districts = (
                    json.loads(current_user.allowed_districts)
                    if isinstance(current_user.allowed_districts, str)
                    else current_user.allowed_districts
                )

                # Create filtered district summaries
                filtered_district_summaries = {}
                for district_name, summary in district_summaries.items():
                    if district_name in allowed_districts:
                        filtered_district_summaries[district_name] = summary

                # Replace the original district_summaries with the filtered version
                district_summaries = filtered_district_summaries
            except (json.JSONDecodeError, TypeError):
                # If parsing fails, show no districts
                district_summaries = {}

        print(
            f"DEBUG: After filtering, district_summaries keys: {list(district_summaries.keys()) if district_summaries else 'None'}"
        )

        # Cache the data if it's a full year query
        if is_full_year:
            # Cache the unfiltered data
            (
                unfiltered_data,
                unfiltered_district_summaries,
                unfiltered_overall_summary,
                _,
            ) = compute_virtual_session_data(
                selected_virtual_year, date_from, date_to, {}  # No filters for cache
            )
            save_virtual_session_cache(
                selected_virtual_year,
                date_from,
                date_to,
                unfiltered_data,
                unfiltered_district_summaries,
                unfiltered_overall_summary,
                filter_options,
            )
            # Set the last refreshed time for this fresh data
            last_refreshed = datetime.now(timezone.utc)

        # Apply sorting and pagination
        # Only group if explicitly requested (group_manual=1)
        # When group_manual=0 or not set, show individual lines (one per teacher registration)
        if current_filters.get("group_manual"):
            session_data = _group_sessions_for_table(session_data)
        session_result = apply_sorting_and_pagination(
            session_data, request.args, current_filters
        )

        return render_template(
            "virtual/virtual_usage.html",
            session_data=session_result["paginated_data"],
            filter_options=filter_options,
            current_filters=current_filters,
            pagination=session_result["pagination"],
            google_sheet_url=get_google_sheet_url(selected_virtual_year),
            district_summaries=district_summaries,
            overall_summary=overall_summary,
            last_refreshed=last_refreshed,
            is_cached=is_cached,
        )

    @virtual_bp.route("/usage/api/search-teachers", methods=["GET"])
    @login_required
    def search_teachers():
        """Search for teachers by name, returning id, name, school, and district."""
        from flask import jsonify

        query = request.args.get("q", "").strip()
        if len(query) < 2:
            return jsonify([])

        # Search by first name, last name, or full name
        search_term = f"%{query}%"
        # NOTE: SQLite doesn't support CONCAT(). Use SQLAlchemy string concat operator instead.
        full_name_expr = Teacher.first_name + " " + Teacher.last_name
        teachers = (
            Teacher.query.filter(
                or_(
                    func.lower(Teacher.first_name).like(func.lower(search_term)),
                    func.lower(Teacher.last_name).like(func.lower(search_term)),
                    func.lower(full_name_expr).like(func.lower(search_term)),
                )
            )
            .options(joinedload(Teacher.school))
            .limit(20)
            .all()
        )

        results = []
        for teacher in teachers:
            school_name = teacher.school.name if teacher.school else None
            district_name = (
                teacher.school.district.name
                if teacher.school and teacher.school.district
                else None
            )
            results.append(
                {
                    "id": teacher.id,
                    "name": f"{teacher.first_name} {teacher.last_name}".strip(),
                    "first_name": teacher.first_name,
                    "last_name": teacher.last_name,
                    "school": school_name,
                    "school_id": teacher.school_id,
                    "district": district_name,
                }
            )

        return jsonify(results)

    @virtual_bp.route("/usage/api/search-presenters", methods=["GET"])
    @login_required
    def search_presenters():
        """Search for volunteers/presenters by name, returning id, name, and organization."""
        from flask import jsonify

        query = request.args.get("q", "").strip()
        if len(query) < 2:
            return jsonify([])

        # Search by first name, last name, or full name
        search_term = f"%{query}%"
        # NOTE: SQLite doesn't support CONCAT(). Use SQLAlchemy string concat operator instead.
        full_name_expr = Volunteer.first_name + " " + Volunteer.last_name
        volunteers = (
            Volunteer.query.filter(
                or_(
                    func.lower(Volunteer.first_name).like(func.lower(search_term)),
                    func.lower(Volunteer.last_name).like(func.lower(search_term)),
                    func.lower(full_name_expr).like(func.lower(search_term)),
                )
            )
            .limit(20)
            .all()
        )

        results = []
        for volunteer in volunteers:
            # Get primary organization through VolunteerOrganization
            vol_org = (
                VolunteerOrganization.query.filter_by(
                    volunteer_id=volunteer.id, is_primary=True
                )
                .options(joinedload(VolunteerOrganization.organization))
                .first()
            )

            org_name = None
            org_id = None
            if vol_org and vol_org.organization:
                org_name = vol_org.organization.name
                org_id = vol_org.organization_id
            else:
                # Fallback to any organization
                vol_org_any = (
                    VolunteerOrganization.query.filter_by(volunteer_id=volunteer.id)
                    .options(joinedload(VolunteerOrganization.organization))
                    .first()
                )
                if vol_org_any and vol_org_any.organization:
                    org_name = vol_org_any.organization.name
                    org_id = vol_org_any.organization_id

            results.append(
                {
                    "id": volunteer.id,
                    "name": f"{volunteer.first_name} {volunteer.last_name}".strip(),
                    "first_name": volunteer.first_name,
                    "last_name": volunteer.last_name,
                    "organization": org_name,
                    "organization_id": org_id,
                }
            )

        return jsonify(results)

    @virtual_bp.route("/usage/api/create-teacher", methods=["POST"])
    @login_required
    def create_teacher_api():
        """Create a new teacher immediately via API."""
        from flask import jsonify

        from routes.virtual.utils import get_or_create_district, get_or_create_school

        def _norm(s):
            return s.strip() if s else ""

        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        f_name = _norm(data.get("first_name", ""))
        l_name = _norm(data.get("last_name", ""))
        school_name = _norm(data.get("school", ""))

        if not f_name or not l_name:
            return jsonify({"error": "First and Last Name are required"}), 400

        try:
            # Check for existing teacher
            teacher = Teacher.query.filter(
                func.lower(Teacher.first_name) == func.lower(f_name),
                func.lower(Teacher.last_name) == func.lower(l_name),
            ).first()

            if not teacher:
                teacher = Teacher(first_name=f_name, last_name=l_name, middle_name="")
                if school_name:
                    district_obj = get_or_create_district(None)
                    school_obj = get_or_create_school(school_name, district_obj)
                    if school_obj:
                        teacher.school_id = school_obj.id
                db.session.add(teacher)
                db.session.commit()

            # Re-query ensure relationships loaded (if needed)
            school_res = teacher.school.name if teacher.school else ""
            district_res = (
                teacher.school.district.name
                if teacher.school and teacher.school.district
                else ""
            )

            return jsonify(
                {
                    "success": True,
                    "id": teacher.id,
                    "name": f"{teacher.first_name} {teacher.last_name}",
                    "school": school_res,
                    "district": district_res,
                }
            )

        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    @virtual_bp.route("/usage/api/create-presenter", methods=["POST"])
    @login_required
    def create_presenter_api():
        """Create a new presenter/volunteer immediately via API."""
        from flask import jsonify

        def _norm(s):
            return s.strip() if s else ""

        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        f_name = _norm(data.get("first_name", ""))
        l_name = _norm(data.get("last_name", ""))
        org_name = _norm(data.get("organization", ""))

        if not f_name:
            return jsonify({"error": "First Name is required"}), 400

        try:
            # Check for existing
            volunteer = Volunteer.query.filter(
                func.lower(Volunteer.first_name) == func.lower(f_name),
                func.lower(Volunteer.last_name) == func.lower(l_name),
            ).first()

            if not volunteer:
                volunteer = Volunteer(first_name=f_name, last_name=l_name)
                db.session.add(volunteer)
                db.session.commit()

            # Handle Org
            if org_name:
                org = Organization.query.filter(
                    func.lower(Organization.name) == func.lower(org_name)
                ).first()
                if not org:
                    org = Organization(name=org_name)
                    db.session.add(org)
                    db.session.commit()

                vol_org = VolunteerOrganization.query.filter_by(
                    volunteer_id=volunteer.id, organization_id=org.id
                ).first()
                if not vol_org:
                    has_primary = VolunteerOrganization.query.filter_by(
                        volunteer_id=volunteer.id, is_primary=True
                    ).first()
                    vol_org = VolunteerOrganization(
                        volunteer_id=volunteer.id,
                        organization_id=org.id,
                        role="Professional",
                        is_primary=not has_primary,
                        status="Current",
                    )
                    db.session.add(vol_org)
                    db.session.commit()

            # Get primary org for display
            display_org = ""
            primary_vol_org = VolunteerOrganization.query.filter_by(
                volunteer_id=volunteer.id, is_primary=True
            ).first()
            if primary_vol_org and primary_vol_org.organization:
                display_org = primary_vol_org.organization.name
            elif not display_org and org_name:
                display_org = org_name

            return jsonify(
                {
                    "success": True,
                    "id": volunteer.id,
                    "name": f"{volunteer.first_name} {volunteer.last_name}",
                    "organization": display_org,
                }
            )

        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    @virtual_bp.route("/usage/update-status/<int:event_id>", methods=["POST"])
    @login_required
    @admin_required
    def update_virtual_session_status(event_id):
        """Quick update of event status only."""
        from flask import jsonify

        event = db.session.get(Event, event_id)
        if not event or event.type != EventType.VIRTUAL_SESSION:
            return jsonify({"error": "Event not found"}), 404

        if event.session_host != "APP":
            return jsonify({"error": "Only app-created sessions can be updated"}), 403

        status_str = (
            request.json.get("status")
            if request.is_json
            else request.form.get("status")
        )
        if not status_str:
            return jsonify({"error": "Status is required"}), 400

        try:
            event.status = EventStatus(status_str)
            db.session.commit()
            return jsonify({"success": True, "status": event.status.value})
        except ValueError:
            return jsonify({"error": "Invalid status"}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    @virtual_bp.route("/usage/edit/<int:event_id>", methods=["GET", "POST"])
    @login_required
    @admin_required
    def edit_virtual_usage_session(event_id):
        """
        Edit an existing Virtual Session event created via the app UI.

        GET: Returns event data as JSON for populating the edit form
        POST: Updates the event with new data
        """
        event = db.session.get(Event, event_id)
        if not event or event.type != EventType.VIRTUAL_SESSION:
            flash("Event not found.", "danger")
            return redirect(url_for("virtual.virtual_usage"))

        # Only allow editing APP-created sessions
        if event.session_host != "APP":
            flash("Only sessions created in the app can be edited here.", "warning")
            return redirect(url_for("virtual.virtual_usage"))

        if request.method == "GET":
            # Return event data as JSON for the edit form
            from flask import jsonify

            # Get teachers
            teachers_data = []
            for reg in event.teacher_registrations:
                teacher = reg.teacher
                if teacher:
                    teachers_data.append(
                        {
                            "id": teacher.id,
                            "name": f"{teacher.first_name} {teacher.last_name}".strip(),
                            "school": teacher.school.name if teacher.school else "",
                        }
                    )

            # Get presenters
            presenters_data = []
            for participation in event.volunteers:
                vol_org = VolunteerOrganization.query.filter_by(
                    volunteer_id=participation.id, is_primary=True
                ).first()
                org_name = (
                    vol_org.organization.name
                    if vol_org and vol_org.organization
                    else None
                )
                presenters_data.append(
                    {
                        "id": participation.id,
                        "name": f"{participation.first_name} {participation.last_name}".strip(),
                        "organization": org_name or "",
                    }
                )

            return jsonify(
                {
                    "id": event.id,
                    "title": event.title,
                    "date": (
                        event.start_date.strftime("%Y-%m-%d")
                        if event.start_date
                        else ""
                    ),
                    "time": (
                        event.start_date.strftime("%H:%M") if event.start_date else ""
                    ),
                    "duration": event.duration or 60,
                    "session_type": event.additional_information or "",
                    "topic_theme": event.series or "",
                    "session_link": event.registration_link or "",
                    "status": event.status.value if event.status else "Draft",
                    "teachers": teachers_data,
                    "presenters": presenters_data,
                }
            )

        # POST: Update event
        def _norm(s: str | None) -> str:
            return (s or "").strip()

        def _split_name(full_name: str) -> tuple[str, str]:
            full_name = _norm(full_name)
            parts = [p for p in full_name.split() if p]
            if not parts:
                return "", ""
            if len(parts) == 1:
                return parts[0], ""
            return " ".join(parts[:-1]), parts[-1]

        # Update basic fields
        title = _norm(request.form.get("title"))
        session_type = _norm(request.form.get("session_type"))
        topic_theme = _norm(request.form.get("topic_theme"))
        registration_link = _norm(request.form.get("session_link"))

        date_str = _norm(request.form.get("date"))
        time_str = _norm(request.form.get("time"))
        duration_minutes = 60
        try:
            duration_minutes = int(_norm(request.form.get("duration")) or "60")
        except ValueError:
            duration_minutes = 60

        if not title or not date_str or not time_str:
            flash("Title, date, and time are required.", "danger")
            return redirect(url_for("virtual.virtual_usage"))

        try:
            start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        except ValueError:
            flash("Invalid date/time.", "danger")
            return redirect(url_for("virtual.virtual_usage"))

        end_dt = start_dt + timedelta(minutes=duration_minutes)

        # Update event
        status_str = _norm(request.form.get("status", "Requested"))
        try:
            event.status = EventStatus(status_str)
        except ValueError:
            event.status = EventStatus.REQUESTED

        event.title = title
        event.start_date = start_dt
        event.end_date = end_dt
        event.duration = duration_minutes
        event.additional_information = session_type or None
        event.series = topic_theme or None
        event.registration_link = registration_link or None

        # Update teachers (remove old, add new)
        EventTeacher.query.filter_by(event_id=event.id).delete()

        from routes.virtual.routes import get_or_create_district, get_or_create_school

        districts_set = set()

        teacher_ids = request.form.getlist("teacher_id[]")
        teacher_names = request.form.getlist("teacher_name[]")
        teacher_schools = request.form.getlist("teacher_school[]")

        for idx, t_id_raw in enumerate(teacher_ids):
            t_id = _norm(t_id_raw)
            t_name = _norm(teacher_names[idx]) if idx < len(teacher_names) else ""

            teacher = None
            if t_id and t_id.isdigit():
                teacher = Teacher.query.get(int(t_id))

            if not teacher and t_name:
                t_first, t_last = _split_name(t_name)
                if t_first and t_last:
                    school_name = (
                        _norm(teacher_schools[idx])
                        if idx < len(teacher_schools)
                        else ""
                    )
                    teacher_q = Teacher.query.filter(
                        func.lower(Teacher.first_name) == func.lower(t_first),
                        func.lower(Teacher.last_name) == func.lower(t_last),
                    )
                    teacher = teacher_q.first()

                    if not teacher:
                        teacher = Teacher(
                            first_name=t_first, last_name=t_last, middle_name=""
                        )
                        if school_name:
                            district_obj = get_or_create_district(None)
                            school_obj = get_or_create_school(school_name, district_obj)
                            if school_obj:
                                teacher.school_id = school_obj.id
                                if school_obj.district:
                                    districts_set.add(school_obj.district.name)
                        db.session.add(teacher)
                        db.session.flush()

            if not teacher:
                continue

            if teacher.school and teacher.school.district:
                districts_set.add(teacher.school.district.name)

            reg = EventTeacher(
                event_id=event.id,
                teacher_id=teacher.id,
                status="registered",
                is_simulcast=False,
                attendance_confirmed_at=None,
            )
            db.session.add(reg)

        # ---- Quick Create Teachers (Edit) ----
        new_t_first = request.form.getlist("new_teacher_first_name[]")
        new_t_last = request.form.getlist("new_teacher_last_name[]")
        new_t_school = request.form.getlist("new_teacher_school[]")

        for idx, f_name in enumerate(new_t_first):
            f_name = _norm(f_name)
            l_name = _norm(new_t_last[idx]) if idx < len(new_t_last) else ""
            if not f_name or not l_name:
                continue

            school_name = _norm(new_t_school[idx]) if idx < len(new_t_school) else ""

            # Check if teacher already exists to avoid dupes
            teacher = Teacher.query.filter(
                func.lower(Teacher.first_name) == func.lower(f_name),
                func.lower(Teacher.last_name) == func.lower(l_name),
            ).first()

            if not teacher:
                teacher = Teacher(first_name=f_name, last_name=l_name, middle_name="")
                if school_name:
                    district_obj = get_or_create_district(None)
                    school_obj = get_or_create_school(school_name, district_obj)
                    if school_obj:
                        teacher.school_id = school_obj.id
                        if school_obj.district:
                            districts_set.add(school_obj.district.name)
                db.session.add(teacher)
                db.session.flush()

            if teacher.school and teacher.school.district:
                districts_set.add(teacher.school.district.name)

            # Link to event if not already linked
            existing_reg = EventTeacher.query.filter_by(
                event_id=event.id, teacher_id=teacher.id
            ).first()
            if not existing_reg:
                reg = EventTeacher(
                    event_id=event.id,
                    teacher_id=teacher.id,
                    status="registered",
                    is_simulcast=False,
                    attendance_confirmed_at=None,
                )
                db.session.add(reg)

        # Update presenters (remove old participations, add new)
        EventParticipation.query.filter_by(event_id=event.id).delete()
        event.volunteers.clear()

        presenter_ids = request.form.getlist("presenter_id[]")
        presenter_names = request.form.getlist("presenter_name[]")
        presenter_orgs = request.form.getlist("presenter_org[]")

        for idx, p_id_raw in enumerate(presenter_ids):
            p_id = _norm(p_id_raw)
            p_name = _norm(presenter_names[idx]) if idx < len(presenter_names) else ""
            org_name = _norm(presenter_orgs[idx]) if idx < len(presenter_orgs) else ""

            volunteer = None
            if p_id and p_id.isdigit():
                volunteer = Volunteer.query.get(int(p_id))

            if not volunteer and p_name:
                p_first, p_last = _split_name(p_name)
                if not p_first:
                    continue
                if not p_last:
                    p_last = ""

                vol_q = Volunteer.query.filter(
                    func.lower(Volunteer.first_name) == func.lower(p_first),
                    func.lower(Volunteer.last_name) == func.lower(p_last),
                )
                volunteer = vol_q.first()

                if not volunteer:
                    volunteer = Volunteer(first_name=p_first, last_name=p_last)
                    db.session.add(volunteer)
                    db.session.flush()

            if not volunteer:
                continue

            if org_name:
                org = Organization.query.filter(
                    func.lower(Organization.name) == func.lower(org_name)
                ).first()
                if not org:
                    org = Organization(name=org_name)
                    db.session.add(org)
                    db.session.flush()

                vol_org = VolunteerOrganization.query.filter_by(
                    volunteer_id=volunteer.id, organization_id=org.id
                ).first()
                if not vol_org:
                    has_primary = VolunteerOrganization.query.filter_by(
                        volunteer_id=volunteer.id, is_primary=True
                    ).first()
                    vol_org = VolunteerOrganization(
                        volunteer_id=volunteer.id,
                        organization_id=org.id,
                        role="Professional",
                        is_primary=not has_primary,
                        status="Current",
                    )
                    db.session.add(vol_org)

            event.add_volunteer(
                volunteer, participant_type="Presenter", status="Confirmed"
            )

        # ---- Quick Create Presenters (Edit) ----
        new_p_first = request.form.getlist("new_presenter_first_name[]")
        new_p_last = request.form.getlist("new_presenter_last_name[]")
        new_p_org = request.form.getlist("new_presenter_org[]")

        for idx, f_name in enumerate(new_p_first):
            f_name = _norm(f_name)
            l_name = _norm(new_p_last[idx]) if idx < len(new_p_last) else ""
            if not f_name:
                continue

            org_name = _norm(new_p_org[idx]) if idx < len(new_p_org) else ""

            # Check if exists
            volunteer = Volunteer.query.filter(
                func.lower(Volunteer.first_name) == func.lower(f_name),
                func.lower(Volunteer.last_name) == func.lower(l_name),
            ).first()

            if not volunteer:
                volunteer = Volunteer(first_name=f_name, last_name=l_name)
                db.session.add(volunteer)
                db.session.flush()

            # Handle Organization
            if org_name:
                org = Organization.query.filter(
                    func.lower(Organization.name) == func.lower(org_name)
                ).first()
                if not org:
                    org = Organization(name=org_name)
                    db.session.add(org)
                    db.session.flush()

                vol_org = VolunteerOrganization.query.filter_by(
                    volunteer_id=volunteer.id, organization_id=org.id
                ).first()
                if not vol_org:
                    has_primary = VolunteerOrganization.query.filter_by(
                        volunteer_id=volunteer.id, is_primary=True
                    ).first()
                    vol_org = VolunteerOrganization(
                        volunteer_id=volunteer.id,
                        organization_id=org.id,
                        role="Professional",
                        is_primary=not has_primary,
                        status="Current",
                    )
                    db.session.add(vol_org)

            # Link to event
            event.add_volunteer(
                volunteer, participant_type="Presenter", status="Confirmed"
            )

        # Update districts
        event.districts.clear()
        for district_name in districts_set:
            district_obj = get_or_create_district(district_name)
            if district_obj:
                try:
                    event.districts.append(district_obj)
                except Exception:
                    pass

        if districts_set:
            event.district_partner = ", ".join(sorted(districts_set))

        db.session.commit()

        # Invalidate caches so the report reflects updated entries
        try:
            from routes.reports.virtual_session import get_current_virtual_year

            year = get_current_virtual_year()
            invalidate_virtual_session_caches(year)
        except Exception:
            pass

        flash("Session updated successfully!", "success")
        return redirect(url_for("virtual.virtual_usage"))

    @virtual_bp.route("/usage/delete/<int:event_id>", methods=["POST"])
    @login_required
    @admin_required
    def delete_virtual_usage_session(event_id):
        """
        Delete a Virtual Session event created via the app UI.

        Only allows deletion of APP-created sessions to prevent accidental
        deletion of imported sessions.
        """
        event = db.session.get(Event, event_id)
        if not event or event.type != EventType.VIRTUAL_SESSION:
            flash("Event not found.", "danger")
            return redirect(url_for("virtual.virtual_usage"))

        # Only allow deleting APP-created sessions
        if event.session_host != "APP":
            flash("Only sessions created in the app can be deleted here.", "warning")
            return redirect(url_for("virtual.virtual_usage"))

        try:
            # Get virtual year for cache invalidation
            from routes.reports.virtual_session import get_current_virtual_year

            year = get_current_virtual_year()

            # Delete EventParticipation records (no cascade relationship)
            EventParticipation.query.filter_by(event_id=event.id).delete()

            # Delete the event (cascades will handle EventTeacher, EventComment, EventAttendanceDetail)
            db.session.delete(event)
            db.session.commit()

            # Invalidate caches so the report reflects the deletion
            try:
                invalidate_virtual_session_caches(year)
            except Exception:
                pass

            flash("Session deleted successfully!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting session: {str(e)}", "danger")

        return redirect(url_for("virtual.virtual_usage"))

    @virtual_bp.route("/usage/create", methods=["POST"])
    @login_required
    @admin_required
    def create_virtual_usage_session():
        """
        Create a new Virtual Session event via the app UI.

        Defaults (per request):
        - Event status: Requested
        - Teacher registrations: registered, attendance not confirmed
        - Presenter participation: Confirmed (as Presenter)
        """

        def _norm(s: str | None) -> str:
            return (s or "").strip()

        def _split_name(full_name: str) -> tuple[str, str]:
            full_name = _norm(full_name)
            parts = [p for p in full_name.split() if p]
            if not parts:
                return "", ""
            if len(parts) == 1:
                return parts[0], ""
            return " ".join(parts[:-1]), parts[-1]

        # ---- Basic event fields ----
        year = _norm(request.form.get("year")) or get_current_virtual_year()
        title = _norm(request.form.get("title"))
        # District will be determined from teachers, not input directly
        session_type = _norm(request.form.get("session_type"))
        topic_theme = _norm(request.form.get("topic_theme"))
        registration_link = _norm(request.form.get("session_link"))

        date_str = _norm(request.form.get("date"))  # YYYY-MM-DD
        time_str = _norm(request.form.get("time"))  # HH:MM
        duration_minutes = 60
        try:
            duration_minutes = int(_norm(request.form.get("duration")) or "60")
        except ValueError:
            duration_minutes = 60
        if duration_minutes <= 0:
            duration_minutes = 60

        if not title or not date_str or not time_str:
            flash("Title, date, and time are required.", "danger")
            return redirect(url_for("virtual.virtual_usage", year=year))

        try:
            start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        except ValueError:
            flash("Invalid date/time.", "danger")
            return redirect(url_for("virtual.virtual_usage", year=year))

        end_dt = start_dt + timedelta(minutes=duration_minutes)

        # ---- District/School helpers (reuse import logic) ----
        from routes.virtual.routes import get_or_create_district, get_or_create_school

        # Collect districts from teachers
        districts_set = set()

        try:
            event = Event(
                title=title,
                type=EventType.VIRTUAL_SESSION,
                format=EventFormat.VIRTUAL,
                start_date=start_dt,
                end_date=end_dt,
                duration=duration_minutes,
                status=EventStatus.REQUESTED,
                additional_information=session_type or None,
                series=topic_theme or None,
                registration_link=registration_link or None,
                session_host="APP",  # Used to distinguish manual/app-entered sessions
            )
            db.session.add(event)
            db.session.flush()  # get event.id

            # ---- Teachers (multi) ----
            teacher_ids = request.form.getlist("teacher_id[]")
            teacher_names = request.form.getlist("teacher_name[]")
            teacher_schools = request.form.getlist("teacher_school[]")

            for idx, t_id_raw in enumerate(teacher_ids):
                t_id = _norm(t_id_raw)
                t_name = _norm(teacher_names[idx]) if idx < len(teacher_names) else ""

                teacher = None

                # If ID provided, use existing teacher
                if t_id and t_id.isdigit():
                    teacher = Teacher.query.get(int(t_id))

                # If no teacher found by ID, try to find/create by name
                if not teacher and t_name:
                    t_first, t_last = _split_name(t_name)
                    if t_first and t_last:
                        school_name = (
                            _norm(teacher_schools[idx])
                            if idx < len(teacher_schools)
                            else ""
                        )

                        # Find existing teacher by name (case-insensitive)
                        teacher_q = Teacher.query.filter(
                            func.lower(Teacher.first_name) == func.lower(t_first),
                            func.lower(Teacher.last_name) == func.lower(t_last),
                        )
                        teacher = teacher_q.first()

                        # Create new teacher if not found
                        if not teacher:
                            teacher = Teacher(
                                first_name=t_first, last_name=t_last, middle_name=""
                            )
                            if school_name:
                                district_obj = get_or_create_district(
                                    None
                                )  # Will be set from school
                                school_obj = get_or_create_school(
                                    school_name, district_obj
                                )
                                if school_obj:
                                    teacher.school_id = school_obj.id
                                    if school_obj.district:
                                        districts_set.add(school_obj.district.name)
                            db.session.add(teacher)
                            db.session.flush()

                if not teacher:
                    continue

                # Get district from teacher's school
                if teacher.school and teacher.school.district:
                    districts_set.add(teacher.school.district.name)

                # Create registration if not exists
                reg = EventTeacher.query.filter_by(
                    event_id=event.id, teacher_id=teacher.id
                ).first()
                if not reg:
                    reg = EventTeacher(
                        event_id=event.id,
                        teacher_id=teacher.id,
                        status="registered",
                        is_simulcast=False,
                        attendance_confirmed_at=None,
                    )
                    db.session.add(reg)

            # ---- Quick Create Teachers (Create) ----
            new_t_first = request.form.getlist("new_teacher_first_name[]")
            new_t_last = request.form.getlist("new_teacher_last_name[]")
            new_t_school = request.form.getlist("new_teacher_school[]")

            for idx, f_name in enumerate(new_t_first):
                f_name = _norm(f_name)
                l_name = _norm(new_t_last[idx]) if idx < len(new_t_last) else ""
                if not f_name or not l_name:
                    continue

                school_name = (
                    _norm(new_t_school[idx]) if idx < len(new_t_school) else ""
                )

                # Check if teacher already exists to avoid dupes
                teacher = Teacher.query.filter(
                    func.lower(Teacher.first_name) == func.lower(f_name),
                    func.lower(Teacher.last_name) == func.lower(l_name),
                ).first()

                if not teacher:
                    teacher = Teacher(
                        first_name=f_name, last_name=l_name, middle_name=""
                    )
                    if school_name:
                        district_obj = get_or_create_district(
                            None
                        )  # Will be set from school
                        school_obj = get_or_create_school(school_name, district_obj)
                        if school_obj:
                            teacher.school_id = school_obj.id
                            if school_obj.district:
                                districts_set.add(school_obj.district.name)
                    db.session.add(teacher)
                    db.session.flush()

                if teacher.school and teacher.school.district:
                    districts_set.add(teacher.school.district.name)

                # Create registration if not exists
                reg = EventTeacher.query.filter_by(
                    event_id=event.id, teacher_id=teacher.id
                ).first()
                if not reg:
                    reg = EventTeacher(
                        event_id=event.id,
                        teacher_id=teacher.id,
                        status="registered",
                        is_simulcast=False,
                        attendance_confirmed_at=None,
                    )
                    db.session.add(reg)

            # Attach districts to event
            for district_name in districts_set:
                district_obj = get_or_create_district(district_name)
                if district_obj:
                    try:
                        event.districts.append(district_obj)
                    except Exception:
                        pass

            if districts_set:
                event.district_partner = ", ".join(sorted(districts_set))

            # ---- Presenters (multi) ----
            presenter_ids = request.form.getlist("presenter_id[]")
            presenter_names = request.form.getlist("presenter_name[]")
            presenter_orgs = request.form.getlist("presenter_org[]")

            for idx, p_id_raw in enumerate(presenter_ids):
                p_id = _norm(p_id_raw)
                p_name = (
                    _norm(presenter_names[idx]) if idx < len(presenter_names) else ""
                )
                org_name = (
                    _norm(presenter_orgs[idx]) if idx < len(presenter_orgs) else ""
                )

                volunteer = None

                # If ID provided, use existing volunteer
                if p_id and p_id.isdigit():
                    volunteer = Volunteer.query.get(int(p_id))

                # If no volunteer found by ID, try to find/create by name
                if not volunteer and p_name:
                    p_first, p_last = _split_name(p_name)
                    if not p_first:
                        continue
                    if not p_last:
                        p_last = ""

                    # Find existing volunteer by name (case-insensitive)
                    vol_q = Volunteer.query.filter(
                        func.lower(Volunteer.first_name) == func.lower(p_first),
                        func.lower(Volunteer.last_name) == func.lower(p_last),
                    )
                    volunteer = vol_q.first()

                    # Create new volunteer if not found
                    if not volunteer:
                        volunteer = Volunteer(first_name=p_first, last_name=p_last)
                        db.session.add(volunteer)
                        db.session.flush()

                if not volunteer:
                    continue

                # Ensure org exists + primary association
                if org_name:
                    org = Organization.query.filter(
                        func.lower(Organization.name) == func.lower(org_name)
                    ).first()
                    if not org:
                        org = Organization(name=org_name)
                        db.session.add(org)
                        db.session.flush()

                    vol_org = VolunteerOrganization.query.filter_by(
                        volunteer_id=volunteer.id, organization_id=org.id
                    ).first()
                    if not vol_org:
                        # Check if volunteer has any primary org, if not make this one primary
                        has_primary = VolunteerOrganization.query.filter_by(
                            volunteer_id=volunteer.id, is_primary=True
                        ).first()
                        vol_org = VolunteerOrganization(
                            volunteer_id=volunteer.id,
                            organization_id=org.id,
                            role="Professional",
                            is_primary=not has_primary,  # Primary if no other primary exists
                            status="Current",
                        )
                        db.session.add(vol_org)

                # Link as presenter participation
                event.add_volunteer(
                    volunteer, participant_type="Presenter", status="Confirmed"
                )

            # ---- Quick Create Presenters (Create) ----
            new_p_first = request.form.getlist("new_presenter_first_name[]")
            new_p_last = request.form.getlist("new_presenter_last_name[]")
            new_p_org = request.form.getlist("new_presenter_org[]")

            for idx, f_name in enumerate(new_p_first):
                f_name = _norm(f_name)
                l_name = _norm(new_p_last[idx]) if idx < len(new_p_last) else ""
                if not f_name:
                    continue

                org_name = _norm(new_p_org[idx]) if idx < len(new_p_org) else ""

                # Check if exists
                volunteer = Volunteer.query.filter(
                    func.lower(Volunteer.first_name) == func.lower(f_name),
                    func.lower(Volunteer.last_name) == func.lower(l_name),
                ).first()

                if not volunteer:
                    volunteer = Volunteer(first_name=f_name, last_name=l_name)
                    db.session.add(volunteer)
                    db.session.flush()

                # Handle Organization
                if org_name:
                    org = Organization.query.filter(
                        func.lower(Organization.name) == func.lower(org_name)
                    ).first()
                    if not org:
                        org = Organization(name=org_name)
                        db.session.add(org)
                        db.session.flush()

                    vol_org = VolunteerOrganization.query.filter_by(
                        volunteer_id=volunteer.id, organization_id=org.id
                    ).first()
                    if not vol_org:
                        has_primary = VolunteerOrganization.query.filter_by(
                            volunteer_id=volunteer.id, is_primary=True
                        ).first()
                        vol_org = VolunteerOrganization(
                            volunteer_id=volunteer.id,
                            organization_id=org.id,
                            role="Professional",
                            is_primary=not has_primary,
                            status="Current",
                        )
                        db.session.add(vol_org)

                event.add_volunteer(
                    volunteer, participant_type="Presenter", status="Confirmed"
                )

            db.session.commit()

            # Invalidate caches so the report reflects new entries
            try:
                invalidate_virtual_session_caches(year)
            except Exception:
                pass

            flash("Virtual session created.", "success")
            return redirect(url_for("virtual.virtual_usage", year=year))

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating virtual session: {e}", "danger")
            return redirect(url_for("virtual.virtual_usage", year=year))

    @virtual_bp.route("/usage/district/<district_name>")
    @login_required
    @district_scoped_required
    def virtual_usage_district(district_name):
        # Get filter parameters: Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Check if refresh is requested
        refresh_requested = request.args.get("refresh", "0") == "1"

        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        date_from = default_date_from
        date_to = default_date_to
        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass
        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass
        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to
        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }
        virtual_year_options = generate_school_year_options()

        # Check if we're using default full year date range (for caching)
        is_full_year = date_from == default_date_from and date_to == default_date_to

        # If refresh is requested, invalidate cache for this district and virtual year
        if refresh_requested and is_full_year:
            # Invalidate district cache
            cached_data = VirtualSessionDistrictCache.query.filter_by(
                district_name=district_name,
                virtual_year=selected_virtual_year,
                date_from=date_from.date(),
                date_to=date_to.date(),
            ).first()
            if cached_data:
                db.session.delete(cached_data)
                db.session.commit()
            print(
                f"Cache invalidated for district {district_name} virtual session report {selected_virtual_year}"
            )

        # Try to get cached data for full year queries
        cached_data = None
        is_cached = False
        last_refreshed = None

        if is_full_year and not refresh_requested:
            cached_data = get_virtual_session_district_cache(
                district_name, selected_virtual_year, date_from, date_to
            )
            if cached_data:
                is_cached = True
                last_refreshed = cached_data.last_updated
                print(
                    f"Using cached data for district {district_name} virtual session report {selected_virtual_year}"
                )

                # Check if cached data has the new teacher ID fields and correct student calculation
                teacher_breakdown = cached_data.teacher_breakdown
                summary_stats = cached_data.summary_stats

                needs_refresh = False

                # Check for missing teacher ID field
                if teacher_breakdown and len(teacher_breakdown) > 0:
                    sample_teacher = teacher_breakdown[0]
                    if "id" not in sample_teacher:
                        print(
                            f"DEBUG: Cached data missing teacher ID field, invalidating cache"
                        )
                        needs_refresh = True

                # Check if student calculation is using old method (not unique teachers * 25)
                if (
                    summary_stats
                    and "total_teachers" in summary_stats
                    and "total_students" in summary_stats
                ):
                    expected_students = summary_stats["total_teachers"] * 25
                    actual_students = summary_stats["total_students"]
                    if actual_students != expected_students:
                        print(
                            f"DEBUG: Cached data using old student calculation ({actual_students} vs expected {expected_students}), invalidating cache"
                        )
                        needs_refresh = True

                # Force refresh to ensure completed session filtering is applied
                print(
                    f"DEBUG: Forcing cache refresh to ensure completed session filtering is applied"
                )
                needs_refresh = True

                if needs_refresh:
                    db.session.delete(cached_data)
                    db.session.commit()
                    # Force fresh data computation
                    (
                        session_data,
                        monthly_stats,
                        school_breakdown,
                        teacher_breakdown,
                        summary_stats,
                    ) = compute_virtual_session_district_data(
                        district_name, selected_virtual_year, date_from, date_to
                    )
                    is_cached = False
                    last_refreshed = datetime.now(timezone.utc)

                    # Cache the new data
                    save_virtual_session_district_cache(
                        district_name,
                        selected_virtual_year,
                        date_from,
                        date_to,
                        session_data,
                        monthly_stats,
                        school_breakdown,
                        teacher_breakdown,
                        summary_stats,
                    )
                else:
                    # Use cached data
                    session_data = cached_data.session_data
                    monthly_stats = cached_data.monthly_stats
                    school_breakdown = cached_data.school_breakdown
                    summary_stats = cached_data.summary_stats

                return render_template(
                    "virtual/virtual_usage_district.html",
                    district_name=district_name,
                    monthly_stats=monthly_stats,
                    current_filters=current_filters,
                    school_year_options=virtual_year_options,
                    total_teachers=summary_stats["total_teachers"],
                    total_students=summary_stats["total_students"],
                    total_unique_sessions=summary_stats["total_unique_sessions"],
                    total_experiences=summary_stats["total_experiences"],
                    total_organizations=summary_stats["total_organizations"],
                    total_professionals=summary_stats["total_professionals"],
                    total_local_professionals=summary_stats.get(
                        "total_local_professionals", 0
                    ),
                    total_professionals_of_color=summary_stats[
                        "total_professionals_of_color"
                    ],
                    total_schools=summary_stats["total_schools"],
                    summary_stats=summary_stats,
                    school_breakdown=school_breakdown,
                    teacher_breakdown=teacher_breakdown,
                    session_data=session_data,
                    last_refreshed=last_refreshed,
                    is_cached=is_cached,
                )

        # If no cache or not full year, compute data fresh
        print(
            f"Computing fresh data for district {district_name} virtual session report {selected_virtual_year}"
        )
        (
            session_data,
            monthly_stats,
            school_breakdown,
            teacher_breakdown,
            summary_stats,
        ) = compute_virtual_session_district_data(
            district_name, selected_virtual_year, date_from, date_to
        )

        # Cache the data if it's a full year query
        if is_full_year:
            save_virtual_session_district_cache(
                district_name,
                selected_virtual_year,
                date_from,
                date_to,
                session_data,
                monthly_stats,
                school_breakdown,
                teacher_breakdown,
                summary_stats,
            )
            # Set the last refreshed time for this fresh data
            last_refreshed = datetime.now(timezone.utc)

        return render_template(
            "virtual/virtual_usage_district.html",
            district_name=district_name,
            monthly_stats=monthly_stats,
            current_filters=current_filters,
            school_year_options=virtual_year_options,
            total_teachers=summary_stats["total_teachers"],
            total_students=summary_stats["total_students"],
            total_unique_sessions=summary_stats["total_unique_sessions"],
            total_experiences=summary_stats["total_experiences"],
            total_organizations=summary_stats["total_organizations"],
            total_professionals=summary_stats["total_professionals"],
            total_local_professionals=summary_stats.get("total_local_professionals", 0),
            total_professionals_of_color=summary_stats["total_professionals_of_color"],
            total_schools=summary_stats["total_schools"],
            summary_stats=summary_stats,
            school_breakdown=school_breakdown,
            teacher_breakdown=teacher_breakdown,
            session_data=session_data,
            last_refreshed=last_refreshed,
            is_cached=is_cached,
        )

    @virtual_bp.route("/usage/export")
    @login_required
    def virtual_usage_export():
        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range based on virtual session year
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
            "career_cluster": request.args.get("career_cluster"),
            "school": request.args.get("school"),
            "district": request.args.get("district"),
            "status": request.args.get("status"),
            "search": request.args.get(
                "search"
            ),  # Text search across teacher, title, presenter
        }

        # Base query for virtual session events (same as main route)
        base_query = Event.query.options(
            joinedload(Event.districts),
            joinedload(Event.teacher_registrations)
            .joinedload(EventTeacher.teacher)
            .joinedload(Teacher.school),
        ).filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )

        # Apply filters (same as main route)
        if current_filters["career_cluster"]:
            base_query = base_query.filter(
                Event.series.ilike(f'%{current_filters["career_cluster"]}%')
            )

        if current_filters["status"]:
            base_query = base_query.filter(Event.status == current_filters["status"])

        # Get all events
        events = base_query.order_by(Event.start_date.desc()).all()

        # Build session data (same logic as main route)
        session_data = []

        for event in events:
            teacher_registrations = event.teacher_registrations

            if teacher_registrations:
                for teacher_reg in teacher_registrations:
                    teacher = teacher_reg.teacher

                    school = None
                    school_name = ""
                    school_level = ""

                    if teacher:
                        if hasattr(teacher, "school_obj") and teacher.school_obj:
                            school = teacher.school_obj
                            school_name = school.name
                            school_level = getattr(school, "level", "")
                        elif teacher.school_id:
                            school = School.query.get(teacher.school_id)
                            if school:
                                school_name = school.name
                                school_level = getattr(school, "level", "")

                    district_name = None
                    if event.districts:
                        district_name = event.districts[0].name
                    elif event.district_partner:
                        district_name = event.district_partner
                    else:
                        district_name = "Unknown District"

                    # Apply filters
                    if (
                        current_filters["district"]
                        and district_name != current_filters["district"]
                    ):
                        continue

                    if (
                        current_filters["school"]
                        and school_name
                        and current_filters["school"].lower() not in school_name.lower()
                    ):
                        continue

                    session_data.append(
                        {
                            "status": teacher_reg.status or "registered",
                            "date": (
                                event.start_date.strftime("%m/%d/%y")
                                if event.start_date
                                else ""
                            ),
                            "time": (
                                event.start_date.strftime("%I:%M %p")
                                if event.start_date
                                else ""
                            ),
                            "session_type": event.additional_information or "",
                            "teacher_name": (
                                f"{teacher.first_name} {teacher.last_name}"
                                if teacher
                                else ""
                            ),
                            "teacher_id": teacher.id if teacher else None,
                            "school_name": school_name,
                            "school_level": school_level,
                            "district": district_name,
                            "session_title": event.title,
                            "presenter": (
                                ", ".join([v.full_name for v in event.volunteers])
                                if event.volunteers
                                else ""
                            ),
                            "presenter_organization": (
                                ", ".join(
                                    [
                                        _get_primary_org_name_for_volunteer(v)
                                        or "Independent"
                                        for v in event.volunteers
                                    ]
                                )
                                if event.volunteers
                                else ""
                            ),
                            "presenter_data": (
                                [
                                    {"id": v.id, "name": v.full_name}
                                    for v in event.volunteers
                                ]
                                if event.volunteers
                                else []
                            ),
                            "topic_theme": event.series or "",
                            "session_link": event.registration_link or "",
                            "participant_count": event.participant_count or 0,
                            "duration": event.duration or 0,
                        }
                    )
            else:
                district_name = None
                if event.districts:
                    district_name = event.districts[0].name
                elif event.district_partner:
                    district_name = event.district_partner
                else:
                    district_name = "Unknown District"

                if (
                    current_filters["district"]
                    and district_name != current_filters["district"]
                ):
                    continue

                session_data.append(
                    {
                        "status": event.status.value if event.status else "",
                        "date": (
                            event.start_date.strftime("%m/%d/%y")
                            if event.start_date
                            else ""
                        ),
                        "time": (
                            event.start_date.strftime("%I:%M %p")
                            if event.start_date
                            else ""
                        ),
                        "session_type": event.additional_information or "",
                        "teacher_name": "",
                        "teacher_id": None,
                        "school_name": "",
                        "school_level": "",
                        "district": district_name,
                        "session_title": event.title,
                        "presenter": (
                            ", ".join([v.full_name for v in event.volunteers])
                            if event.volunteers
                            else ""
                        ),
                        "presenter_organization": (
                            ", ".join(
                                [
                                    _get_primary_org_name_for_volunteer(v)
                                    or "Independent"
                                    for v in event.volunteers
                                ]
                            )
                            if event.volunteers
                            else ""
                        ),
                        "presenter_data": (
                            [
                                {"id": v.id, "name": v.full_name}
                                for v in event.volunteers
                            ]
                            if event.volunteers
                            else []
                        ),
                        "topic_theme": event.series or "",
                        "session_link": event.registration_link or "",
                        "participant_count": event.participant_count or 0,
                        "duration": event.duration or 0,
                    }
                )

        # Apply runtime filters (including search)
        session_data = apply_runtime_filters(session_data, current_filters)

        # Apply sorting (same logic as main route)
        sort_column = request.args.get("sort", "date")
        sort_order = request.args.get("order", "desc")

        sortable_columns = {
            "status": "status",
            "date": "date",
            "time": "time",
            "session_type": "session_type",
            "teacher_name": "teacher_name",
            "school_name": "school_name",
            "school_level": "school_level",
            "district": "district",
            "session_title": "session_title",
            "presenter": "presenter",
            "presenter_organization": "presenter_organization",
            "topic_theme": "topic_theme",
        }

        if sort_column in sortable_columns:
            reverse_order = sort_order == "desc"

            if sort_column == "date":

                def date_sort_key(session):
                    try:
                        if session["date"]:
                            date_parts = session["date"].split("/")
                            if len(date_parts) == 3:
                                month, day, year = (
                                    int(date_parts[0]),
                                    int(date_parts[1]),
                                    int(date_parts[2]),
                                )
                                if year < 50:
                                    year += 2000
                                else:
                                    year += 1900
                                return datetime(year, month, day)
                            elif len(date_parts) == 2:
                                month, day = int(date_parts[0]), int(date_parts[1])
                                virtual_year_start = int(
                                    selected_virtual_year.split("-")[0]
                                )
                                year = (
                                    virtual_year_start
                                    if month >= 7
                                    else virtual_year_start + 1
                                )
                                return datetime(year, month, day)
                        return datetime.min
                    except (ValueError, IndexError):
                        return datetime.min

                session_data.sort(key=date_sort_key, reverse=reverse_order)

            elif sort_column == "time":

                def time_sort_key(session):
                    try:
                        if session["time"]:
                            return datetime.strptime(session["time"], "%I:%M %p").time()
                        return datetime.min.time()
                    except ValueError:
                        return datetime.min.time()

                session_data.sort(key=time_sort_key, reverse=reverse_order)

            else:

                def string_sort_key(session):
                    value = session.get(sortable_columns[sort_column], "") or ""
                    return str(value).lower()

                session_data.sort(key=string_sort_key, reverse=reverse_order)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Virtual Session Usage"

        # Define headers
        headers = [
            "Status",
            "Date",
            "Time",
            "Session Type",
            "Teacher Name",
            "School Name",
            "School Level",
            "District",
            "Session Title",
            "Presenter",
            "Topic/Theme",
            "Session Link",
            "Participant Count",
            "Duration (min)",
        ]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="495057", end_color="495057", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Add data
        for row_idx, session in enumerate(session_data, 2):
            data_row = [
                session["status"],
                session["date"],
                session["time"],
                session["session_type"],
                session["teacher_name"],
                session["school_name"],
                session["school_level"],
                session["district"],
                session["session_title"],
                session["presenter"],
                session.get("presenter_organization", ""),
                session["topic_theme"],
                session["session_link"],
                session["participant_count"],
                session["duration"],
            ]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

        # Add summary info at the top
        ws.insert_rows(1, 4)

        # Report title
        title_cell = ws.cell(row=1, column=1, value="Virtual Session Usage Report")
        title_cell.font = Font(bold=True, size=16)

        # Export date
        export_date_cell = ws.cell(
            row=2,
            column=1,
            value=f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )
        export_date_cell.font = Font(italic=True)

        # Filters applied
        filter_info = f"Virtual Year: {selected_virtual_year}"
        if current_filters["date_from"] and current_filters["date_to"]:
            filter_info += f" | Date Range: {current_filters['date_from'].strftime('%Y-%m-%d')} to {current_filters['date_to'].strftime('%Y-%m-%d')}"
        if current_filters["district"]:
            filter_info += f" | District: {current_filters['district']}"
        if current_filters["career_cluster"]:
            filter_info += f" | Career Cluster: {current_filters['career_cluster']}"
        if current_filters["status"]:
            filter_info += f" | Status: {current_filters['status']}"

        filter_cell = ws.cell(row=3, column=1, value=f"Filters: {filter_info}")
        filter_cell.font = Font(italic=True, size=10)

        # Total records
        total_cell = ws.cell(
            row=4, column=1, value=f"Total Records: {len(session_data)}"
        )
        total_cell.font = Font(bold=True)

        # Create file in memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"virtual_session_usage_{selected_virtual_year}_{timestamp}.xlsx"

        # Create response
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @virtual_bp.route("/usage/breakdown")
    @login_required
    def virtual_breakdown():
        """
        Detailed breakdown of virtual sessions by month and category.

        Categories:
        - Successfully completed sessions
        - Number of simulcast + Sessions
        - Number of teacher canceled sessions
        - Number of teacher no shows
        - Number of pathful professional canceled/no shows sessions
        - Number of local professional canceled/no show sessions
        - Number of unfilled sessions
        """
        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range based on virtual session year
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Base query for virtual session events
        base_query = Event.query.options(
            joinedload(Event.districts),
            joinedload(Event.teacher_registrations).joinedload(EventTeacher.teacher),
            joinedload(Event.volunteers),
            joinedload(
                Event.school_obj
            ),  # Add school relationship for grade level analysis
        ).filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )

        # Apply district filtering for district-scoped users
        if current_user.scope_type == "district" and current_user.allowed_districts:
            import json

            try:
                allowed_districts = (
                    json.loads(current_user.allowed_districts)
                    if isinstance(current_user.allowed_districts, str)
                    else current_user.allowed_districts
                )
                # Filter events by allowed districts
                base_query = base_query.filter(
                    Event.districts.any(District.name.in_(allowed_districts))
                )
            except (json.JSONDecodeError, TypeError):
                # If parsing fails, return no events
                base_query = base_query.filter(False)

        events = base_query.all()

        # Initialize monthly breakdown data
        monthly_breakdown = {}

        # Create month entries for the full virtual year (August to July)
        virtual_year_start = int(selected_virtual_year.split("-")[0])
        months = [
            (virtual_year_start, 8, "August"),
            (virtual_year_start, 9, "September"),
            (virtual_year_start, 10, "October"),
            (virtual_year_start, 11, "November"),
            (virtual_year_start, 12, "December"),
            (virtual_year_start + 1, 1, "January"),
            (virtual_year_start + 1, 2, "February"),
            (virtual_year_start + 1, 3, "March"),
            (virtual_year_start + 1, 4, "April"),
            (virtual_year_start + 1, 5, "May"),
            (virtual_year_start + 1, 6, "June"),
            (virtual_year_start + 1, 7, "July"),
        ]

        for year, month_num, month_name in months:
            month_key = f"{year}-{month_num:02d}"
            monthly_breakdown[month_key] = {
                "month_name": month_name,
                "year": year,
                "month": month_num,
                "successfully_completed": 0,
                "simulcast_sessions": 0,
                "teacher_canceled": 0,
                "teacher_no_shows": 0,
                "pathful_professional_canceled_no_shows": 0,
                "local_professional_canceled_no_shows": 0,
                "unfilled_sessions": 0,
                # New fields for volunteer diversity metrics - 4-column structure
                # Sessions Only (completed sessions)
                "local_volunteer_count_sessions_only": 0,
                "local_volunteer_percentage_sessions_only": 0.0,
                "poc_volunteer_count_sessions_only": 0,
                "poc_volunteer_percentage_sessions_only": 0.0,
                # Sessions + Simulcast (completed + simulcast sessions)
                "local_volunteer_count_sessions_simulcast": 0,
                "local_volunteer_percentage_sessions_simulcast": 0.0,
                "poc_volunteer_count_sessions_simulcast": 0,
                "poc_volunteer_percentage_sessions_simulcast": 0.0,
                # Track sessions filled by local/POC volunteers
                "_sessions_with_local_volunteers": set(),
                "_sessions_with_poc_volunteers": set(),
                "_simulcast_sessions_with_local_volunteers": set(),
                "_simulcast_sessions_with_poc_volunteers": set(),
            }

        # Debug: Collect unique status strings to understand what we're working with
        unique_statuses = set()
        unique_additional_info = set()
        unique_series = set()
        unique_school_levels = set()

        # Process each event
        for event in events:
            if not event.start_date:
                continue

            month_key = f"{event.start_date.year}-{event.start_date.month:02d}"

            # Skip if not in our breakdown range
            if month_key not in monthly_breakdown:
                continue

            # Get the original status string for analysis
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()

            # Collect unique statuses for debugging
            if original_status:
                unique_statuses.add(original_status)

            # Collect unique additional info (session types)
            if event.additional_information:
                unique_additional_info.add(event.additional_information)

            # Collect unique series (topics)
            if event.series:
                unique_series.add(event.series)

            # Collect unique school levels
            if event.school_obj and event.school_obj.level:
                unique_school_levels.add(event.school_obj.level)

            # Categorize based on original status string (primary source of truth)
            if "successfully completed" in original_status:
                monthly_breakdown[month_key]["successfully_completed"] += 1

            elif "simulcast" in original_status:
                monthly_breakdown[month_key]["simulcast_sessions"] += 1

            elif "teacher" in original_status and (
                "cancelation" in original_status or "cancellation" in original_status
            ):
                monthly_breakdown[month_key]["teacher_canceled"] += 1

            elif (
                "teacher no-show" in original_status
                or "teacher no show" in original_status
            ):
                monthly_breakdown[month_key]["teacher_no_shows"] += 1

            elif "pathful professional" in original_status:
                monthly_breakdown[month_key][
                    "pathful_professional_canceled_no_shows"
                ] += 1

            elif "local professional" in original_status:
                monthly_breakdown[month_key][
                    "local_professional_canceled_no_shows"
                ] += 1

            elif (
                "technical difficulties" in original_status
                or original_status == "count"
            ):
                monthly_breakdown[month_key]["unfilled_sessions"] += 1

            # Fallback to event status if no original status string
            elif not original_status:
                event_status = event.status
                event_status_str = event_status.value.lower() if event_status else ""

                if event_status == EventStatus.COMPLETED:
                    monthly_breakdown[month_key]["successfully_completed"] += 1
                elif event_status == EventStatus.SIMULCAST:
                    monthly_breakdown[month_key]["simulcast_sessions"] += 1
                elif event_status == EventStatus.NO_SHOW and not event.volunteers:
                    monthly_breakdown[month_key]["unfilled_sessions"] += 1
                elif event_status == EventStatus.CANCELLED and not event.volunteers:
                    monthly_breakdown[month_key]["unfilled_sessions"] += 1

                # Check for professional statuses in main event status
                elif "pathful professional" in event_status_str:
                    monthly_breakdown[month_key][
                        "pathful_professional_canceled_no_shows"
                    ] += 1
                elif "local professional" in event_status_str:
                    monthly_breakdown[month_key][
                        "local_professional_canceled_no_shows"
                    ] += 1

            # Check teacher registrations for additional status information
            for teacher_reg in event.teacher_registrations:
                tr_status = (teacher_reg.status or "").lower().strip()

                # Check for professional statuses FIRST (before teacher-specific logic)
                if "pathful professional" in tr_status:
                    monthly_breakdown[month_key][
                        "pathful_professional_canceled_no_shows"
                    ] += 1
                elif "local professional" in tr_status:
                    monthly_breakdown[month_key][
                        "local_professional_canceled_no_shows"
                    ] += 1

                # Count teacher-specific cancellations and no-shows (only if not professional)
                elif "cancel" in tr_status and "teacher" not in original_status:
                    monthly_breakdown[month_key]["teacher_canceled"] += 1
                elif (
                    "no-show" in tr_status or "no show" in tr_status
                ) and "teacher" not in original_status:
                    monthly_breakdown[month_key]["teacher_no_shows"] += 1

            # Check for simulcast from teacher registrations (additional simulcast detection)
            simulcast_teacher_count = sum(
                1 for tr in event.teacher_registrations if tr.is_simulcast
            )
            if simulcast_teacher_count > 0 and "simulcast" not in original_status:
                monthly_breakdown[month_key]["simulcast_sessions"] += 1

            # Calculate volunteer diversity metrics for successfully completed sessions
            # (which includes simulcast sessions - same volunteer talking to multiple classes)
            if "successfully completed" in original_status or (
                not original_status and event.status == EventStatus.COMPLETED
            ):

                # Track sessions filled by local and POC volunteers
                event_id = event.id or event.title
                if event_id:
                    has_local_volunteer = False
                    has_poc_volunteer = False

                    for volunteer in event.volunteers or []:
                        # Check if volunteer is local
                        try:
                            from models.contact import LocalStatusEnum

                            if (
                                getattr(volunteer, "local_status", None)
                                == LocalStatusEnum.local
                            ):
                                has_local_volunteer = True
                        except Exception:
                            try:
                                if (
                                    str(getattr(volunteer, "local_status", ""))
                                    .lower()
                                    .endswith("local")
                                ):
                                    has_local_volunteer = True
                            except Exception:
                                pass

                        # Check if volunteer is People of Color
                        if getattr(volunteer, "is_people_of_color", False):
                            has_poc_volunteer = True

                    # Track for completed sessions (sessions only)
                    if has_local_volunteer:
                        monthly_breakdown[month_key][
                            "_sessions_with_local_volunteers"
                        ].add(event_id)
                    if has_poc_volunteer:
                        monthly_breakdown[month_key][
                            "_sessions_with_poc_volunteers"
                        ].add(event_id)

                    # Also track for sessions + simulcast (if this is a simulcast session)
                    if "simulcast" in original_status or (
                        not original_status and event.status == EventStatus.SIMULCAST
                    ):
                        if has_local_volunteer:
                            monthly_breakdown[month_key][
                                "_simulcast_sessions_with_local_volunteers"
                            ].add(event_id)
                        if has_poc_volunteer:
                            monthly_breakdown[month_key][
                                "_simulcast_sessions_with_poc_volunteers"
                            ].add(event_id)

        # Calculate final monthly volunteer counts and percentages from unique sets
        for month_key, month_data in monthly_breakdown.items():
            # Calculate session-based metrics for both scenarios

            # 1. Sessions Only (completed sessions)
            month_data["local_volunteer_count_sessions_only"] = len(
                month_data["_sessions_with_local_volunteers"]
            )
            month_data["poc_volunteer_count_sessions_only"] = len(
                month_data["_sessions_with_poc_volunteers"]
            )

            # 2. Sessions + Simulcast (completed + simulcast sessions)
            month_data["local_volunteer_count_sessions_simulcast"] = len(
                month_data["_sessions_with_local_volunteers"]
            ) + len(month_data["_simulcast_sessions_with_local_volunteers"])
            month_data["poc_volunteer_count_sessions_simulcast"] = len(
                month_data["_sessions_with_poc_volunteers"]
            ) + len(month_data["_simulcast_sessions_with_poc_volunteers"])

            # Calculate total sessions for percentage calculations
            total_completed_sessions = month_data["successfully_completed"]
            total_sessions_simulcast = (
                month_data["successfully_completed"] + month_data["simulcast_sessions"]
            )

            # Calculate percentages for Sessions Only
            if total_completed_sessions > 0:
                month_data["local_volunteer_percentage_sessions_only"] = round(
                    (
                        month_data["local_volunteer_count_sessions_only"]
                        / total_completed_sessions
                    )
                    * 100,
                    1,
                )
                month_data["poc_volunteer_percentage_sessions_only"] = round(
                    (
                        month_data["poc_volunteer_count_sessions_only"]
                        / total_completed_sessions
                    )
                    * 100,
                    1,
                )
            else:
                month_data["local_volunteer_percentage_sessions_only"] = 0.0
                month_data["poc_volunteer_percentage_sessions_only"] = 0.0

            # Calculate percentages for Sessions + Simulcast
            if total_sessions_simulcast > 0:
                month_data["local_volunteer_percentage_sessions_simulcast"] = round(
                    (
                        month_data["local_volunteer_count_sessions_simulcast"]
                        / total_sessions_simulcast
                    )
                    * 100,
                    1,
                )
                month_data["poc_volunteer_percentage_sessions_simulcast"] = round(
                    (
                        month_data["poc_volunteer_count_sessions_simulcast"]
                        / total_sessions_simulcast
                    )
                    * 100,
                    1,
                )
            else:
                month_data["local_volunteer_percentage_sessions_simulcast"] = 0.0
                month_data["poc_volunteer_percentage_sessions_simulcast"] = 0.0

            # Clean up tracking sets
            del month_data["_sessions_with_local_volunteers"]
            del month_data["_sessions_with_poc_volunteers"]
            del month_data["_simulcast_sessions_with_local_volunteers"]
            del month_data["_simulcast_sessions_with_poc_volunteers"]

        # Calculate year-to-date totals
        ytd_totals = {
            "successfully_completed": 0,
            "simulcast_sessions": 0,
            "teacher_canceled": 0,
            "teacher_no_shows": 0,
            "pathful_professional_canceled_no_shows": 0,
            "local_professional_canceled_no_shows": 0,
            "unfilled_sessions": 0,
            # New fields for volunteer diversity metrics - 4-column structure
            # Sessions Only (completed sessions)
            "local_volunteer_count_sessions_only": 0,
            "local_volunteer_percentage_sessions_only": 0.0,
            "poc_volunteer_count_sessions_only": 0,
            "poc_volunteer_percentage_sessions_only": 0.0,
            # Sessions + Simulcast (completed + simulcast sessions)
            "local_volunteer_count_sessions_simulcast": 0,
            "local_volunteer_percentage_sessions_simulcast": 0.0,
            "poc_volunteer_count_sessions_simulcast": 0,
            "poc_volunteer_percentage_sessions_simulcast": 0.0,
        }

        for month_data in monthly_breakdown.values():
            for key in ytd_totals:
                if key in month_data:
                    ytd_totals[key] += month_data[key]

        # Calculate YTD percentages for volunteer diversity metrics
        # Track sessions filled by local and POC volunteers across all months
        all_ytd_sessions_with_local = set()
        all_ytd_sessions_with_poc = set()
        all_ytd_simulcast_with_local = set()
        all_ytd_simulcast_with_poc = set()

        # Re-process events to collect YTD session data
        for event in events:
            if not event.start_date:
                continue

            # Only include events in our breakdown range
            month_key = f"{event.start_date.year}-{event.start_date.month:02d}"
            if month_key not in monthly_breakdown:
                continue

            # Only count successfully completed sessions
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()

            if "successfully completed" in original_status or (
                not original_status and event.status == EventStatus.COMPLETED
            ):
                event_id = event.id or event.title
                if event_id:
                    has_local_volunteer = False
                    has_poc_volunteer = False

                    for volunteer in event.volunteers or []:
                        # Check if volunteer is local
                        try:
                            from models.contact import LocalStatusEnum

                            if (
                                getattr(volunteer, "local_status", None)
                                == LocalStatusEnum.local
                            ):
                                has_local_volunteer = True
                        except Exception:
                            try:
                                if (
                                    str(getattr(volunteer, "local_status", ""))
                                    .lower()
                                    .endswith("local")
                                ):
                                    has_local_volunteer = True
                            except Exception:
                                pass

                        # Check if volunteer is People of Color
                        if getattr(volunteer, "is_people_of_color", False):
                            has_poc_volunteer = True

                    # Track for completed sessions (sessions only)
                    if has_local_volunteer:
                        all_ytd_sessions_with_local.add(event_id)
                    if has_poc_volunteer:
                        all_ytd_sessions_with_poc.add(event_id)

                    # Also track for sessions + simulcast (if this is a simulcast session)
                    if "simulcast" in original_status or (
                        not original_status and event.status == EventStatus.SIMULCAST
                    ):
                        if has_local_volunteer:
                            all_ytd_simulcast_with_local.add(event_id)
                        if has_poc_volunteer:
                            all_ytd_simulcast_with_poc.add(event_id)

        # Calculate YTD totals from session counts
        ytd_totals["local_volunteer_count_sessions_only"] = len(
            all_ytd_sessions_with_local
        )
        ytd_totals["poc_volunteer_count_sessions_only"] = len(all_ytd_sessions_with_poc)

        ytd_totals["local_volunteer_count_sessions_simulcast"] = len(
            all_ytd_sessions_with_local
        ) + len(all_ytd_simulcast_with_local)
        ytd_totals["poc_volunteer_count_sessions_simulcast"] = len(
            all_ytd_sessions_with_poc
        ) + len(all_ytd_simulcast_with_poc)

        # Calculate total YTD sessions for percentage calculation
        total_ytd_completed = ytd_totals["successfully_completed"]
        total_ytd_sessions_simulcast = (
            ytd_totals["successfully_completed"] + ytd_totals["simulcast_sessions"]
        )

        # Calculate YTD percentages for Sessions Only
        if total_ytd_completed > 0:
            ytd_totals["local_volunteer_percentage_sessions_only"] = round(
                (
                    ytd_totals["local_volunteer_count_sessions_only"]
                    / total_ytd_completed
                )
                * 100,
                1,
            )
            ytd_totals["poc_volunteer_percentage_sessions_only"] = round(
                (ytd_totals["poc_volunteer_count_sessions_only"] / total_ytd_completed)
                * 100,
                1,
            )
        else:
            ytd_totals["local_volunteer_percentage_sessions_only"] = 0.0
            ytd_totals["poc_volunteer_percentage_sessions_only"] = 0.0

        # Calculate YTD percentages for Sessions + Simulcast
        if total_ytd_sessions_simulcast > 0:
            ytd_totals["local_volunteer_percentage_sessions_simulcast"] = round(
                (
                    ytd_totals["local_volunteer_count_sessions_simulcast"]
                    / total_ytd_sessions_simulcast
                )
                * 100,
                1,
            )
            ytd_totals["poc_volunteer_percentage_sessions_simulcast"] = round(
                (
                    ytd_totals["poc_volunteer_count_sessions_simulcast"]
                    / total_ytd_sessions_simulcast
                )
                * 100,
                1,
            )
        else:
            ytd_totals["local_volunteer_percentage_sessions_simulcast"] = 0.0
            ytd_totals["poc_volunteer_percentage_sessions_simulcast"] = 0.0

        # COMPREHENSIVE ANALYTICS SECTIONS

        # 1. Current Running Count - detailed status breakdown
        running_count = {
            "unfilled": 0,
            "successfully_completed": 0,
            "teacher_cancelation": 0,
            "formerly_in_person_completed": 0,
            "formerly_in_person_canceled": 0,
            "professional_cancel_no_show": 0,
            "inclement_weather_cancellation": 0,
            "withdrawn_time_constraint": 0,
            "moved_to_in_person_session": 0,
            "teacher_no_show": 0,
            "covid19_cancelation": 0,
            "white_label_completed": 0,
            "white_label_unfilled": 0,
            "technical_difficulties": 0,
            "local_professional_cancellation": 0,
            "pathful_professional_cancellation": 0,
            "local_professional_no_show": 0,
            "pathful_professional_no_show": 0,
            "formerly_in_person_duplicate": 0,
            "white_label_canceled": 0,
            "simulcast": 0,
            "count": 0,
        }

        # 2. District-wise completed sessions
        district_completed = {}

        # 3. Topic-wise session counts
        topic_counts = {}

        # 4. Session type breakdown
        session_types = {
            "teacher_requested": {"all": 0, "completed": 0},
            "industry_chat": {"all": 0, "completed": 0},
            "formerly_in_person": {"all": 0, "completed": 0},
            "other": {"all": 0, "completed": 0},
            "kansas_city_series": {"all": 0, "completed": 0},
        }

        # 5. Grade level breakdown
        grade_levels = {
            "elementary": {"all": 0, "completed_no_simulcast": 0},
            "middle": {"all": 0, "completed_no_simulcast": 0},
            "high": {"all": 0, "completed_no_simulcast": 0},
            "grade_level_not_set": {"all": 0, "completed_no_simulcast": 0},
            "pre_k": {"all": 0, "completed_no_simulcast": 0},
        }

        # 6. Teacher attendance metrics
        teacher_attendance = {
            "more_than_one_teacher_present": 0,
            "total_number_of_sessions": len(events),
        }

        # Process events for comprehensive analytics
        for event in events:
            # Get status information
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()
            event_status = event.status

            # Running count analysis - check multiple sources like monthly breakdown
            event_categorized = False

            # Check original status first
            if "successfully completed" in original_status:
                running_count["successfully_completed"] += 1
                event_categorized = True
            elif "simulcast" in original_status:
                running_count["simulcast"] += 1
                event_categorized = True
            elif "teacher" in original_status and (
                "cancelation" in original_status or "cancellation" in original_status
            ):
                running_count["teacher_cancelation"] += 1
                event_categorized = True
            elif (
                "teacher no-show" in original_status
                or "teacher no show" in original_status
            ):
                running_count["teacher_no_show"] += 1
                event_categorized = True
            elif "pathful professional cancellation" in original_status:
                running_count["pathful_professional_cancellation"] += 1
                event_categorized = True
            elif (
                "pathful professional no-show" in original_status
                or "pathful professional no show" in original_status
            ):
                running_count["pathful_professional_no_show"] += 1
                event_categorized = True
            elif "local professional cancellation" in original_status:
                running_count["local_professional_cancellation"] += 1
                event_categorized = True
            elif (
                "local professional no-show" in original_status
                or "local professional no show" in original_status
            ):
                running_count["local_professional_no_show"] += 1
                event_categorized = True
            elif "technical difficulties" in original_status:
                running_count["technical_difficulties"] += 1
                event_categorized = True
            elif "inclement weather" in original_status:
                running_count["inclement_weather_cancellation"] += 1
                event_categorized = True
            elif "moved to in-person" in original_status:
                running_count["moved_to_in_person_session"] += 1
                event_categorized = True
            elif (
                "white label completed" in original_status
                or "white lable completed" in original_status
            ):
                running_count["white_label_completed"] += 1
                event_categorized = True
            elif (
                "white label unfilled" in original_status
                or "white lable unfilled" in original_status
            ):
                running_count["white_label_unfilled"] += 1
                event_categorized = True
            elif (
                "formerly in-person" in original_status
                and "completed" in original_status
            ):
                running_count["formerly_in_person_completed"] += 1
                event_categorized = True
            elif "formerly in-person" in original_status and (
                "canceled" in original_status or "cancelled" in original_status
            ):
                running_count["formerly_in_person_canceled"] += 1
                event_categorized = True
            elif "covid" in original_status:
                running_count["covid19_cancelation"] += 1
                event_categorized = True
            elif original_status == "count":
                running_count["count"] += 1
                event_categorized = True

            # If not categorized by original status, check teacher registrations
            if not event_categorized:
                for teacher_reg in event.teacher_registrations:
                    tr_status = (teacher_reg.status or "").lower().strip()

                    # Check for professional statuses FIRST
                    if "pathful professional" in tr_status:
                        if "no-show" in tr_status or "no show" in tr_status:
                            running_count["pathful_professional_no_show"] += 1
                        else:
                            running_count["pathful_professional_cancellation"] += 1
                        event_categorized = True
                        break
                    elif "local professional" in tr_status:
                        if "no-show" in tr_status or "no show" in tr_status:
                            running_count["local_professional_no_show"] += 1
                        else:
                            running_count["local_professional_cancellation"] += 1
                        event_categorized = True
                        break
                    # Check for teacher-specific statuses
                    elif "cancel" in tr_status and "professional" not in tr_status:
                        running_count["teacher_cancelation"] += 1
                        event_categorized = True
                        break
                    elif (
                        "no-show" in tr_status or "no show" in tr_status
                    ) and "professional" not in tr_status:
                        running_count["teacher_no_show"] += 1
                        event_categorized = True
                        break

                # Check for simulcast from teacher registrations
                if not event_categorized:
                    simulcast_teacher_count = sum(
                        1 for tr in event.teacher_registrations if tr.is_simulcast
                    )
                    if simulcast_teacher_count > 0:
                        running_count["simulcast"] += 1
                        event_categorized = True

            # If still not categorized, check event status
            if not event_categorized:
                if event_status == EventStatus.COMPLETED:
                    running_count["successfully_completed"] += 1
                    event_categorized = True
                elif event_status == EventStatus.SIMULCAST:
                    running_count["simulcast"] += 1
                    event_categorized = True
                elif event_status == EventStatus.CANCELLED and not event.volunteers:
                    running_count["unfilled"] += 1
                    event_categorized = True
                elif event_status == EventStatus.NO_SHOW and not event.volunteers:
                    running_count["unfilled"] += 1
                    event_categorized = True

            # Default to unfilled if still not categorized
            if not event_categorized:
                running_count["unfilled"] += 1

            # District-wise completed sessions - check if this event is successfully completed
            event_is_completed = False

            # Check multiple sources for "successfully completed" status
            if "successfully completed" in original_status:
                event_is_completed = True
            elif event_status == EventStatus.COMPLETED:
                event_is_completed = True
            else:
                # Check teacher registrations for completion indicators
                for teacher_reg in event.teacher_registrations:
                    tr_status = (teacher_reg.status or "").lower().strip()
                    if (
                        "successfully completed" in tr_status
                        or "attended" in tr_status
                        or "completed" in tr_status
                    ):
                        event_is_completed = True
                        break

            # Count by district if successfully completed
            if event_is_completed and event.districts:
                for district in event.districts:
                    district_name = district.name
                    if district_name not in district_completed:
                        district_completed[district_name] = 0
                    district_completed[district_name] += 1

            # Topic-wise session counts (using series field which contains Topic/Theme)
            if event.series and event.series.strip():
                topic = event.series.strip()
                if topic not in topic_counts:
                    topic_counts[topic] = 0
                topic_counts[topic] += 1

            # Session type analysis (using additional_information field)
            session_type = event.additional_information or "other"
            session_type_clean = (
                session_type.lower().replace(" ", "_").replace("-", "_")
            )

            # Map to our predefined session types or use 'other'
            if session_type_clean in session_types:
                session_types[session_type_clean]["all"] += 1
                if "successfully completed" in original_status:
                    session_types[session_type_clean]["completed"] += 1
            else:
                session_types["other"]["all"] += 1
                if "successfully completed" in original_status:
                    session_types["other"]["completed"] += 1

            # Grade level analysis (using school level)
            grade_level = None
            if event.school_obj and event.school_obj.level:
                grade_level = event.school_obj.level.lower().strip()

            if grade_level:
                # Map school levels to our grade level categories
                if grade_level in ["elementary", "elem"]:
                    grade_key = "elementary"
                elif grade_level in ["middle", "mid"]:
                    grade_key = "middle"
                elif grade_level in ["high", "secondary"]:
                    grade_key = "high"
                elif grade_level in ["pre-k", "prek", "pre_k"]:
                    grade_key = "pre_k"
                else:
                    grade_key = "grade_level_not_set"

                grade_levels[grade_key]["all"] += 1
                # Count completed (no simulcast)
                if "successfully completed" in original_status:
                    grade_levels[grade_key]["completed_no_simulcast"] += 1
            else:
                grade_levels["grade_level_not_set"]["all"] += 1
                if "successfully completed" in original_status:
                    grade_levels["grade_level_not_set"]["completed_no_simulcast"] += 1

            # Teacher attendance analysis
            teacher_count = len(event.teacher_registrations)
            if teacher_count > 1:
                teacher_attendance["more_than_one_teacher_present"] += 1

        # Calculate professional cancel/no-show total for running count
        running_count["professional_cancel_no_show"] = (
            running_count["pathful_professional_cancellation"]
            + running_count["pathful_professional_no_show"]
            + running_count["local_professional_cancellation"]
            + running_count["local_professional_no_show"]
        )

        # Sort districts and topics for consistent display
        district_completed = dict(sorted(district_completed.items()))
        topic_counts = dict(sorted(topic_counts.items()))

        # Debug output to understand the data
        print(f"DEBUG: Found {len(events)} events")
        print(f"DEBUG: Unique statuses: {sorted(unique_statuses)}")
        print(f"DEBUG: Unique additional info: {sorted(unique_additional_info)}")
        print(f"DEBUG: Unique series: {sorted(unique_series)}")
        print(f"DEBUG: Unique school levels: {sorted(unique_school_levels)}")
        print(f"DEBUG: Running count totals: {running_count}")
        print(f"DEBUG: Topic counts: {topic_counts}")
        print(f"DEBUG: District completed: {district_completed}")

        # Show overall completion stats
        total_completed = sum(district_completed.values()) if district_completed else 0
        total_successfully_completed = running_count.get("successfully_completed", 0)
        print(f"DEBUG: Total completed by district: {total_completed}")
        print(
            f"DEBUG: Total successfully completed in running count: {total_successfully_completed}"
        )

        # Prepare data for template
        virtual_year_options = generate_school_year_options()

        return render_template(
            "virtual/virtual_breakdown.html",
            monthly_breakdown=monthly_breakdown,
            ytd_totals=ytd_totals,
            running_count=running_count,
            district_completed=district_completed,
            topic_counts=topic_counts,
            session_types=session_types,
            grade_levels=grade_levels,
            teacher_attendance=teacher_attendance,
            current_filters=current_filters,
            virtual_year_options=virtual_year_options,
            months=months,
        )

    @virtual_bp.route("/usage/google-sheets")
    @login_required
    def virtual_google_sheets():
        """Manage Google Sheets for virtual district reports"""
        virtual_year = request.args.get("year", get_current_virtual_year())

        # Get all Google Sheets for virtual district reports for this year
        sheets = (
            GoogleSheet.query.filter_by(
                academic_year=virtual_year, purpose="virtual_district_reports"
            )
            .order_by(GoogleSheet.sheet_name)
            .all()
        )

        # Get only allowed districts for dropdown
        allowed_district_names = {
            "Hickman Mills School District",
            "Grandview School District",
            "Kansas City Kansas Public Schools",
        }
        districts = (
            District.query.filter(District.name.in_(allowed_district_names))
            .order_by(District.name)
            .all()
        )

        return render_template(
            "virtual/virtual_google_sheets.html",
            sheets=sheets,
            districts=districts,
            virtual_year=virtual_year,
            virtual_year_options=generate_school_year_options(),
        )

    @virtual_bp.route("/usage/google-sheets/create", methods=["POST"])
    @login_required
    def create_virtual_google_sheet():
        """Create a new Google Sheet for virtual district reports"""
        try:
            virtual_year = request.form.get("virtual_year")
            district_name = request.form.get("district_name")
            sheet_id = request.form.get("sheet_id")
            sheet_name = request.form.get("sheet_name")

            if not all([virtual_year, district_name, sheet_id, sheet_name]):
                flash("All fields are required.", "error")
                return redirect(
                    url_for("virtual.virtual_google_sheets", year=virtual_year)
                )

            # Check if sheet already exists for this district and year
            existing_sheet = GoogleSheet.query.filter_by(
                academic_year=virtual_year,
                purpose="virtual_district_reports",
                sheet_name=sheet_name,
            ).first()

            if existing_sheet:
                flash(
                    f"A Google Sheet with this name already exists for {virtual_year}.",
                    "error",
                )
                return redirect(
                    url_for("virtual.virtual_google_sheets", year=virtual_year)
                )

            # Create new Google Sheet record
            new_sheet = GoogleSheet(
                academic_year=virtual_year,
                purpose="virtual_district_reports",
                sheet_id=sheet_id,
                sheet_name=sheet_name,
                created_by=current_user.id,
            )

            db.session.add(new_sheet)
            db.session.commit()

            flash(
                f'Google Sheet "{sheet_name}" created successfully for {district_name}.',
                "success",
            )
            return redirect(url_for("virtual.virtual_google_sheets", year=virtual_year))

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating Google Sheet: {str(e)}", "error")
            return redirect(url_for("virtual.virtual_google_sheets", year=virtual_year))

    @virtual_bp.route("/usage/google-sheets/<int:sheet_id>/update", methods=["POST"])
    @login_required
    def update_virtual_google_sheet(sheet_id):
        """Update an existing Google Sheet"""
        try:
            sheet = GoogleSheet.query.get_or_404(sheet_id)

            sheet_url = request.form.get("sheet_id")
            sheet_name = request.form.get("sheet_name")

            if not all([sheet_url, sheet_name]):
                flash("Sheet URL and name are required.", "error")
                return redirect(
                    url_for("virtual.virtual_google_sheets", year=sheet.academic_year)
                )

            sheet.update_sheet_id(sheet_url)
            sheet.sheet_name = sheet_name

            db.session.commit()

            flash(f'Google Sheet "{sheet_name}" updated successfully.', "success")
            return redirect(
                url_for("virtual.virtual_google_sheets", year=sheet.academic_year)
            )

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating Google Sheet: {str(e)}", "error")
            return redirect(url_for("virtual.virtual_google_sheets"))

    @virtual_bp.route("/usage/google-sheets/<int:sheet_id>/delete", methods=["POST"])
    @login_required
    def delete_virtual_google_sheet(sheet_id):
        """Delete a Google Sheet"""
        try:
            sheet = GoogleSheet.query.get_or_404(sheet_id)
            virtual_year = sheet.academic_year
            sheet_name = sheet.sheet_name

            db.session.delete(sheet)
            db.session.commit()

            flash(f'Google Sheet "{sheet_name}" deleted successfully.', "success")
            return redirect(url_for("virtual.virtual_google_sheets", year=virtual_year))

        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting Google Sheet: {str(e)}", "error")
            return redirect(url_for("virtual.virtual_google_sheets"))

    @virtual_bp.route("/usage/google-sheets/<int:sheet_id>/view")
    @login_required
    def view_virtual_google_sheet(sheet_id):
        """Redirect to the Google Sheet"""
        sheet = GoogleSheet.query.get_or_404(sheet_id)

        if not sheet.decrypted_sheet_id:
            flash("Google Sheet URL not available.", "error")
            return redirect(
                url_for("virtual.virtual_google_sheets", year=sheet.academic_year)
            )

        # Create Google Sheets URL
        google_sheets_url = (
            f"https://docs.google.com/spreadsheets/d/{sheet.decrypted_sheet_id}/edit"
        )

        return redirect(google_sheets_url)

    @virtual_bp.route("/usage/district/<district_name>/google-sheet")
    @login_required
    @district_scoped_required
    def get_district_google_sheet(district_name):
        """Get Google Sheet for a specific district and year"""
        virtual_year = request.args.get("year", get_current_virtual_year())

        # Look for a sheet that matches this district in the name
        sheet = GoogleSheet.query.filter(
            GoogleSheet.academic_year == virtual_year,
            GoogleSheet.purpose == "virtual_district_reports",
            GoogleSheet.sheet_name.ilike(f"%{district_name}%"),
        ).first()

        if sheet:
            return redirect(
                url_for("virtual.view_virtual_google_sheet", sheet_id=sheet.id)
            )
        else:
            flash(
                f"No Google Sheet found for {district_name} in {virtual_year}.",
                "warning",
            )
            return redirect(
                url_for(
                    "virtual.virtual_usage_district",
                    district_name=district_name,
                    year=virtual_year,
                )
            )

    @virtual_bp.route("/usage/usage/district/<district_name>/teachers")
    @login_required
    @district_scoped_required
    def virtual_district_teacher_breakdown(district_name):
        """
        Show detailed teacher breakdown by school for a specific district.

        Args:
            district_name: Name of the district

        Returns:
            Rendered template with teachers grouped by school
        """
        # Get filter parameters
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle date parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Get detailed teacher breakdown data
        teacher_school_breakdown = compute_teacher_school_breakdown(
            district_name, selected_virtual_year, date_from, date_to
        )

        virtual_year_options = generate_school_year_options()

        return render_template(
            "virtual/virtual_teacher_breakdown.html",
            district_name=district_name,
            teacher_school_breakdown=teacher_school_breakdown,
            current_filters=current_filters,
            virtual_year_options=virtual_year_options,
        )

    @virtual_bp.route("/usage/usage/district/<district_name>/teacher-progress")
    @login_required
    @district_scoped_required
    def virtual_district_teacher_progress(district_name):
        """
        Show teacher progress tracking for specific teachers in Kansas City Kansas Public Schools.
        This view tracks progress for a predefined set of teachers to ensure district goals are met.

        Args:
            district_name: Name of the district (restricted to Kansas City Kansas Public Schools)

        Returns:
            Rendered template with teacher progress data
        """
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Get filter parameters
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle date parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Get teacher progress tracking data
        teacher_progress_data = compute_teacher_progress_tracking(
            district_name, selected_virtual_year, date_from, date_to
        )

        virtual_year_options = generate_school_year_options()

        # Compute the last time virtual session data was updated (any event update)
        from sqlalchemy.sql import func as _sql_func

        from models.event import Event, EventType

        last_virtual_update = (
            db.session.query(_sql_func.max(Event.updated_at))
            .filter(Event.type == EventType.VIRTUAL_SESSION)
            .scalar()
        )

        return render_template(
            "virtual/virtual_teacher_progress.html",
            district_name=district_name,
            teacher_progress_data=teacher_progress_data,
            current_filters=current_filters,
            virtual_year_options=virtual_year_options,
            last_virtual_update=last_virtual_update,
        )

    @virtual_bp.route("/usage/district/<district_name>/teacher-progress/google-sheets")
    @login_required
    @admin_required
    def virtual_teacher_progress_google_sheets(district_name):
        """Manage Google Sheets for teacher progress tracking"""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Prevent district-scoped users from accessing management UI
        if current_user.scope_type == "district":
            flash("Access denied for district-scoped accounts.", "error")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        virtual_year = request.args.get("year", get_current_virtual_year())

        # Get all Google Sheets for teacher progress tracking for this year
        sheets = (
            GoogleSheet.query.filter_by(
                academic_year=virtual_year, purpose="teacher_progress_tracking"
            )
            .order_by(GoogleSheet.sheet_name)
            .all()
        )

        return render_template(
            "virtual/virtual_teacher_progress_google_sheets.html",
            sheets=sheets,
            district_name=district_name,
            virtual_year=virtual_year,
            virtual_year_options=generate_school_year_options(),
        )

    @virtual_bp.route(
        "/usage/district/<district_name>/teacher-progress/google-sheets/create",
        methods=["POST"],
    )
    @login_required
    @admin_required
    def create_teacher_progress_google_sheet(district_name):
        """Create a new Google Sheet for teacher progress tracking"""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Block district-scoped users
        if current_user.scope_type == "district":
            flash("Access denied for district-scoped accounts.", "error")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        try:
            virtual_year = request.form.get("virtual_year")
            sheet_id = request.form.get("sheet_id")
            sheet_name = request.form.get("sheet_name")

            if not all([virtual_year, sheet_id, sheet_name]):
                flash("All fields are required.", "error")
                return redirect(
                    url_for(
                        "virtual.virtual_teacher_progress_google_sheets",
                        district_name=district_name,
                        year=virtual_year,
                    )
                )

            # Check if sheet already exists for this year
            existing_sheet = GoogleSheet.query.filter_by(
                academic_year=virtual_year,
                purpose="teacher_progress_tracking",
            ).first()

            if existing_sheet:
                flash(
                    f"A Google Sheet already exists for teacher progress tracking in {virtual_year}.",
                    "error",
                )
                return redirect(
                    url_for(
                        "virtual.virtual_teacher_progress_google_sheets",
                        district_name=district_name,
                        year=virtual_year,
                    )
                )

            # Create new Google Sheet record
            new_sheet = GoogleSheet(
                academic_year=virtual_year,
                purpose="teacher_progress_tracking",
                sheet_id=sheet_id,
                sheet_name=sheet_name,
                created_by=current_user.id,
            )

            db.session.add(new_sheet)
            db.session.commit()

            flash(
                f'Google Sheet "{sheet_name}" created successfully for teacher progress tracking.',
                "success",
            )
            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=virtual_year,
                )
            )

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating Google Sheet: {str(e)}", "error")
            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=virtual_year,
                )
            )

    @virtual_bp.route(
        "/usage/district/<district_name>/teacher-progress/google-sheets/<int:sheet_id>/import",
        methods=["POST"],
    )
    @login_required
    @admin_required
    def import_teacher_progress_data(district_name, sheet_id):
        """Import teacher progress data from Google Sheet"""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Block district-scoped users
        if current_user.scope_type == "district":
            flash("Access denied for district-scoped accounts.", "error")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        try:
            sheet = GoogleSheet.query.get_or_404(sheet_id)

            if sheet.purpose != "teacher_progress_tracking":
                flash("Invalid sheet type for teacher progress tracking.", "error")
                return redirect(
                    url_for(
                        "virtual.virtual_teacher_progress_google_sheets",
                        district_name=district_name,
                        year=sheet.academic_year,
                    )
                )

            # Import teacher progress data from Google Sheet
            import pandas as pd

            # Get the sheet ID from the GoogleSheet record
            sheet_id = sheet.sheet_id
            if not sheet_id:
                flash("No sheet ID found for this Google Sheet.", "error")
                return redirect(
                    url_for(
                        "virtual.virtual_teacher_progress_google_sheets",
                        district_name=district_name,
                        year=sheet.academic_year,
                    )
                )

            # Create CSV URL for Google Sheets - use export format which works better
            csv_url = (
                f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            )

            try:
                # Read data from Google Sheet
                df = pd.read_csv(csv_url)
            except Exception as e:
                # Try alternative URL format if the first one fails
                csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
                try:
                    df = pd.read_csv(csv_url)
                except Exception as e2:
                    flash(f"Error reading Google Sheet: {str(e2)}", "error")
                    return redirect(
                        url_for(
                            "virtual.virtual_teacher_progress_google_sheets",
                            district_name=district_name,
                            year=sheet.academic_year,
                        )
                    )

            # Convert DataFrame to list of dictionaries
            # Expected columns: Building, Name, Email, Grade
            sample_teachers = []
            for _, row in df.iterrows():
                # Skip empty rows
                if pd.isna(row.get("Building")) or pd.isna(row.get("Name")):
                    continue

                teacher_data = {
                    "building": str(row.get("Building", "")).strip(),
                    "name": str(row.get("Name", "")).strip(),
                    "email": str(row.get("Email", "")).strip(),
                    "grade": str(row.get("Grade", "")).strip(),
                }

                # Only add if we have required fields
                if (
                    teacher_data["building"]
                    and teacher_data["name"]
                    and teacher_data["email"]
                ):
                    sample_teachers.append(teacher_data)

            # Clear existing data for this academic year
            from models import TeacherProgress

            TeacherProgress.query.filter_by(
                academic_year=sheet.academic_year, virtual_year=sheet.academic_year
            ).delete()

            # Import new data
            imported_count = 0
            for teacher_data in sample_teachers:
                teacher_progress = TeacherProgress(
                    academic_year=sheet.academic_year,
                    virtual_year=sheet.academic_year,
                    building=teacher_data["building"],
                    name=teacher_data["name"],
                    email=teacher_data["email"],
                    grade=teacher_data["grade"],
                    target_sessions=1,  # Default target of 1 session
                    created_by=current_user.id,
                )
                db.session.add(teacher_progress)
                imported_count += 1

            db.session.commit()

            flash(
                f"Successfully imported {imported_count} teachers from '{sheet.sheet_name}' for {sheet.academic_year}.",
                "success",
            )

            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=sheet.academic_year,
                )
            )

        except Exception as e:
            flash(f"Error importing teacher progress data: {str(e)}", "error")
            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=sheet.academic_year,
                )
            )

    @virtual_bp.route("/usage/usage/district/<district_name>/teacher-progress/export")
    @login_required
    @district_scoped_required
    def virtual_district_teacher_progress_export(district_name):
        """
        Export teacher progress tracking data to Excel for Kansas City Kansas Public Schools.
        This exports both summary data and detailed teacher breakdown.

        Args:
            district_name: Name of the district (restricted to Kansas City Kansas Public Schools)

        Returns:
            Excel file with teacher progress data
        """
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Get filter parameters
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle date parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            except ValueError:
                date_from = default_date_from
        else:
            date_from = default_date_from

        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                date_to = default_date_to
        else:
            date_to = default_date_to

        # Get teacher progress data
        teacher_progress_data = compute_teacher_progress_tracking(
            district_name, selected_virtual_year, date_from, date_to
        )

        if not teacher_progress_data:
            flash("No teacher progress data available for export.", "warning")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        # Generate Excel file
        excel_data = generate_teacher_progress_excel(
            teacher_progress_data,
            district_name,
            selected_virtual_year,
            date_from,
            date_to,
        )

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Teacher_Progress_Report_{timestamp}.xlsx"

        # Return Excel file
        return Response(
            excel_data,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @virtual_bp.route(
        "/usage/district/<district_name>/teacher-progress/google-sheets/<int:sheet_id>/delete",
        methods=["POST"],
    )
    @login_required
    @admin_required
    def delete_teacher_progress_google_sheet(district_name, sheet_id):
        """Delete a Google Sheet for teacher progress tracking"""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Block district-scoped users
        if current_user.scope_type == "district":
            flash("Access denied for district-scoped accounts.", "error")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        try:
            sheet = GoogleSheet.query.get_or_404(sheet_id)
            virtual_year = sheet.academic_year
            sheet_name = sheet.sheet_name

            db.session.delete(sheet)
            db.session.commit()

            flash(f'Google Sheet "{sheet_name}" deleted successfully.', "success")
            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=virtual_year,
                )
            )

        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting Google Sheet: {str(e)}", "error")
            return redirect(
                url_for(
                    "virtual.virtual_teacher_progress_google_sheets",
                    district_name=district_name,
                    year=virtual_year,
                )
            )

    @virtual_bp.route("/usage/recruitment")
    @login_required
    def virtual_recruitment():
        """
        Virtual Sessions Presenter Recruitment View

        Displays upcoming virtual events without assigned presenters to enable
        proactive recruitment of volunteers. Only accessible to Admin users and
        global-scoped regular users (not district or school-scoped users).

        Returns:
            Rendered template with filtered list of events needing presenters
        """
        # Access Control: Allow Admin (any scope) OR User with global scope
        # Deny: District-scoped or School-scoped users
        if not (current_user.is_admin or current_user.scope_type == "global"):
            flash(
                "Access denied. This feature is only available to administrators and global users.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range based on virtual session year (Aug 1 - Jul 31)
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        # Current timestamp for filtering future events
        # Use both timezone-aware and naive versions for compatibility
        now = datetime.now(timezone.utc)
        now_naive = datetime.now()  # For databases with naive datetime storage

        # Build query for virtual events without presenter assignments
        # Subquery to check for presenter assignments
        from sqlalchemy import and_
        from sqlalchemy import exists as sql_exists

        presenter_exists = sql_exists().where(
            and_(
                EventParticipation.event_id == Event.id,
                EventParticipation.participant_type == "Presenter",
            )
        )

        # Main query: Virtual events, future only, without presenters
        # Use OR condition to handle both timezone-aware and naive datetimes
        query = Event.query.filter(
            Event.type == EventType.VIRTUAL_SESSION,
            or_(Event.start_date > now, Event.start_date > now_naive),
            ~presenter_exists,  # NOT EXISTS
        )

        # Apply date range filter
        query = query.filter(Event.start_date >= date_from, Event.start_date <= date_to)

        # Apply school filter
        school_id = request.args.get("school")
        if school_id:
            query = query.filter(Event.school == school_id)

        # Apply district filter (via school relationship)
        district_name = request.args.get("district")
        if district_name:
            query = query.join(School).filter(School.district_id == district_name)

        # Apply event type filter (allows filtering to specific virtual event subtypes if needed)
        event_type = request.args.get("event_type")
        if event_type:
            try:
                query = query.filter(Event.type == EventType(event_type))
            except (ValueError, KeyError):
                pass  # Invalid event type, ignore filter

        # Apply search filter across title and teacher names
        search_term = request.args.get("search", "").strip()
        if search_term:
            query = query.filter(
                or_(
                    Event.title.ilike(f"%{search_term}%"),
                    Event.educators.ilike(f"%{search_term}%"),
                )
            )

        # Sort by start date ascending (earliest events first = highest priority)
        query = query.order_by(Event.start_date.asc())

        # Execute query
        events = query.all()

        # Prepare event data with calculated fields
        event_data = []
        for event in events:
            # Ensure event.start_date is timezone-aware for comparison
            event_start = event.start_date
            if event_start.tzinfo is None:
                # If naive, assume UTC
                event_start = event_start.replace(tzinfo=timezone.utc)

            # Calculate days until event
            days_until = (event_start - now).days

            # Count tagged teachers
            teacher_count = EventTeacher.query.filter_by(event_id=event.id).count()

            # Get school and district info
            school = School.query.get(event.school) if event.school else None
            district_display = (
                school.district.name if school and school.district else "N/A"
            )
            school_name = school.name if school else "N/A"

            # Format date for display (use original event.start_date for formatting)
            try:
                formatted_date = event.start_date.strftime("%b %d, %Y %I:%M %p")
            except:
                formatted_date = str(event.start_date)

            event_data.append(
                {
                    "id": event.id,
                    "title": event.title,
                    "start_date": event.start_date,
                    "school_name": school_name,
                    "district_name": district_display,
                    "teacher_count": teacher_count,
                    "days_until": days_until,
                    "formatted_date": formatted_date,
                    "status": (
                        event.status.value
                        if hasattr(event.status, "value")
                        else str(event.status)
                    ),
                }
            )

        # Prepare filter options for dropdowns
        # Get all schools with virtual events
        schools_query = (
            db.session.query(School)
            .join(Event)
            .filter(Event.type == EventType.VIRTUAL_SESSION)
            .distinct()
            .order_by(School.name)
        )
        schools = schools_query.all()

        # Get all districts with virtual events
        districts_query = (
            db.session.query(District)
            .join(School)
            .join(Event)
            .filter(Event.type == EventType.VIRTUAL_SESSION)
            .distinct()
            .order_by(District.name)
        )
        districts = districts_query.all()

        # Generate school year options
        school_years = generate_school_year_options()

        # Build filter options
        filter_options = {
            "schools": schools,
            "districts": districts,
            "school_years": school_years,
            "event_types": [et for et in EventType],  # All event types
        }

        # Current filters for template
        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
            "school": school_id,
            "district": district_name,
            "event_type": event_type,
            "search": search_term,
        }

        return render_template(
            "virtual/virtual_recruitment.html",
            events=event_data,
            filter_options=filter_options,
            current_filters=current_filters,
            event_count=len(event_data),
        )

    @virtual_bp.route("/usage/district/<district_name>/teacher-progress/matching")
    @login_required
    @admin_required
    def virtual_teacher_progress_matching(district_name):
        """
        Admin interface for matching TeacherProgress entries to Teacher records.

        Args:
            district_name: Name of the district (restricted to Kansas City Kansas Public Schools)

        Returns:
            Rendered template with matching interface
        """
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            flash(
                "This view is only available for Kansas City Kansas Public Schools.",
                "error",
            )
            return redirect(url_for("virtual.virtual_usage"))

        # Block district-scoped users
        if current_user.scope_type == "district":
            flash("Access denied for district-scoped accounts.", "error")
            return redirect(
                url_for(
                    "virtual.virtual_district_teacher_progress",
                    district_name=district_name,
                )
            )

        virtual_year = request.args.get("year", get_current_virtual_year())
        show_unmatched_only = (
            request.args.get("unmatched_only", "false").lower() == "true"
        )
        search_query = request.args.get("search", "").strip()

        # Get all TeacherProgress entries for this virtual year
        query = TeacherProgress.query.filter_by(virtual_year=virtual_year)

        # Filter by unmatched only if requested
        if show_unmatched_only:
            query = query.filter_by(teacher_id=None)

        # Apply search filter if provided
        if search_query:
            search_pattern = f"%{search_query}%"
            query = query.filter(
                or_(
                    TeacherProgress.name.ilike(search_pattern),
                    TeacherProgress.email.ilike(search_pattern),
                    TeacherProgress.building.ilike(search_pattern),
                )
            )

        teacher_progress_entries = query.order_by(
            TeacherProgress.building, TeacherProgress.name
        ).all()

        # Get all teachers for manual matching dropdown
        teachers = Teacher.query.order_by(Teacher.last_name, Teacher.first_name).all()

        # Get match statistics
        total_entries = TeacherProgress.query.filter_by(
            virtual_year=virtual_year
        ).count()
        matched_entries = (
            TeacherProgress.query.filter_by(virtual_year=virtual_year)
            .filter(TeacherProgress.teacher_id.isnot(None))
            .count()
        )
        unmatched_entries = total_entries - matched_entries

        return render_template(
            "virtual/virtual_teacher_progress_matching.html",
            district_name=district_name,
            teacher_progress_entries=teacher_progress_entries,
            teachers=teachers,
            virtual_year=virtual_year,
            virtual_year_options=generate_school_year_options(),
            show_unmatched_only=show_unmatched_only,
            search_query=search_query,
            total_entries=total_entries,
            matched_entries=matched_entries,
            unmatched_entries=unmatched_entries,
        )

    @virtual_bp.route(
        "/usage/district/<district_name>/teacher-progress/matching/match",
        methods=["POST"],
    )
    @login_required
    @admin_required
    def match_teacher_progress(district_name):
        """Manually match a TeacherProgress entry to a Teacher record."""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            return jsonify({"success": False, "message": "Invalid district"}), 400

        # Block district-scoped users
        if current_user.scope_type == "district":
            return jsonify({"success": False, "message": "Access denied"}), 403

        try:
            teacher_progress_id = request.json.get("teacher_progress_id")
            teacher_id = request.json.get("teacher_id")

            if not teacher_progress_id:
                return (
                    jsonify(
                        {"success": False, "message": "Missing teacher_progress_id"}
                    ),
                    400,
                )

            teacher_progress = TeacherProgress.query.get_or_404(teacher_progress_id)

            if teacher_id:
                # Match to teacher
                teacher = Teacher.query.get_or_404(teacher_id)
                teacher_progress.teacher_id = teacher.id
                message = f'Matched "{teacher_progress.name}" to "{teacher.first_name} {teacher.last_name}"'
            else:
                # Unmatch (set to None)
                teacher_progress.teacher_id = None
                message = f'Removed match for "{teacher_progress.name}"'

            db.session.commit()
            return jsonify({"success": True, "message": message})

        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)}), 500

    @virtual_bp.route(
        "/usage/district/<district_name>/teacher-progress/matching/auto-match",
        methods=["POST"],
    )
    @login_required
    @admin_required
    def auto_match_teacher_progress(district_name):
        """Run automatic matching for TeacherProgress entries."""
        # Restrict access to Kansas City Kansas Public Schools only
        if district_name != "Kansas City Kansas Public Schools":
            return jsonify({"success": False, "message": "Invalid district"}), 400

        # Block district-scoped users
        if current_user.scope_type == "district":
            return jsonify({"success": False, "message": "Access denied"}), 403

        try:
            virtual_year = request.json.get("virtual_year", get_current_virtual_year())
            stats = match_teacher_progress_to_teachers(virtual_year=virtual_year)

            return jsonify(
                {
                    "success": True,
                    "message": f"Auto-matching completed: {stats['matched_by_email']} by email, {stats['matched_by_name']} by name, {stats['unmatched']} unmatched",
                    "stats": stats,
                }
            )

        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500


def compute_teacher_school_breakdown(district_name, virtual_year, date_from, date_to):
    """
    Compute teacher breakdown grouped by school for a specific district.
    Includes both completed sessions, no-show sessions, and upcoming sessions for each teacher.

    Args:
        district_name: Name of the district
        virtual_year: The virtual year
        date_from: Start date
        date_to: End date

    Returns:
        Dictionary with schools as keys and teacher data as values
    """
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)
    # Base query for virtual session events
    base_query = Event.query.options(
        joinedload(Event.districts),
        joinedload(Event.teacher_registrations)
        .joinedload(EventTeacher.teacher)
        .joinedload(Teacher.school),
    ).filter(
        Event.type == EventType.VIRTUAL_SESSION,
        Event.start_date >= date_from,
        Event.start_date <= date_to,
    )

    events = base_query.all()

    # Dictionary to store school -> teachers -> session counts
    school_teacher_data = {}

    for event in events:
        # Process each teacher registration for both completed and no-show sessions
        for teacher_reg in event.teacher_registrations:
            teacher = teacher_reg.teacher
            if not teacher:
                continue

            # Get teacher info
            teacher_name = f"{teacher.first_name} {teacher.last_name}"
            teacher_id = teacher.id

            # Determine teacher's school and district (use teacher's school for accuracy)
            school_name = "Unknown School"
            teacher_district_name = None
            if hasattr(teacher, "school_obj") and teacher.school_obj:
                school_obj = teacher.school_obj
                school_name = school_obj.name
                if hasattr(school_obj, "district") and school_obj.district:
                    teacher_district_name = school_obj.district.name
            elif teacher.school_id:
                school_obj = School.query.get(teacher.school_id)
                if school_obj:
                    school_name = school_obj.name
                    if hasattr(school_obj, "district") and school_obj.district:
                        teacher_district_name = school_obj.district.name

            # Only include registrations for teachers in the requested district
            # Use helper function to handle aliases (e.g., "KCPS (MO)" vs "Kansas City Public Schools (MO)")
            if not _district_name_matches(district_name, teacher_district_name):
                continue

            # Initialize school if not exists
            if school_name not in school_teacher_data:
                school_teacher_data[school_name] = {}

            # Initialize teacher if not exists
            if teacher_id not in school_teacher_data[school_name]:
                school_teacher_data[school_name][teacher_id] = {
                    "id": teacher_id,
                    "name": teacher_name,
                    "sessions": 0,
                    "no_shows": 0,
                    "upcoming_sessions": 0,
                }

            # Determine classification for this registration
            def _norm(x):
                return (x or "").strip().lower()

            def _sanitize(text: str) -> str:
                t = (text or "").lower().strip()
                for ch in ["_", "-", "/", "\\", ",", ".", "  "]:
                    t = t.replace(ch, " ")
                while "  " in t:
                    t = t.replace("  ", " ")
                return t

            original_status = _sanitize(getattr(event, "original_status_string", None))
            moved_to_in_person = "moved to in-person" in original_status
            # Consider common variations for teacher no-show at the event level
            event_teacher_no_show = ("teacher no show" in original_status) or (
                "teacher did not attend" in original_status
            )
            tr_status = _sanitize(getattr(teacher_reg, "status", None))

            is_teacher_no_show = (
                "teacher no show" in tr_status
                or "no show" in tr_status
                or "did not attend" in tr_status
            )
            is_teacher_cancel = (
                "cancel" in tr_status
                or "withdraw" in tr_status
                or "inclement weather" in tr_status
                or "technical" in tr_status
            )
            completed_by_status = ("attended" in tr_status) or (
                "completed" in tr_status
            )

            counted_completed = False
            counted_no_show = False

            # Check for upcoming sessions (Teacher requested or Industry chat + Draft status)
            # Only count upcoming sessions that are actually in the future
            session_type = getattr(event, "additional_information", "")
            event_status = getattr(event, "status", "")
            teacher_reg_status = getattr(teacher_reg, "status", "")

            # Handle enum status values
            event_status_str = (
                str(event_status.value)
                if hasattr(event_status, "value")
                else str(event_status)
            )
            teacher_reg_status_str = (
                str(teacher_reg_status) if teacher_reg_status else ""
            )

            # Check if this is an upcoming session AND it's actually in the future
            # Ensure both datetimes are timezone-aware for comparison
            event_start_date = event.start_date
            if event_start_date and event_start_date.tzinfo is None:
                # If event.start_date is timezone-naive, assume it's in UTC
                event_start_date = event_start_date.replace(tzinfo=timezone.utc)

            is_upcoming_session = (
                (
                    "teacher requested" in (session_type or "").lower()
                    or "industry chat" in (session_type or "").lower()
                )
                and (
                    "draft" in event_status_str.lower()
                    or "draft" in teacher_reg_status_str.lower()
                )
                and event_start_date
                and event_start_date > now  # Only future sessions
            )

            if is_upcoming_session:
                school_teacher_data[school_name][teacher_id]["upcoming_sessions"] += 1
            # Prioritize explicit no-show status over attendance_confirmed_at
            elif is_teacher_no_show or (
                event_teacher_no_show and not moved_to_in_person
            ):
                school_teacher_data[school_name][teacher_id]["no_shows"] += 1
                counted_no_show = True
            elif (
                moved_to_in_person
                or completed_by_status
                or teacher_reg.attendance_confirmed_at
            ):
                school_teacher_data[school_name][teacher_id]["sessions"] += 1
                counted_completed = True
            else:
                # Ignore cancellations and indeterminate statuses (do not count as no-show)
                if not is_teacher_cancel:
                    # If truly indeterminate and attendance not confirmed, do nothing
                    pass
                else:
                    pass

    # Convert to sorted structure for template
    school_breakdown = {}
    for school_name, teachers_dict in school_teacher_data.items():
        # Convert teachers dict to sorted list
        teachers_list = sorted(
            teachers_dict.values(),
            key=lambda x: (
                -(
                    x["sessions"] + x["no_shows"] + x["upcoming_sessions"]
                ),  # Sort by total sessions (completed + no-shows + upcoming) desc
                x["name"],
            ),  # Then by name asc
        )

        # Include teachers with at least 1 session (completed, no-show, or upcoming)
        teachers_with_activity = [
            t
            for t in teachers_list
            if t["sessions"] > 0 or t["no_shows"] > 0 or t["upcoming_sessions"] > 0
        ]

        if (
            teachers_with_activity
        ):  # Only include schools that have teachers with activity
            school_breakdown[school_name] = {
                "teachers": teachers_with_activity,
                "total_teachers": len(teachers_with_activity),
                "total_sessions": sum(t["sessions"] for t in teachers_with_activity),
                "total_no_shows": sum(t["no_shows"] for t in teachers_with_activity),
                "total_upcoming_sessions": sum(
                    t["upcoming_sessions"] for t in teachers_with_activity
                ),
            }

    # Sort schools by total activity (sessions + no-shows + upcoming_sessions) (descending)
    sorted_schools = dict(
        sorted(
            school_breakdown.items(),
            key=lambda x: (
                -(
                    x[1]["total_sessions"]
                    + x[1]["total_no_shows"]
                    + x[1]["total_upcoming_sessions"]
                ),
                x[0],
            ),  # Sort by total activity desc, then school name asc
        )
    )

    return sorted_schools


def compute_teacher_progress_tracking(district_name, virtual_year, date_from, date_to):
    """
    Compute teacher progress tracking for specific teachers in Kansas City Kansas Public Schools.
    This function will track progress for a predefined set of teachers imported from a spreadsheet.

    Args:
        district_name: Name of the district (should be "Kansas City Kansas Public Schools")
        virtual_year: The virtual year
        date_from: Start date
        date_to: End date

    Returns:
        Dictionary with teacher progress data grouped by school
    """
    from datetime import datetime, timezone

    from models import Event, EventTeacher, School, Teacher, TeacherProgress
    from models.event import EventStatus, EventType

    # Get all teachers from the progress tracking table for this virtual year
    teachers = TeacherProgress.query.filter_by(virtual_year=virtual_year).all()

    if not teachers:
        return {}

    # Get virtual session events for the date range
    events = Event.query.filter(
        Event.type == EventType.VIRTUAL_SESSION,
        Event.start_date >= date_from,
        Event.start_date <= date_to,
    ).all()

    # Create mappings for unique teachers and their name variations
    teacher_progress_map = {}
    teacher_alias_map = {}
    for teacher in teachers:
        teacher_progress_map[teacher.id] = {
            "teacher": teacher,
            "completed_sessions": 0,
            "planned_sessions": 0,
            "no_show_count": 0,  # Track no-shows to affect status calculation
        }

        # Store multiple possible name variations for matching
        # Normalize names by replacing hyphens with spaces for better matching
        base_name = teacher.name.lower().strip()
        normalized_name = base_name.replace("-", " ").replace(".", "").replace(",", "")

        name_variations = [
            base_name,  # Original with hyphens
            normalized_name,  # Normalized (hyphens -> spaces)
            base_name.replace(".", "").replace(",", "").strip(),
            # Add first + last name variation if different from stored name
            (
                f"{teacher.name.split()[0]} {teacher.name.split()[-1]}".lower()
                if len(teacher.name.split()) > 1
                else teacher.name.lower()
            ),
            # Also add normalized first + last
            (
                f"{teacher.name.split()[0].lower()} {teacher.name.split()[-1].lower()}".replace(
                    "-", " "
                )
                if len(teacher.name.split()) > 1
                else normalized_name
            ),
        ]

        # Ensure aliases point back to the unique teacher entry
        for name_var in name_variations:
            if name_var:
                teacher_alias_map.setdefault(name_var, teacher.id)

    # Count completed and planned sessions for each teacher
    for event in events:
        for teacher_reg in event.teacher_registrations:
            if teacher_reg.teacher:
                teacher_name = (
                    f"{teacher_reg.teacher.first_name} {teacher_reg.teacher.last_name}"
                )
                teacher_key = teacher_name.lower().strip()
                # Normalize the key by replacing hyphens with spaces for better matching
                teacher_key_normalized = teacher_key.replace("-", " ")

                # Try exact match first against aliases (both original and normalized)
                teacher_id = teacher_alias_map.get(
                    teacher_key
                ) or teacher_alias_map.get(teacher_key_normalized)
                progress_data = (
                    teacher_progress_map.get(teacher_id) if teacher_id else None
                )

                if not progress_data:
                    # Try flexible matching - look for partial matches
                    # Compare both original and normalized versions
                    for name_key, alias_teacher_id in teacher_alias_map.items():
                        name_key_normalized = name_key.replace("-", " ")
                        # Check if either version matches (original or normalized)
                        if (
                            teacher_key in name_key
                            or name_key in teacher_key
                            or teacher_key_normalized in name_key_normalized
                            or name_key_normalized in teacher_key_normalized
                        ) and len(teacher_key) > 3:
                            progress_data = teacher_progress_map.get(alias_teacher_id)
                            if progress_data:
                                break

                if progress_data:
                    # Normalize status strings for robust matching (similar to compute_teacher_school_breakdown)
                    def _sanitize_status(text):
                        """Normalize status text for comparison"""
                        if not text:
                            return ""
                        t = str(text).lower().strip()
                        # Replace common separators with spaces
                        for ch in ["_", "-", "/", "\\", ",", ".", "  "]:
                            t = t.replace(ch, " ")
                        # Normalize multiple spaces
                        while "  " in t:
                            t = t.replace("  ", " ")
                        return t

                    # Check teacher's individual registration status first
                    teacher_reg_status = _sanitize_status(
                        getattr(teacher_reg, "status", "")
                    )
                    # Also check event-level status for teacher no-shows
                    event_original_status = _sanitize_status(
                        getattr(event, "original_status_string", "")
                    )
                    # Check if event status enum is NO_SHOW
                    event_status_is_no_show = event.status == EventStatus.NO_SHOW

                    # Comprehensive no-show detection
                    # Check teacher registration status first (most specific)
                    is_teacher_no_show = (
                        "teacher no show" in teacher_reg_status
                        or "no show" in teacher_reg_status
                        or "did not attend" in teacher_reg_status
                    )

                    # If teacher registration doesn't have no-show, check event-level status
                    if not is_teacher_no_show:
                        is_teacher_no_show = (
                            "teacher no show" in event_original_status
                            or "teacher did not attend" in event_original_status
                            # Check if event status enum is NO_SHOW and original status mentions teacher
                            or (
                                event_status_is_no_show
                                and "teacher" in event_original_status
                            )
                        )
                    is_teacher_cancel = (
                        "cancel" in teacher_reg_status
                        or "withdraw" in teacher_reg_status
                        or "cancelled" in teacher_reg_status
                    )

                    # Track no-shows - they should not count as completed or planned,
                    # but they affect the status calculation (teacher needs to replan)
                    if is_teacher_no_show:
                        progress_data["no_show_count"] += 1
                        continue
                    if is_teacher_cancel:
                        continue

                    # Check if teacher registration has "count" status - this should count as completed
                    # even if the event status is NO_SHOW (as long as it's not a teacher no-show)
                    is_count_status = (
                        teacher_reg_status == "count" or "count" in teacher_reg_status
                    )

                    # Only count as completed if teacher actually attended (not no-show or cancelled)
                    # Check if this is a completed session
                    if (
                        event.status == EventStatus.COMPLETED
                        or (event.status == EventStatus.SIMULCAST)
                        or (getattr(event, "original_status_string", "") or "").lower()
                        in ["completed", "successfully completed"]
                        # Also count if teacher registration has "count" status
                        or is_count_status
                    ):
                        progress_data["completed_sessions"] += 1
                    # Check if this is a planned/upcoming session
                    elif event.status == EventStatus.DRAFT or (
                        getattr(event, "original_status_string", "") or ""
                    ).lower() in ["draft", "registered"]:
                        progress_data["planned_sessions"] += 1

    # Group teachers by building/school
    school_data = {}
    for progress_data in teacher_progress_map.values():
        teacher = progress_data["teacher"]
        building = teacher.building

        if building not in school_data:
            school_data[building] = {
                "teachers": [],
                "total_teachers": 0,
                "goals_achieved": 0,
                "goals_in_progress": 0,
                "goals_not_started": 0,
            }

        # Calculate progress status
        # If teacher has any no-shows and hasn't completed target, they need to replan
        no_show_count = progress_data.get("no_show_count", 0)
        completed = progress_data["completed_sessions"]
        planned = progress_data["planned_sessions"]
        total_sessions = completed + planned

        # If teacher has no-shows and hasn't completed target, check if they need to replan
        # Only force "Needs Planning" if they don't have enough planned sessions to meet target
        if no_show_count > 0 and completed < teacher.target_sessions:
            # If they have enough planned sessions to meet target, they're "In Progress"
            # Otherwise, they need to replan
            if total_sessions >= teacher.target_sessions:
                # They have enough planned sessions, so they're "In Progress"
                progress_status = teacher.get_progress_status(completed, planned)
            else:
                # Not enough planned sessions, force "Needs Planning" - no-shows mean they need to replan
                progress_status = {
                    "status": "not_started",
                    "status_text": "Needs Planning",
                    "status_class": "not_started",
                    "progress_percentage": (
                        min(100, (completed / teacher.target_sessions) * 100)
                        if teacher.target_sessions > 0
                        else 0
                    ),
                    "completed_sessions": completed,
                    "planned_sessions": planned,
                    "needed_sessions": max(
                        0, teacher.target_sessions - completed - planned
                    ),
                }
        else:
            progress_status = teacher.get_progress_status(completed, planned)

        teacher_info = {
            "id": teacher.id,
            "name": teacher.name,
            "email": teacher.email,
            "grade": teacher.grade,
            "target_sessions": teacher.target_sessions,
            "matched_teacher_id": teacher.teacher_id,  # ID of matched Teacher record
            "completed_sessions": progress_data["completed_sessions"],
            "planned_sessions": progress_data["planned_sessions"],
            "needed_sessions": progress_status["needed_sessions"],
            "progress_percentage": progress_status["progress_percentage"],
            "goal_status_class": progress_status["status_class"],
            "goal_status_text": progress_status["status_text"],
            "progress_class": (
                "danger"
                if progress_status["progress_percentage"] < 50
                else "warning" if progress_status["progress_percentage"] < 100 else ""
            ),
        }

        school_data[building]["teachers"].append(teacher_info)
        school_data[building]["total_teachers"] += 1

        # Count goal statuses
        if progress_status["status"] == "achieved":
            school_data[building]["goals_achieved"] += 1
        elif progress_status["status"] == "in_progress":
            school_data[building]["goals_in_progress"] += 1
        else:
            school_data[building]["goals_not_started"] += 1

    # Sort teachers within each school by progress (achieved first, then by name)
    for building, data in school_data.items():
        data["teachers"].sort(
            key=lambda x: (
                -x["completed_sessions"],  # Completed sessions first
                -x["planned_sessions"],  # Then planned sessions
                x["name"],  # Then alphabetically
            )
        )

    return school_data


def normalize_name(name):
    """
    Normalize a name for matching purposes.

    Args:
        name: Name string to normalize

    Returns:
        Normalized name string (lowercase, no punctuation, single spaces)
    """
    if not name:
        return ""
    # Convert to lowercase
    normalized = name.lower().strip()
    # Remove common punctuation
    for char in [".", ",", "-", "(", ")", "'", '"']:
        normalized = normalized.replace(char, " ")
    # Replace multiple spaces with single space
    while "  " in normalized:
        normalized = normalized.replace("  ", " ")
    return normalized.strip()


def name_similarity(name1, name2):
    """
    Calculate similarity between two names using SequenceMatcher.

    Args:
        name1: First name to compare
        name2: Second name to compare

    Returns:
        Similarity ratio between 0.0 and 1.0
    """
    norm1 = normalize_name(name1)
    norm2 = normalize_name(name2)
    return difflib.SequenceMatcher(None, norm1, norm2).ratio()


def match_teacher_progress_to_teachers(virtual_year=None, min_similarity=0.85):
    """
    Automatically match TeacherProgress entries to Teacher records.

    Matching strategy:
    1. Match by email (exact match on primary email) - highest priority
    2. Match by name (fuzzy matching with similarity threshold) - secondary

    Args:
        virtual_year: Virtual year to match teachers for (None = all years)
        min_similarity: Minimum similarity ratio for name matching (default: 0.85)

    Returns:
        dict: Statistics about matching results
    """
    from models.teacher_progress import TeacherProgress

    # Get all TeacherProgress entries (optionally filtered by virtual_year)
    query = TeacherProgress.query
    if virtual_year:
        query = query.filter_by(virtual_year=virtual_year)

    teacher_progress_entries = query.filter_by(teacher_id=None).all()

    # Get all teachers with their primary emails
    teachers = Teacher.query.all()

    # Build email lookup map (email -> teacher)
    email_to_teacher = {}
    for teacher in teachers:
        primary_email = teacher.emails.filter_by(primary=True).first()
        if primary_email and primary_email.email:
            email_lower = primary_email.email.lower().strip()
            email_to_teacher[email_lower] = teacher

    stats = {
        "total_processed": len(teacher_progress_entries),
        "matched_by_email": 0,
        "matched_by_name": 0,
        "unmatched": 0,
        "errors": [],
    }

    for tp_entry in teacher_progress_entries:
        try:
            matched_teacher = None

            # Strategy 1: Match by email (exact match)
            if tp_entry.email:
                email_lower = tp_entry.email.lower().strip()
                matched_teacher = email_to_teacher.get(email_lower)
                if matched_teacher:
                    stats["matched_by_email"] += 1

            # Strategy 2: Match by name (fuzzy matching) if email didn't match
            if not matched_teacher and tp_entry.name:
                best_match = None
                best_similarity = 0.0

                for teacher in teachers:
                    # Build full name from teacher
                    teacher_full_name = (
                        f"{teacher.first_name} {teacher.last_name}".strip()
                    )
                    if not teacher_full_name:
                        continue

                    similarity = name_similarity(tp_entry.name, teacher_full_name)
                    if similarity > best_similarity and similarity >= min_similarity:
                        best_similarity = similarity
                        best_match = teacher

                if best_match:
                    matched_teacher = best_match
                    stats["matched_by_name"] += 1

            # Update TeacherProgress entry with matched teacher_id
            if matched_teacher:
                tp_entry.teacher_id = matched_teacher.id
            else:
                stats["unmatched"] += 1

        except Exception as e:
            stats["errors"].append(f"Error matching {tp_entry.name}: {str(e)}")
            continue

    # Commit all changes
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        stats["errors"].append(f"Error committing matches: {str(e)}")

    return stats
