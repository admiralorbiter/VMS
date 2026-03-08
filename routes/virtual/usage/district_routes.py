"""
District reporting and export routes for virtual sessions.

Contains routes for district-level session views, Excel exports,
breakdown reports, and teacher-by-school breakdowns. Registered on virtual_bp.
"""

import io
from datetime import datetime, timedelta, timezone

import openpyxl
from flask import (
    Response,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload

from models import db
from models.contact import LocalStatusEnum
from models.district_model import District
from models.event import Event, EventFormat, EventStatus, EventTeacher, EventType
from models.organization import Organization, VolunteerOrganization
from models.reports import VirtualSessionDistrictCache
from models.school_model import School
from models.teacher import Teacher
from models.volunteer import EventParticipation, Volunteer
from routes.decorators import admin_required, district_scoped_required
from routes.reports.common import (
    generate_school_year_options,
    get_current_virtual_year,
    get_school_year_date_range,
    get_virtual_year_dates,
    is_cache_valid,
)
from routes.virtual.routes import virtual_bp

from .cache import (
    get_virtual_session_district_cache,
    invalidate_virtual_session_caches,
    save_virtual_session_district_cache,
)
from .computation import (
    _district_name_matches,
    _get_primary_org_name_for_volunteer,
    apply_runtime_filters,
    apply_sorting_and_pagination,
    calculate_summaries_from_sessions,
    compute_virtual_session_data,
    compute_virtual_session_district_data,
    get_google_sheet_url,
)
from .teacher_progress import compute_teacher_school_breakdown


def load_district_routes():
    @virtual_bp.route("/usage/district/<district_name>")
    @login_required
    @district_scoped_required
    def virtual_usage_district(district_name):
        # Get filter parameters: Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Check if refresh is requested
        refresh_requested = request.args.get("refresh", "0") == "1"

        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        date_from = default_date_from
        date_to = default_date_to
        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass
        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass
        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to
        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }
        virtual_year_options = generate_school_year_options()

        # Check if we're using default full year date range (for caching)
        is_full_year = date_from == default_date_from and date_to == default_date_to

        # If refresh is requested, invalidate cache for this district and virtual year
        if refresh_requested and is_full_year:
            # Invalidate district cache
            cached_data = VirtualSessionDistrictCache.query.filter_by(
                district_name=district_name,
                virtual_year=selected_virtual_year,
                date_from=date_from.date(),
                date_to=date_to.date(),
            ).first()
            if cached_data:
                db.session.delete(cached_data)
                db.session.commit()
            print(
                f"Cache invalidated for district {district_name} virtual session report {selected_virtual_year}"
            )

        # Try to get cached data for full year queries
        cached_data = None
        is_cached = False
        last_refreshed = None

        if is_full_year and not refresh_requested:
            cached_data = get_virtual_session_district_cache(
                district_name, selected_virtual_year, date_from, date_to
            )
            if cached_data:
                is_cached = True
                last_refreshed = cached_data.last_updated
                print(
                    f"Using cached data for district {district_name} virtual session report {selected_virtual_year}"
                )

                # Check if cached data has the new teacher ID fields and correct student calculation
                teacher_breakdown = cached_data.teacher_breakdown
                summary_stats = cached_data.summary_stats

                needs_refresh = False

                # Check for missing teacher ID field
                if teacher_breakdown and len(teacher_breakdown) > 0:
                    sample_teacher = teacher_breakdown[0]
                    if "id" not in sample_teacher:
                        print(
                            f"DEBUG: Cached data missing teacher ID field, invalidating cache"
                        )
                        needs_refresh = True

                # Check if student calculation is using old method (not unique teachers * 25)
                if (
                    summary_stats
                    and "total_teachers" in summary_stats
                    and "total_students" in summary_stats
                ):
                    expected_students = summary_stats["total_teachers"] * 25
                    actual_students = summary_stats["total_students"]
                    if actual_students != expected_students:
                        print(
                            f"DEBUG: Cached data using old student calculation ({actual_students} vs expected {expected_students}), invalidating cache"
                        )
                        needs_refresh = True

                # Force refresh to ensure completed session filtering is applied
                print(
                    f"DEBUG: Forcing cache refresh to ensure completed session filtering is applied"
                )
                needs_refresh = True

                if needs_refresh:
                    db.session.delete(cached_data)
                    db.session.commit()
                    # Force fresh data computation
                    (
                        session_data,
                        monthly_stats,
                        school_breakdown,
                        teacher_breakdown,
                        summary_stats,
                    ) = compute_virtual_session_district_data(
                        district_name, selected_virtual_year, date_from, date_to
                    )
                    is_cached = False
                    last_refreshed = datetime.now(timezone.utc)

                    # Cache the new data
                    save_virtual_session_district_cache(
                        district_name,
                        selected_virtual_year,
                        date_from,
                        date_to,
                        session_data,
                        monthly_stats,
                        school_breakdown,
                        teacher_breakdown,
                        summary_stats,
                    )
                else:
                    # Use cached data
                    session_data = cached_data.session_data
                    monthly_stats = cached_data.monthly_stats
                    school_breakdown = cached_data.school_breakdown
                    summary_stats = cached_data.summary_stats

                return render_template(
                    "virtual/usage/district.html",
                    district_name=district_name,
                    monthly_stats=monthly_stats,
                    current_filters=current_filters,
                    school_year_options=virtual_year_options,
                    total_teachers=summary_stats["total_teachers"],
                    total_students=summary_stats["total_students"],
                    total_unique_sessions=summary_stats["total_unique_sessions"],
                    total_experiences=summary_stats["total_experiences"],
                    total_organizations=summary_stats["total_organizations"],
                    total_professionals=summary_stats["total_professionals"],
                    total_local_professionals=summary_stats.get(
                        "total_local_professionals", 0
                    ),
                    total_professionals_of_color=summary_stats[
                        "total_professionals_of_color"
                    ],
                    total_schools=summary_stats["total_schools"],
                    summary_stats=summary_stats,
                    school_breakdown=school_breakdown,
                    teacher_breakdown=teacher_breakdown,
                    session_data=session_data,
                    last_refreshed=last_refreshed,
                    is_cached=is_cached,
                )

        # If no cache or not full year, compute data fresh
        print(
            f"Computing fresh data for district {district_name} virtual session report {selected_virtual_year}"
        )
        (
            session_data,
            monthly_stats,
            school_breakdown,
            teacher_breakdown,
            summary_stats,
        ) = compute_virtual_session_district_data(
            district_name, selected_virtual_year, date_from, date_to
        )

        # Cache the data if it's a full year query
        if is_full_year:
            save_virtual_session_district_cache(
                district_name,
                selected_virtual_year,
                date_from,
                date_to,
                session_data,
                monthly_stats,
                school_breakdown,
                teacher_breakdown,
                summary_stats,
            )
            # Set the last refreshed time for this fresh data
            last_refreshed = datetime.now(timezone.utc)

        return render_template(
            "virtual/usage/district.html",
            district_name=district_name,
            monthly_stats=monthly_stats,
            current_filters=current_filters,
            school_year_options=virtual_year_options,
            total_teachers=summary_stats["total_teachers"],
            total_students=summary_stats["total_students"],
            total_unique_sessions=summary_stats["total_unique_sessions"],
            total_experiences=summary_stats["total_experiences"],
            total_organizations=summary_stats["total_organizations"],
            total_professionals=summary_stats["total_professionals"],
            total_local_professionals=summary_stats.get("total_local_professionals", 0),
            total_professionals_of_color=summary_stats["total_professionals_of_color"],
            total_schools=summary_stats["total_schools"],
            summary_stats=summary_stats,
            school_breakdown=school_breakdown,
            teacher_breakdown=teacher_breakdown,
            session_data=session_data,
            last_refreshed=last_refreshed,
            is_cached=is_cached,
        )

    @virtual_bp.route("/usage/export")
    @login_required
    def virtual_usage_export():
        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range based on virtual session year
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
            "career_cluster": request.args.get("career_cluster"),
            "school": request.args.get("school"),
            "district": request.args.get("district"),
            "status": request.args.get("status"),
            "search": request.args.get(
                "search"
            ),  # Text search across teacher, title, presenter
        }

        # Base query for virtual session events (same as main route)
        base_query = Event.query.options(
            joinedload(Event.districts),
            joinedload(Event.teacher_registrations)
            .joinedload(EventTeacher.teacher)
            .joinedload(Teacher.school),
        ).filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )

        # Apply filters (same as main route)
        if current_filters["career_cluster"]:
            base_query = base_query.filter(
                Event.series.ilike(f'%{current_filters["career_cluster"]}%')
            )

        if current_filters["status"]:
            base_query = base_query.filter(Event.status == current_filters["status"])

        # Get all events
        events = base_query.order_by(Event.start_date.desc()).all()

        # Build session data (same logic as main route)
        session_data = []

        for event in events:
            teacher_registrations = event.teacher_registrations

            if teacher_registrations:
                for teacher_reg in teacher_registrations:
                    teacher = teacher_reg.teacher

                    school = None
                    school_name = ""
                    school_level = ""

                    if teacher:
                        if hasattr(teacher, "school_obj") and teacher.school_obj:
                            school = teacher.school_obj
                            school_name = school.name
                            school_level = getattr(school, "level", "")
                        elif teacher.school_id:
                            school = School.query.get(teacher.school_id)
                            if school:
                                school_name = school.name
                                school_level = getattr(school, "level", "")

                    district_name = None
                    if event.districts:
                        district_name = event.districts[0].name
                    elif event.district_partner:
                        district_name = event.district_partner
                    else:
                        district_name = "Unknown District"

                    # Apply filters
                    if (
                        current_filters["district"]
                        and district_name != current_filters["district"]
                    ):
                        continue

                    if (
                        current_filters["school"]
                        and school_name
                        and current_filters["school"].lower() not in school_name.lower()
                    ):
                        continue

                    session_data.append(
                        {
                            "status": teacher_reg.status or "registered",
                            "date": (
                                event.start_date.strftime("%m/%d/%y")
                                if event.start_date
                                else ""
                            ),
                            "time": (
                                event.start_date.strftime("%I:%M %p")
                                if event.start_date
                                else ""
                            ),
                            "session_type": event.additional_information or "",
                            "teacher_name": (
                                f"{teacher.first_name} {teacher.last_name}"
                                if teacher
                                else ""
                            ),
                            "teacher_id": teacher.id if teacher else None,
                            "school_name": school_name,
                            "school_level": school_level,
                            "district": district_name,
                            "session_title": event.title,
                            "presenter": (
                                ", ".join([v.full_name for v in event.volunteers])
                                if event.volunteers
                                else ""
                            ),
                            "presenter_organization": (
                                ", ".join(
                                    [
                                        _get_primary_org_name_for_volunteer(v)
                                        or "Independent"
                                        for v in event.volunteers
                                    ]
                                )
                                if event.volunteers
                                else ""
                            ),
                            "presenter_data": (
                                [
                                    {"id": v.id, "name": v.full_name}
                                    for v in event.volunteers
                                ]
                                if event.volunteers
                                else []
                            ),
                            "topic_theme": event.series or "",
                            "session_link": event.registration_link or "",
                            "participant_count": event.participant_count or 0,
                            "duration": event.duration or 0,
                        }
                    )
            else:
                district_name = None
                if event.districts:
                    district_name = event.districts[0].name
                elif event.district_partner:
                    district_name = event.district_partner
                else:
                    district_name = "Unknown District"

                if (
                    current_filters["district"]
                    and district_name != current_filters["district"]
                ):
                    continue

                session_data.append(
                    {
                        "status": event.status.value if event.status else "",
                        "date": (
                            event.start_date.strftime("%m/%d/%y")
                            if event.start_date
                            else ""
                        ),
                        "time": (
                            event.start_date.strftime("%I:%M %p")
                            if event.start_date
                            else ""
                        ),
                        "session_type": event.additional_information or "",
                        "teacher_name": "",
                        "teacher_id": None,
                        "school_name": "",
                        "school_level": "",
                        "district": district_name,
                        "session_title": event.title,
                        "presenter": (
                            ", ".join([v.full_name for v in event.volunteers])
                            if event.volunteers
                            else ""
                        ),
                        "presenter_organization": (
                            ", ".join(
                                [
                                    _get_primary_org_name_for_volunteer(v)
                                    or "Independent"
                                    for v in event.volunteers
                                ]
                            )
                            if event.volunteers
                            else ""
                        ),
                        "presenter_data": (
                            [
                                {"id": v.id, "name": v.full_name}
                                for v in event.volunteers
                            ]
                            if event.volunteers
                            else []
                        ),
                        "topic_theme": event.series or "",
                        "session_link": event.registration_link or "",
                        "participant_count": event.participant_count or 0,
                        "duration": event.duration or 0,
                    }
                )

        # Apply runtime filters (including search)
        session_data = apply_runtime_filters(session_data, current_filters)

        # Apply sorting (same logic as main route)
        sort_column = request.args.get("sort", "date")
        sort_order = request.args.get("order", "desc")

        sortable_columns = {
            "status": "status",
            "date": "date",
            "time": "time",
            "session_type": "session_type",
            "teacher_name": "teacher_name",
            "school_name": "school_name",
            "school_level": "school_level",
            "district": "district",
            "session_title": "session_title",
            "presenter": "presenter",
            "presenter_organization": "presenter_organization",
            "topic_theme": "topic_theme",
        }

        if sort_column in sortable_columns:
            reverse_order = sort_order == "desc"

            if sort_column == "date":

                def date_sort_key(session):
                    try:
                        if session["date"]:
                            date_parts = session["date"].split("/")
                            if len(date_parts) == 3:
                                month, day, year = (
                                    int(date_parts[0]),
                                    int(date_parts[1]),
                                    int(date_parts[2]),
                                )
                                if year < 50:
                                    year += 2000
                                else:
                                    year += 1900
                                return datetime(year, month, day)
                            elif len(date_parts) == 2:
                                month, day = int(date_parts[0]), int(date_parts[1])
                                virtual_year_start = int(
                                    selected_virtual_year.split("-")[0]
                                )
                                year = (
                                    virtual_year_start
                                    if month >= 7
                                    else virtual_year_start + 1
                                )
                                return datetime(year, month, day)
                        return datetime.min
                    except (ValueError, IndexError):
                        return datetime.min

                session_data.sort(key=date_sort_key, reverse=reverse_order)

            elif sort_column == "time":

                def time_sort_key(session):
                    try:
                        if session["time"]:
                            return datetime.strptime(session["time"], "%I:%M %p").time()
                        return datetime.min.time()
                    except ValueError:
                        return datetime.min.time()

                session_data.sort(key=time_sort_key, reverse=reverse_order)

            else:

                def string_sort_key(session):
                    value = session.get(sortable_columns[sort_column], "") or ""
                    return str(value).lower()

                session_data.sort(key=string_sort_key, reverse=reverse_order)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Virtual Session Usage"

        # Define headers
        headers = [
            "Status",
            "Date",
            "Time",
            "Session Type",
            "Teacher Name",
            "School Name",
            "School Level",
            "District",
            "Session Title",
            "Presenter",
            "Topic/Theme",
            "Session Link",
            "Participant Count",
            "Duration (min)",
        ]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="495057", end_color="495057", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Add data
        for row_idx, session in enumerate(session_data, 2):
            data_row = [
                session["status"],
                session["date"],
                session["time"],
                session["session_type"],
                session["teacher_name"],
                session["school_name"],
                session["school_level"],
                session["district"],
                session["session_title"],
                session["presenter"],
                session.get("presenter_organization", ""),
                session["topic_theme"],
                session["session_link"],
                session["participant_count"],
                session["duration"],
            ]

            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

        # Add summary info at the top
        ws.insert_rows(1, 4)

        # Report title
        title_cell = ws.cell(row=1, column=1, value="Virtual Session Usage Report")
        title_cell.font = Font(bold=True, size=16)

        # Export date
        export_date_cell = ws.cell(
            row=2,
            column=1,
            value=f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )
        export_date_cell.font = Font(italic=True)

        # Filters applied
        filter_info = f"Virtual Year: {selected_virtual_year}"
        if current_filters["date_from"] and current_filters["date_to"]:
            filter_info += f" | Date Range: {current_filters['date_from'].strftime('%Y-%m-%d')} to {current_filters['date_to'].strftime('%Y-%m-%d')}"
        if current_filters["district"]:
            filter_info += f" | District: {current_filters['district']}"
        if current_filters["career_cluster"]:
            filter_info += f" | Career Cluster: {current_filters['career_cluster']}"
        if current_filters["status"]:
            filter_info += f" | Status: {current_filters['status']}"

        filter_cell = ws.cell(row=3, column=1, value=f"Filters: {filter_info}")
        filter_cell.font = Font(italic=True, size=10)

        # Total records
        total_cell = ws.cell(
            row=4, column=1, value=f"Total Records: {len(session_data)}"
        )
        total_cell.font = Font(bold=True)

        # Create file in memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"virtual_session_usage_{selected_virtual_year}_{timestamp}.xlsx"

        # Create response
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @virtual_bp.route("/usage/breakdown")
    @login_required
    def virtual_breakdown():
        """
        Detailed breakdown of virtual sessions by month and category.

        Categories:
        - Successfully completed sessions
        - Number of simulcast + Sessions
        - Number of teacher canceled sessions
        - Number of teacher no shows
        - Number of pathful professional canceled/no shows sessions
        - Number of local professional canceled/no show sessions
        - Number of unfilled sessions
        """
        # Get filter parameters - Use virtual year instead of school year
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range based on virtual session year
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle explicit date range parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")

        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Base query for virtual session events
        base_query = Event.query.options(
            joinedload(Event.districts),
            joinedload(Event.teacher_registrations).joinedload(EventTeacher.teacher),
            joinedload(Event.volunteers),
            joinedload(
                Event.school_obj
            ),  # Add school relationship for grade level analysis
        ).filter(
            Event.type == EventType.VIRTUAL_SESSION,
            Event.start_date >= date_from,
            Event.start_date <= date_to,
        )

        # Apply district filtering for district-scoped users
        if current_user.scope_type == "district" and current_user.allowed_districts:
            import json

            try:
                allowed_districts = (
                    json.loads(current_user.allowed_districts)
                    if isinstance(current_user.allowed_districts, str)
                    else current_user.allowed_districts
                )
                # Filter events by allowed districts
                base_query = base_query.filter(
                    Event.districts.any(District.name.in_(allowed_districts))
                )
            except (json.JSONDecodeError, TypeError):
                # If parsing fails, return no events
                base_query = base_query.filter(False)

        events = base_query.all()

        # Initialize monthly breakdown data
        monthly_breakdown = {}

        # Create month entries for the full virtual year (August to July)
        virtual_year_start = int(selected_virtual_year.split("-")[0])
        months = [
            (virtual_year_start, 8, "August"),
            (virtual_year_start, 9, "September"),
            (virtual_year_start, 10, "October"),
            (virtual_year_start, 11, "November"),
            (virtual_year_start, 12, "December"),
            (virtual_year_start + 1, 1, "January"),
            (virtual_year_start + 1, 2, "February"),
            (virtual_year_start + 1, 3, "March"),
            (virtual_year_start + 1, 4, "April"),
            (virtual_year_start + 1, 5, "May"),
            (virtual_year_start + 1, 6, "June"),
            (virtual_year_start + 1, 7, "July"),
        ]

        for year, month_num, month_name in months:
            month_key = f"{year}-{month_num:02d}"
            monthly_breakdown[month_key] = {
                "month_name": month_name,
                "year": year,
                "month": month_num,
                "successfully_completed": 0,
                "simulcast_sessions": 0,
                "teacher_canceled": 0,
                "teacher_no_shows": 0,
                "pathful_professional_canceled_no_shows": 0,
                "local_professional_canceled_no_shows": 0,
                "unfilled_sessions": 0,
                # New fields for volunteer diversity metrics - 4-column structure
                # Sessions Only (completed sessions)
                "local_volunteer_count_sessions_only": 0,
                "local_volunteer_percentage_sessions_only": 0.0,
                "poc_volunteer_count_sessions_only": 0,
                "poc_volunteer_percentage_sessions_only": 0.0,
                # Sessions + Simulcast (completed + simulcast sessions)
                "local_volunteer_count_sessions_simulcast": 0,
                "local_volunteer_percentage_sessions_simulcast": 0.0,
                "poc_volunteer_count_sessions_simulcast": 0,
                "poc_volunteer_percentage_sessions_simulcast": 0.0,
                # Track sessions filled by local/POC volunteers
                "_sessions_with_local_volunteers": set(),
                "_sessions_with_poc_volunteers": set(),
                "_simulcast_sessions_with_local_volunteers": set(),
                "_simulcast_sessions_with_poc_volunteers": set(),
            }

        # Debug: Collect unique status strings to understand what we're working with
        unique_statuses = set()
        unique_additional_info = set()
        unique_series = set()
        unique_school_levels = set()

        # Process each event
        for event in events:
            if not event.start_date:
                continue

            month_key = f"{event.start_date.year}-{event.start_date.month:02d}"

            # Skip if not in our breakdown range
            if month_key not in monthly_breakdown:
                continue

            # Get the original status string for analysis
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()

            # Collect unique statuses for debugging
            if original_status:
                unique_statuses.add(original_status)

            # Collect unique additional info (session types)
            if event.additional_information:
                unique_additional_info.add(event.additional_information)

            # Collect unique series (topics)
            if event.series:
                unique_series.add(event.series)

            # Collect unique school levels
            if event.school_obj and event.school_obj.level:
                unique_school_levels.add(event.school_obj.level)

            # Categorize based on original status string (primary source of truth)
            if "successfully completed" in original_status:
                monthly_breakdown[month_key]["successfully_completed"] += 1

            elif "simulcast" in original_status:
                monthly_breakdown[month_key]["simulcast_sessions"] += 1

            elif "teacher" in original_status and (
                "cancelation" in original_status or "cancellation" in original_status
            ):
                monthly_breakdown[month_key]["teacher_canceled"] += 1

            elif (
                "teacher no-show" in original_status
                or "teacher no show" in original_status
            ):
                monthly_breakdown[month_key]["teacher_no_shows"] += 1

            elif "pathful professional" in original_status:
                monthly_breakdown[month_key][
                    "pathful_professional_canceled_no_shows"
                ] += 1

            elif "local professional" in original_status:
                monthly_breakdown[month_key][
                    "local_professional_canceled_no_shows"
                ] += 1

            elif (
                "technical difficulties" in original_status
                or original_status == "count"
            ):
                monthly_breakdown[month_key]["unfilled_sessions"] += 1

            # Fallback to event status if no original status string
            elif not original_status:
                event_status = event.status
                event_status_str = event_status.value.lower() if event_status else ""

                if event_status == EventStatus.COMPLETED:
                    monthly_breakdown[month_key]["successfully_completed"] += 1
                elif event_status == EventStatus.SIMULCAST:
                    monthly_breakdown[month_key]["simulcast_sessions"] += 1
                elif event_status == EventStatus.NO_SHOW and not event.volunteers:
                    monthly_breakdown[month_key]["unfilled_sessions"] += 1
                elif event_status == EventStatus.CANCELLED and not event.volunteers:
                    monthly_breakdown[month_key]["unfilled_sessions"] += 1

                # Check for professional statuses in main event status
                elif "pathful professional" in event_status_str:
                    monthly_breakdown[month_key][
                        "pathful_professional_canceled_no_shows"
                    ] += 1
                elif "local professional" in event_status_str:
                    monthly_breakdown[month_key][
                        "local_professional_canceled_no_shows"
                    ] += 1

            # Check teacher registrations for additional status information
            for teacher_reg in event.teacher_registrations:
                tr_status = (teacher_reg.status or "").lower().strip()

                # Check for professional statuses FIRST (before teacher-specific logic)
                if "pathful professional" in tr_status:
                    monthly_breakdown[month_key][
                        "pathful_professional_canceled_no_shows"
                    ] += 1
                elif "local professional" in tr_status:
                    monthly_breakdown[month_key][
                        "local_professional_canceled_no_shows"
                    ] += 1

                # Count teacher-specific cancellations and no-shows (only if not professional)
                elif "cancel" in tr_status and "teacher" not in original_status:
                    monthly_breakdown[month_key]["teacher_canceled"] += 1
                elif (
                    "no-show" in tr_status or "no show" in tr_status
                ) and "teacher" not in original_status:
                    monthly_breakdown[month_key]["teacher_no_shows"] += 1

            # Check for simulcast from teacher registrations (additional simulcast detection)
            simulcast_teacher_count = sum(
                1 for tr in event.teacher_registrations if tr.is_simulcast
            )
            if simulcast_teacher_count > 0 and "simulcast" not in original_status:
                monthly_breakdown[month_key]["simulcast_sessions"] += 1

            # Calculate volunteer diversity metrics for successfully completed sessions
            # (which includes simulcast sessions - same volunteer talking to multiple classes)
            if "successfully completed" in original_status or (
                not original_status and event.status == EventStatus.COMPLETED
            ):

                # Track sessions filled by local and POC volunteers
                event_id = event.id or event.title
                if event_id:
                    has_local_volunteer = False
                    has_poc_volunteer = False

                    for volunteer in event.volunteers or []:
                        # Check if volunteer is local
                        try:
                            from models.contact import LocalStatusEnum

                            if (
                                getattr(volunteer, "local_status", None)
                                == LocalStatusEnum.local
                            ):
                                has_local_volunteer = True
                        except Exception:
                            try:
                                if (
                                    str(getattr(volunteer, "local_status", ""))
                                    .lower()
                                    .endswith("local")
                                ):
                                    has_local_volunteer = True
                            except Exception:
                                pass

                        # Check if volunteer is People of Color
                        if getattr(volunteer, "is_people_of_color", False):
                            has_poc_volunteer = True

                    # Track for completed sessions (sessions only)
                    if has_local_volunteer:
                        monthly_breakdown[month_key][
                            "_sessions_with_local_volunteers"
                        ].add(event_id)
                    if has_poc_volunteer:
                        monthly_breakdown[month_key][
                            "_sessions_with_poc_volunteers"
                        ].add(event_id)

                    # Also track for sessions + simulcast (if this is a simulcast session)
                    if "simulcast" in original_status or (
                        not original_status and event.status == EventStatus.SIMULCAST
                    ):
                        if has_local_volunteer:
                            monthly_breakdown[month_key][
                                "_simulcast_sessions_with_local_volunteers"
                            ].add(event_id)
                        if has_poc_volunteer:
                            monthly_breakdown[month_key][
                                "_simulcast_sessions_with_poc_volunteers"
                            ].add(event_id)

        # Calculate final monthly volunteer counts and percentages from unique sets
        for month_key, month_data in monthly_breakdown.items():
            # Calculate session-based metrics for both scenarios

            # 1. Sessions Only (completed sessions)
            month_data["local_volunteer_count_sessions_only"] = len(
                month_data["_sessions_with_local_volunteers"]
            )
            month_data["poc_volunteer_count_sessions_only"] = len(
                month_data["_sessions_with_poc_volunteers"]
            )

            # 2. Sessions + Simulcast (completed + simulcast sessions)
            month_data["local_volunteer_count_sessions_simulcast"] = len(
                month_data["_sessions_with_local_volunteers"]
            ) + len(month_data["_simulcast_sessions_with_local_volunteers"])
            month_data["poc_volunteer_count_sessions_simulcast"] = len(
                month_data["_sessions_with_poc_volunteers"]
            ) + len(month_data["_simulcast_sessions_with_poc_volunteers"])

            # Calculate total sessions for percentage calculations
            total_completed_sessions = month_data["successfully_completed"]
            total_sessions_simulcast = (
                month_data["successfully_completed"] + month_data["simulcast_sessions"]
            )

            # Calculate percentages for Sessions Only
            if total_completed_sessions > 0:
                month_data["local_volunteer_percentage_sessions_only"] = round(
                    (
                        month_data["local_volunteer_count_sessions_only"]
                        / total_completed_sessions
                    )
                    * 100,
                    1,
                )
                month_data["poc_volunteer_percentage_sessions_only"] = round(
                    (
                        month_data["poc_volunteer_count_sessions_only"]
                        / total_completed_sessions
                    )
                    * 100,
                    1,
                )
            else:
                month_data["local_volunteer_percentage_sessions_only"] = 0.0
                month_data["poc_volunteer_percentage_sessions_only"] = 0.0

            # Calculate percentages for Sessions + Simulcast
            if total_sessions_simulcast > 0:
                month_data["local_volunteer_percentage_sessions_simulcast"] = round(
                    (
                        month_data["local_volunteer_count_sessions_simulcast"]
                        / total_sessions_simulcast
                    )
                    * 100,
                    1,
                )
                month_data["poc_volunteer_percentage_sessions_simulcast"] = round(
                    (
                        month_data["poc_volunteer_count_sessions_simulcast"]
                        / total_sessions_simulcast
                    )
                    * 100,
                    1,
                )
            else:
                month_data["local_volunteer_percentage_sessions_simulcast"] = 0.0
                month_data["poc_volunteer_percentage_sessions_simulcast"] = 0.0

            # Clean up tracking sets
            del month_data["_sessions_with_local_volunteers"]
            del month_data["_sessions_with_poc_volunteers"]
            del month_data["_simulcast_sessions_with_local_volunteers"]
            del month_data["_simulcast_sessions_with_poc_volunteers"]

        # Calculate year-to-date totals
        ytd_totals = {
            "successfully_completed": 0,
            "simulcast_sessions": 0,
            "teacher_canceled": 0,
            "teacher_no_shows": 0,
            "pathful_professional_canceled_no_shows": 0,
            "local_professional_canceled_no_shows": 0,
            "unfilled_sessions": 0,
            # New fields for volunteer diversity metrics - 4-column structure
            # Sessions Only (completed sessions)
            "local_volunteer_count_sessions_only": 0,
            "local_volunteer_percentage_sessions_only": 0.0,
            "poc_volunteer_count_sessions_only": 0,
            "poc_volunteer_percentage_sessions_only": 0.0,
            # Sessions + Simulcast (completed + simulcast sessions)
            "local_volunteer_count_sessions_simulcast": 0,
            "local_volunteer_percentage_sessions_simulcast": 0.0,
            "poc_volunteer_count_sessions_simulcast": 0,
            "poc_volunteer_percentage_sessions_simulcast": 0.0,
        }

        for month_data in monthly_breakdown.values():
            for key in ytd_totals:
                if key in month_data:
                    ytd_totals[key] += month_data[key]

        # Calculate YTD percentages for volunteer diversity metrics
        # Track sessions filled by local and POC volunteers across all months
        all_ytd_sessions_with_local = set()
        all_ytd_sessions_with_poc = set()
        all_ytd_simulcast_with_local = set()
        all_ytd_simulcast_with_poc = set()

        # Re-process events to collect YTD session data
        for event in events:
            if not event.start_date:
                continue

            # Only include events in our breakdown range
            month_key = f"{event.start_date.year}-{event.start_date.month:02d}"
            if month_key not in monthly_breakdown:
                continue

            # Only count successfully completed sessions
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()

            if "successfully completed" in original_status or (
                not original_status and event.status == EventStatus.COMPLETED
            ):
                event_id = event.id or event.title
                if event_id:
                    has_local_volunteer = False
                    has_poc_volunteer = False

                    for volunteer in event.volunteers or []:
                        # Check if volunteer is local
                        try:
                            from models.contact import LocalStatusEnum

                            if (
                                getattr(volunteer, "local_status", None)
                                == LocalStatusEnum.local
                            ):
                                has_local_volunteer = True
                        except Exception:
                            try:
                                if (
                                    str(getattr(volunteer, "local_status", ""))
                                    .lower()
                                    .endswith("local")
                                ):
                                    has_local_volunteer = True
                            except Exception:
                                pass

                        # Check if volunteer is People of Color
                        if getattr(volunteer, "is_people_of_color", False):
                            has_poc_volunteer = True

                    # Track for completed sessions (sessions only)
                    if has_local_volunteer:
                        all_ytd_sessions_with_local.add(event_id)
                    if has_poc_volunteer:
                        all_ytd_sessions_with_poc.add(event_id)

                    # Also track for sessions + simulcast (if this is a simulcast session)
                    if "simulcast" in original_status or (
                        not original_status and event.status == EventStatus.SIMULCAST
                    ):
                        if has_local_volunteer:
                            all_ytd_simulcast_with_local.add(event_id)
                        if has_poc_volunteer:
                            all_ytd_simulcast_with_poc.add(event_id)

        # Calculate YTD totals from session counts
        ytd_totals["local_volunteer_count_sessions_only"] = len(
            all_ytd_sessions_with_local
        )
        ytd_totals["poc_volunteer_count_sessions_only"] = len(all_ytd_sessions_with_poc)

        ytd_totals["local_volunteer_count_sessions_simulcast"] = len(
            all_ytd_sessions_with_local
        ) + len(all_ytd_simulcast_with_local)
        ytd_totals["poc_volunteer_count_sessions_simulcast"] = len(
            all_ytd_sessions_with_poc
        ) + len(all_ytd_simulcast_with_poc)

        # Calculate total YTD sessions for percentage calculation
        total_ytd_completed = ytd_totals["successfully_completed"]
        total_ytd_sessions_simulcast = (
            ytd_totals["successfully_completed"] + ytd_totals["simulcast_sessions"]
        )

        # Calculate YTD percentages for Sessions Only
        if total_ytd_completed > 0:
            ytd_totals["local_volunteer_percentage_sessions_only"] = round(
                (
                    ytd_totals["local_volunteer_count_sessions_only"]
                    / total_ytd_completed
                )
                * 100,
                1,
            )
            ytd_totals["poc_volunteer_percentage_sessions_only"] = round(
                (ytd_totals["poc_volunteer_count_sessions_only"] / total_ytd_completed)
                * 100,
                1,
            )
        else:
            ytd_totals["local_volunteer_percentage_sessions_only"] = 0.0
            ytd_totals["poc_volunteer_percentage_sessions_only"] = 0.0

        # Calculate YTD percentages for Sessions + Simulcast
        if total_ytd_sessions_simulcast > 0:
            ytd_totals["local_volunteer_percentage_sessions_simulcast"] = round(
                (
                    ytd_totals["local_volunteer_count_sessions_simulcast"]
                    / total_ytd_sessions_simulcast
                )
                * 100,
                1,
            )
            ytd_totals["poc_volunteer_percentage_sessions_simulcast"] = round(
                (
                    ytd_totals["poc_volunteer_count_sessions_simulcast"]
                    / total_ytd_sessions_simulcast
                )
                * 100,
                1,
            )
        else:
            ytd_totals["local_volunteer_percentage_sessions_simulcast"] = 0.0
            ytd_totals["poc_volunteer_percentage_sessions_simulcast"] = 0.0

        # COMPREHENSIVE ANALYTICS SECTIONS

        # 1. Current Running Count - detailed status breakdown
        running_count = {
            "unfilled": 0,
            "successfully_completed": 0,
            "teacher_cancelation": 0,
            "formerly_in_person_completed": 0,
            "formerly_in_person_canceled": 0,
            "professional_cancel_no_show": 0,
            "inclement_weather_cancellation": 0,
            "withdrawn_time_constraint": 0,
            "moved_to_in_person_session": 0,
            "teacher_no_show": 0,
            "covid19_cancelation": 0,
            "white_label_completed": 0,
            "white_label_unfilled": 0,
            "technical_difficulties": 0,
            "local_professional_cancellation": 0,
            "pathful_professional_cancellation": 0,
            "local_professional_no_show": 0,
            "pathful_professional_no_show": 0,
            "formerly_in_person_duplicate": 0,
            "white_label_canceled": 0,
            "simulcast": 0,
            "count": 0,
        }

        # 2. District-wise completed sessions
        district_completed = {}

        # 3. Topic-wise session counts
        topic_counts = {}

        # 4. Session type breakdown
        session_types = {
            "teacher_requested": {"all": 0, "completed": 0},
            "industry_chat": {"all": 0, "completed": 0},
            "formerly_in_person": {"all": 0, "completed": 0},
            "other": {"all": 0, "completed": 0},
            "kansas_city_series": {"all": 0, "completed": 0},
        }

        # 5. Grade level breakdown
        grade_levels = {
            "elementary": {"all": 0, "completed_no_simulcast": 0},
            "middle": {"all": 0, "completed_no_simulcast": 0},
            "high": {"all": 0, "completed_no_simulcast": 0},
            "grade_level_not_set": {"all": 0, "completed_no_simulcast": 0},
            "pre_k": {"all": 0, "completed_no_simulcast": 0},
        }

        # 6. Teacher attendance metrics
        teacher_attendance = {
            "more_than_one_teacher_present": 0,
            "total_number_of_sessions": len(events),
        }

        # Process events for comprehensive analytics
        for event in events:
            # Get status information
            original_status_raw = getattr(event, "original_status_string", None)
            original_status = (original_status_raw or "").lower().strip()
            event_status = event.status

            # Running count analysis - check multiple sources like monthly breakdown
            event_categorized = False

            # Check original status first
            if "successfully completed" in original_status:
                running_count["successfully_completed"] += 1
                event_categorized = True
            elif "simulcast" in original_status:
                running_count["simulcast"] += 1
                event_categorized = True
            elif "teacher" in original_status and (
                "cancelation" in original_status or "cancellation" in original_status
            ):
                running_count["teacher_cancelation"] += 1
                event_categorized = True
            elif (
                "teacher no-show" in original_status
                or "teacher no show" in original_status
            ):
                running_count["teacher_no_show"] += 1
                event_categorized = True
            elif "pathful professional cancellation" in original_status:
                running_count["pathful_professional_cancellation"] += 1
                event_categorized = True
            elif (
                "pathful professional no-show" in original_status
                or "pathful professional no show" in original_status
            ):
                running_count["pathful_professional_no_show"] += 1
                event_categorized = True
            elif "local professional cancellation" in original_status:
                running_count["local_professional_cancellation"] += 1
                event_categorized = True
            elif (
                "local professional no-show" in original_status
                or "local professional no show" in original_status
            ):
                running_count["local_professional_no_show"] += 1
                event_categorized = True
            elif "technical difficulties" in original_status:
                running_count["technical_difficulties"] += 1
                event_categorized = True
            elif "inclement weather" in original_status:
                running_count["inclement_weather_cancellation"] += 1
                event_categorized = True
            elif "moved to in-person" in original_status:
                running_count["moved_to_in_person_session"] += 1
                event_categorized = True
            elif (
                "white label completed" in original_status
                or "white lable completed" in original_status
            ):
                running_count["white_label_completed"] += 1
                event_categorized = True
            elif (
                "white label unfilled" in original_status
                or "white lable unfilled" in original_status
            ):
                running_count["white_label_unfilled"] += 1
                event_categorized = True
            elif (
                "formerly in-person" in original_status
                and "completed" in original_status
            ):
                running_count["formerly_in_person_completed"] += 1
                event_categorized = True
            elif "formerly in-person" in original_status and (
                "canceled" in original_status or "cancelled" in original_status
            ):
                running_count["formerly_in_person_canceled"] += 1
                event_categorized = True
            elif "covid" in original_status:
                running_count["covid19_cancelation"] += 1
                event_categorized = True
            elif original_status == "count":
                running_count["count"] += 1
                event_categorized = True

            # If not categorized by original status, check teacher registrations
            if not event_categorized:
                for teacher_reg in event.teacher_registrations:
                    tr_status = (teacher_reg.status or "").lower().strip()

                    # Check for professional statuses FIRST
                    if "pathful professional" in tr_status:
                        if "no-show" in tr_status or "no show" in tr_status:
                            running_count["pathful_professional_no_show"] += 1
                        else:
                            running_count["pathful_professional_cancellation"] += 1
                        event_categorized = True
                        break
                    elif "local professional" in tr_status:
                        if "no-show" in tr_status or "no show" in tr_status:
                            running_count["local_professional_no_show"] += 1
                        else:
                            running_count["local_professional_cancellation"] += 1
                        event_categorized = True
                        break
                    # Check for teacher-specific statuses
                    elif "cancel" in tr_status and "professional" not in tr_status:
                        running_count["teacher_cancelation"] += 1
                        event_categorized = True
                        break
                    elif (
                        "no-show" in tr_status or "no show" in tr_status
                    ) and "professional" not in tr_status:
                        running_count["teacher_no_show"] += 1
                        event_categorized = True
                        break

                # Check for simulcast from teacher registrations
                if not event_categorized:
                    simulcast_teacher_count = sum(
                        1 for tr in event.teacher_registrations if tr.is_simulcast
                    )
                    if simulcast_teacher_count > 0:
                        running_count["simulcast"] += 1
                        event_categorized = True

            # If still not categorized, check event status
            if not event_categorized:
                if event_status == EventStatus.COMPLETED:
                    running_count["successfully_completed"] += 1
                    event_categorized = True
                elif event_status == EventStatus.SIMULCAST:
                    running_count["simulcast"] += 1
                    event_categorized = True
                elif event_status == EventStatus.CANCELLED and not event.volunteers:
                    running_count["unfilled"] += 1
                    event_categorized = True
                elif event_status == EventStatus.NO_SHOW and not event.volunteers:
                    running_count["unfilled"] += 1
                    event_categorized = True

            # Default to unfilled if still not categorized
            if not event_categorized:
                running_count["unfilled"] += 1

            # District-wise completed sessions - check if this event is successfully completed
            event_is_completed = False

            # Check multiple sources for "successfully completed" status
            if "successfully completed" in original_status:
                event_is_completed = True
            elif event_status == EventStatus.COMPLETED:
                event_is_completed = True
            else:
                # Check teacher registrations for completion indicators
                for teacher_reg in event.teacher_registrations:
                    tr_status = (teacher_reg.status or "").lower().strip()
                    if (
                        "successfully completed" in tr_status
                        or "attended" in tr_status
                        or "completed" in tr_status
                    ):
                        event_is_completed = True
                        break

            # Count by district if successfully completed
            if event_is_completed and event.districts:
                for district in event.districts:
                    district_name = district.name
                    if district_name not in district_completed:
                        district_completed[district_name] = 0
                    district_completed[district_name] += 1

            # Topic-wise session counts (using series field which contains Topic/Theme)
            if event.series and event.series.strip():
                topic = event.series.strip()
                if topic not in topic_counts:
                    topic_counts[topic] = 0
                topic_counts[topic] += 1

            # Session type analysis (using additional_information field)
            session_type = event.additional_information or "other"
            session_type_clean = (
                session_type.lower().replace(" ", "_").replace("-", "_")
            )

            # Map to our predefined session types or use 'other'
            if session_type_clean in session_types:
                session_types[session_type_clean]["all"] += 1
                if "successfully completed" in original_status:
                    session_types[session_type_clean]["completed"] += 1
            else:
                session_types["other"]["all"] += 1
                if "successfully completed" in original_status:
                    session_types["other"]["completed"] += 1

            # Grade level analysis (using school level)
            grade_level = None
            if event.school_obj and event.school_obj.level:
                grade_level = event.school_obj.level.lower().strip()

            if grade_level:
                # Map school levels to our grade level categories
                if grade_level in ["elementary", "elem"]:
                    grade_key = "elementary"
                elif grade_level in ["middle", "mid"]:
                    grade_key = "middle"
                elif grade_level in ["high", "secondary"]:
                    grade_key = "high"
                elif grade_level in ["pre-k", "prek", "pre_k"]:
                    grade_key = "pre_k"
                else:
                    grade_key = "grade_level_not_set"

                grade_levels[grade_key]["all"] += 1
                # Count completed (no simulcast)
                if "successfully completed" in original_status:
                    grade_levels[grade_key]["completed_no_simulcast"] += 1
            else:
                grade_levels["grade_level_not_set"]["all"] += 1
                if "successfully completed" in original_status:
                    grade_levels["grade_level_not_set"]["completed_no_simulcast"] += 1

            # Teacher attendance analysis
            teacher_count = len(event.teacher_registrations)
            if teacher_count > 1:
                teacher_attendance["more_than_one_teacher_present"] += 1

        # Calculate professional cancel/no-show total for running count
        running_count["professional_cancel_no_show"] = (
            running_count["pathful_professional_cancellation"]
            + running_count["pathful_professional_no_show"]
            + running_count["local_professional_cancellation"]
            + running_count["local_professional_no_show"]
        )

        # Sort districts and topics for consistent display
        district_completed = dict(sorted(district_completed.items()))
        topic_counts = dict(sorted(topic_counts.items()))

        # Debug output to understand the data
        print(f"DEBUG: Found {len(events)} events")
        print(f"DEBUG: Unique statuses: {sorted(unique_statuses)}")
        print(f"DEBUG: Unique additional info: {sorted(unique_additional_info)}")
        print(f"DEBUG: Unique series: {sorted(unique_series)}")
        print(f"DEBUG: Unique school levels: {sorted(unique_school_levels)}")
        print(f"DEBUG: Running count totals: {running_count}")
        print(f"DEBUG: Topic counts: {topic_counts}")
        print(f"DEBUG: District completed: {district_completed}")

        # Show overall completion stats
        total_completed = sum(district_completed.values()) if district_completed else 0
        total_successfully_completed = running_count.get("successfully_completed", 0)
        print(f"DEBUG: Total completed by district: {total_completed}")
        print(
            f"DEBUG: Total successfully completed in running count: {total_successfully_completed}"
        )

        # Prepare data for template
        virtual_year_options = generate_school_year_options()

        return render_template(
            "virtual/usage/breakdown.html",
            monthly_breakdown=monthly_breakdown,
            ytd_totals=ytd_totals,
            running_count=running_count,
            district_completed=district_completed,
            topic_counts=topic_counts,
            session_types=session_types,
            grade_levels=grade_levels,
            teacher_attendance=teacher_attendance,
            current_filters=current_filters,
            virtual_year_options=virtual_year_options,
            months=months,
        )

    @virtual_bp.route("/usage/usage/district/<district_name>/teachers")
    @login_required
    @district_scoped_required
    def virtual_district_teacher_breakdown(district_name):
        """
        Show detailed teacher breakdown by school for a specific district.

        Args:
            district_name: Name of the district

        Returns:
            Rendered template with teachers grouped by school
        """
        # Get filter parameters
        default_virtual_year = get_current_virtual_year()
        selected_virtual_year = request.args.get("year", default_virtual_year)

        # Calculate default date range
        default_date_from, default_date_to = get_virtual_year_dates(
            selected_virtual_year
        )

        # Handle date parameters
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        date_from = default_date_from
        date_to = default_date_to

        if date_from_str:
            try:
                parsed_date_from = datetime.strptime(date_from_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_from.date()
                    <= default_date_to.date()
                ):
                    date_from = parsed_date_from.replace(hour=0, minute=0, second=0)
            except ValueError:
                pass

        if date_to_str:
            try:
                parsed_date_to = datetime.strptime(date_to_str, "%Y-%m-%d")
                if (
                    default_date_from.date()
                    <= parsed_date_to.date()
                    <= default_date_to.date()
                ):
                    date_to = parsed_date_to.replace(hour=23, minute=59, second=59)
            except ValueError:
                pass

        if date_from > date_to:
            date_from = default_date_from
            date_to = default_date_to

        current_filters = {
            "year": selected_virtual_year,
            "date_from": date_from,
            "date_to": date_to,
        }

        # Get detailed teacher breakdown data
        teacher_school_breakdown = compute_teacher_school_breakdown(
            district_name, selected_virtual_year, date_from, date_to
        )

        virtual_year_options = generate_school_year_options()

        return render_template(
            "virtual/teacher_progress/breakdown.html",
            district_name=district_name,
            teacher_school_breakdown=teacher_school_breakdown,
            current_filters=current_filters,
            virtual_year_options=virtual_year_options,
        )
