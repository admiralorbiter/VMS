"""
Events Routes Module
===================

This module provides all the route handlers for event management in the VMS system.
It includes functionality for listing, creating, editing, viewing, and importing events,
as well as syncing data with Salesforce.

Key Features:
- Event CRUD operations (Create, Read, Update, Delete)
- Advanced filtering and sorting of events
- Salesforce integration for data synchronization
- Volunteer and student participation tracking
- Skills management and association

Main Endpoints:
- /events: List all events with filtering and pagination
- /events/add: Create new events
- /events/view/<id>: View detailed event information
- /events/edit/<id>: Edit existing events
- /events/import-from-salesforce: Import events and participants from Salesforce
- /events/sync-student-participants: Sync student participation data from Salesforce

Dependencies:
- Salesforce API integration via simple_salesforce
- Event, Student, School, District, Skill, and Volunteer models
- Various utility functions from routes.utils
"""

import io
import json
from datetime import date, datetime, timedelta, timezone

import openpyxl
from flask import (
    Blueprint,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from simple_salesforce.api import Salesforce
from simple_salesforce.exceptions import SalesforceAuthenticationFailed
from sqlalchemy.orm import joinedload
from sqlalchemy.sql import func

from config import Config
from forms import EventForm
from models import db, eagerload_event_bundle
from models.district_model import District
from models.event import (
    Event,
    EventFormat,
    EventStatus,
    EventStudentParticipation,
    EventTeacher,
    EventType,
)
from models.school_model import School
from models.student import Student
from models.sync_log import SyncLog, SyncStatus
from models.teacher import Teacher
from models.volunteer import EventParticipation, Skill, Volunteer
from routes.decorators import global_users_only
from routes.utils import (
    DISTRICT_MAPPINGS,
    admin_required,
    log_audit_action,
    map_cancellation_reason,
    map_event_format,
    map_session_type,
    parse_date,
    parse_event_skills,
)
from services.scoping import can_edit_event, get_editable_fields, is_tenant_user
from utils.cache_refresh_scheduler import refresh_all_caches


def safe_parse_delivery_hours(value):
    """
    Safely parse delivery hours from Salesforce data

    Args:
        value: Raw value from Salesforce Delivery_Hours__c field

    Returns:
        float or None: Parsed hours value or None if invalid
    """
    if value is None:
        return None

    # Convert to string and strip whitespace
    if isinstance(value, str):
        value = value.strip()
        if not value:  # Empty string
            return None

    try:
        hours = float(value)
        # Return None for zero or negative values (they might indicate missing data)
        return hours if hours > 0 else None
    except (ValueError, TypeError):
        return None


# Create blueprint for events functionality
events_bp = Blueprint("events", __name__)


def process_event_row(row, success_count, error_count, errors, skipped_count):
    """
    Process a single event row from Salesforce data

    This function handles the processing of individual event records from
    Salesforce API responses. It validates required fields, creates or updates
    Event objects, and handles relationships with districts, schools, and skills.

    Args:
        row (dict): Event data from Salesforce
        success_count (int): Running count of successful operations
        error_count (int): Running count of failed operations
        errors (list): List to collect error messages
        skipped_count (int): Running count of skipped events (already exist)

    Returns:
        tuple: Updated (success_count, error_count, skipped_count)
    """
    try:
        # Get initial data to check
        parent_account = row.get("Parent_Account__c")
        district_name = row.get("District__c")
        school_id = row.get("School__c")
        event_name = row.get("Name", "").strip()
        session_type = row.get("Session_Type__c", "")

        # Only skip events that are completely unusable
        if not event_name:
            event_id = row.get("Id", "unknown")
            errors.append(
                f"Skipping event {event_id}: Missing event name/title (required field)"
            )
            return success_count, error_count + 1, skipped_count

        # Get or create event
        event_id = row.get("Id")
        event = Event.query.filter_by(salesforce_id=event_id).first()
        is_new = event is None

        if not event:
            event = Event(
                title=event_name,
                salesforce_id=event_id,
            )
            db.session.add(event)
        else:
            # Event already exists - update it instead of skipping
            pass

        # Update event fields (handles both new and existing events)
        event.title = event_name

        # Update basic event fields
        event.type = map_session_type(session_type)
        event.format = map_event_format(row.get("Format__c", ""))
        event.start_date = parse_date(row.get("Start_Date_and_Time__c")) or datetime(
            2000, 1, 1
        )
        event.end_date = parse_date(row.get("End_Date_and_Time__c")) or datetime(
            2000, 1, 1
        )
        event.status = row.get("Session_Status__c", "Draft")
        event.location = row.get("Location_Information__c", "")
        event.description = row.get("Description__c", "")
        event.cancellation_reason = map_cancellation_reason(
            row.get("Cancellation_Reason__c")
        )
        event.participant_count = int(
            float(row.get("Non_Scheduled_Students_Count__c", 0))
            if row.get("Non_Scheduled_Students_Count__c") is not None
            else 0
        )
        event.additional_information = row.get("Additional_Information__c", "")
        event.session_host = row.get("Session_Host__c", "")

        # Handle numeric fields - FIX for negative available_slots
        def safe_convert_to_int(value, default=0):
            if value is None:
                return default
            try:
                converted = int(float(value))
                # Convert negative values to 0 to avoid constraint violations
                return max(0, converted)
            except (ValueError, TypeError):
                return default

        event.total_requested_volunteer_jobs = safe_convert_to_int(
            row.get("Total_Requested_Volunteer_Jobs__c")
        )
        event.available_slots = safe_convert_to_int(row.get("Available_Slots__c"))

        # Handle School relationship (using Salesforce ID)
        school_district = None
        if school_id:
            school = School.query.get(school_id)
            if school:
                event.school = school_id  # Store the Salesforce ID
                school_district = school.district

        # Handle District relationship
        # If District__c is empty, use Parent_Account__c instead
        if not district_name and parent_account:
            district_name = parent_account

        # Try to get the district
        district = None
        if district_name and district_name in DISTRICT_MAPPINGS:
            mapped_name = DISTRICT_MAPPINGS[district_name]
            district = District.query.filter_by(name=mapped_name).first()

        # Clear and reassign districts if we have a valid district to assign
        if district or school_district:
            event.districts = []  # Clear existing districts
            if school_district:
                event.districts.append(school_district)
            if district and district not in event.districts:
                event.districts.append(district)

        # Store original district name for reference
        event.district_partner = district_name if district_name else None

        # Handle skills
        skills_covered = parse_event_skills(
            row.get("Legacy_Skill_Covered_for_the_Session__c", "")
        )
        skills_needed = parse_event_skills(row.get("Legacy_Skills_Needed__c", ""))
        requested_skills = parse_event_skills(row.get("Requested_Skills__c", ""))

        # Combine all skills and remove duplicates
        all_skills = set(skills_covered + skills_needed + requested_skills)

        # Get existing skills to avoid duplicates
        existing_skill_names = {skill.name for skill in event.skills}

        # Only process new skills
        new_skill_names = all_skills - existing_skill_names
        for skill_name in new_skill_names:
            skill = Skill.query.filter_by(name=skill_name).first()
            if not skill:
                skill = Skill(name=skill_name)
                db.session.add(skill)
            if skill not in event.skills:
                event.skills.append(skill)

        db.session.flush()  # Flush to catch validation errors early

        # Count as processed (both new and existing)
        return success_count + 1, error_count, skipped_count

    except Exception as e:
        db.session.rollback()
        errors.append(f"Error processing event {row.get('Id', 'unknown')}: {str(e)}")
        return success_count, error_count + 1, skipped_count


def process_participation_row(row, success_count, error_count, errors):
    """
    Process a single participation row from Salesforce data

    This function handles volunteer participation records from Salesforce
    sync operations. It creates EventParticipation records linking
    volunteers to events with status and delivery hours.

    Args:
        row (dict): Participation data from Salesforce
        success_count (int): Running count of successful operations
        error_count (int): Running count of failed operations
        errors (list): List to collect error messages

    Returns:
        tuple: Updated (success_count, error_count)
    """
    try:
        # Check if participation already exists
        existing = EventParticipation.query.filter_by(salesforce_id=row["Id"]).first()
        if existing:
            # Update existing record with current Salesforce data
            existing.status = row["Status__c"]
            existing.delivery_hours = safe_parse_delivery_hours(
                row.get("Delivery_Hours__c")
            )
            # Update other fields that might have changed
            if row.get("Email__c"):
                existing.email = row["Email__c"]
            if row.get("Title__c"):
                existing.title = row["Title__c"]
            if row.get("Age_Group__c"):
                existing.age_group = row["Age_Group__c"]
            # print(f"Successfully updated participation: {row['Id']}")  # Debug log
            return success_count + 1, error_count

        # Find the volunteer and event by their Salesforce IDs
        volunteer = Volunteer.query.filter_by(
            salesforce_individual_id=row["Contact__c"]
        ).first()
        event = Event.query.filter_by(salesforce_id=row["Session__c"]).first()

        if not volunteer or not event:
            error_msg = (
                f"Could not find volunteer or event for participation {row['Id']}"
            )
            errors.append(error_msg)
            return success_count, error_count + 1

        # Create new participation record
        participation = EventParticipation(
            volunteer_id=volunteer.id,
            event_id=event.id,
            status=row["Status__c"],
            delivery_hours=safe_parse_delivery_hours(row.get("Delivery_Hours__c")),
            salesforce_id=row["Id"],
        )

        db.session.add(participation)
        # print(f"Successfully added participation: {row['Id']}")  # Debug log
        return success_count + 1, error_count

    except Exception as e:
        error_msg = f"Error processing participation row: {str(e)}"
        print(error_msg)  # Debug log
        db.session.rollback()
        errors.append(error_msg)
        return success_count, error_count + 1


def process_student_participation_row(row, success_count, error_count, errors):
    """
    Process a single student participation row from Salesforce data

    This function handles student participation records from Salesforce sync
    operations. It creates EventStudentParticipation records linking students
    to events with status, delivery hours, and age group information.

    Args:
        row (dict): Student participation data from Salesforce
        success_count (int): Running count of successful operations
        error_count (int): Running count of failed operations
        errors (list): List to collect error messages

    Returns:
        tuple: Updated (success_count, error_count)
    """
    try:
        event_sf_id = row.get("Session__c")
        student_contact_sf_id = row.get(
            "Contact__c"
        )  # Salesforce Contact ID for the student
        participation_sf_id = row.get("Id")  # Salesforce Session_Participant__c ID
        status = row.get("Status__c")
        delivery_hours_str = row.get("Delivery_Hours__c")
        age_group = row.get("Age_Group__c")

        # Validate required Salesforce IDs
        if not all([event_sf_id, student_contact_sf_id, participation_sf_id]):
            errors.append(
                f"Skipping student participation {participation_sf_id or 'unknown'}: Missing required Salesforce IDs (Session__c, Contact__c, Id)"
            )
            return success_count, error_count + 1

        # --- Find Event ---
        event = Event.query.filter_by(salesforce_id=event_sf_id).first()
        if not event:
            errors.append(
                f"Event with Salesforce ID {event_sf_id} not found for participation {participation_sf_id}"
            )
            return success_count, error_count + 1

        # --- Find Student ---
        # Student inherits from Contact, so we query Student via Contact's salesforce_individual_id
        student = Student.query.filter(
            Student.salesforce_individual_id == student_contact_sf_id
        ).first()
        if not student:
            # If a student record doesn't exist in your DB for the given SF Contact ID,
            # we log an error. You could enhance this later to create the student if needed.
            errors.append(
                f"Student with Salesforce Contact ID {student_contact_sf_id} not found for participation {participation_sf_id}"
            )
            return success_count, error_count + 1

        # --- Check for existing participation record ---
        existing_participation = EventStudentParticipation.query.filter_by(
            salesforce_id=participation_sf_id
        ).first()

        if existing_participation:
            # Skip if this specific Salesforce participation record already exists in the DB
            # Optionally, you could add logic here to update the existing record if needed.
            return success_count, error_count
        else:
            # Before creating, guard against duplicate event/student pair
            pair_participation = EventStudentParticipation.query.filter_by(
                event_id=event.id, student_id=student.id
            ).first()

            delivery_hours = safe_parse_delivery_hours(delivery_hours_str)

            if pair_participation:
                # Update existing pair record if needed and attach Salesforce ID
                if not pair_participation.salesforce_id:
                    pair_participation.salesforce_id = participation_sf_id
                # Optionally refresh fields if provided
                if status:
                    pair_participation.status = status
                if delivery_hours is not None:
                    pair_participation.delivery_hours = delivery_hours
                if age_group:
                    pair_participation.age_group = age_group
                return success_count + 1, error_count
            else:
                # --- Create New Participation Record ---
                new_participation = EventStudentParticipation(
                    event_id=event.id,
                    student_id=student.id,  # Use the primary key of the found student
                    status=status,
                    delivery_hours=delivery_hours,
                    age_group=age_group,
                    salesforce_id=participation_sf_id,  # Store the unique SF participation ID
                )
                db.session.add(new_participation)
                db.session.commit()  # Commit the new participation record
                return success_count + 1, error_count

    except Exception as e:
        db.session.rollback()  # Rollback the transaction for this row on error
        errors.append(
            f"Error processing student participation {row.get('Id', 'unknown')}: {str(e)}"
        )
        # Consider logging the full traceback for debugging complex errors
        # import traceback
        # print(traceback.format_exc())
        return success_count, error_count + 1


def fix_missing_participation_records(event):
    """
    Fix missing EventParticipation records for volunteers linked to an event.

    This function ensures that all volunteers linked to an event through the
    many-to-many relationship have proper EventParticipation records with
    appropriate delivery hours.

    Args:
        event: Event object to fix participation records for
    """
    from models.volunteer import EventParticipation

    # Get all volunteers linked to this event
    event_volunteers = event.volunteers

    for volunteer in event_volunteers:
        # Check if participation record exists
        participation = EventParticipation.query.filter_by(
            event_id=event.id, volunteer_id=volunteer.id
        ).first()

        if not participation:
            # Create missing participation record
            # Calculate delivery hours based on event duration
            delivery_hours = None
            if event.duration:
                delivery_hours = event.duration / 60  # Convert minutes to hours
            elif event.start_date and event.end_date:
                # Calculate from start/end times
                duration_minutes = (
                    event.end_date - event.start_date
                ).total_seconds() / 60
                delivery_hours = max(1.0, duration_minutes / 60)  # Minimum 1 hour
            else:
                delivery_hours = 0.0  # Default to 0 hours if no timing data

            # Determine status based on event status
            status = "Attended"
            if event.status == EventStatus.COMPLETED:
                status = "Completed"
            elif event.status == EventStatus.CANCELLED:
                status = "Cancelled"
            elif event.status == EventStatus.NO_SHOW:
                status = "No Show"

            # Create new participation record
            participation = EventParticipation(
                volunteer_id=volunteer.id,
                event_id=event.id,
                status=status,
                delivery_hours=delivery_hours,
                participant_type="Volunteer",
            )
            db.session.add(participation)
            print(
                f"Created missing participation record for volunteer {volunteer.first_name} {volunteer.last_name} in event {event.title}"
            )

        elif participation.delivery_hours is None:
            # Fix existing participation record with missing delivery hours
            if event.duration:
                participation.delivery_hours = event.duration / 60
            elif event.start_date and event.end_date:
                duration_minutes = (
                    event.end_date - event.start_date
                ).total_seconds() / 60
                participation.delivery_hours = max(1.0, duration_minutes / 60)
            else:
                participation.delivery_hours = 0.0
            print(
                f"Fixed delivery hours for volunteer {volunteer.first_name} {volunteer.last_name} in event {event.title}"
            )

    # Commit changes
    try:
        db.session.commit()
        print(f"Successfully fixed participation records for event: {event.title}")
    except Exception as e:
        db.session.rollback()
        print(f"Error fixing participation records: {str(e)}")


@events_bp.route("/events")
@login_required
@global_users_only
def events():
    """
    List all events with filtering, sorting, and pagination
    Now supports filtering by academic year (8/1 to 7/31).
    """
    # Get pagination parameters
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 25, type=int)
    allowed_per_page = [10, 25, 50, 100]
    if per_page not in allowed_per_page:
        per_page = 25

    # Academic year calculation
    today = date.today()
    if today.month >= 8:
        default_academic_year = f"{today.year}-{today.year+1}"
    else:
        default_academic_year = f"{today.year-1}-{today.year}"
    academic_year = request.args.get("academic_year", default_academic_year)

    # Build list of available academic years from events
    min_event = Event.query.order_by(Event.start_date.asc()).first()
    max_event = Event.query.order_by(Event.start_date.desc()).first()
    academic_years = []
    if min_event and max_event:
        min_year = (
            min_event.start_date.year
            if min_event.start_date.month >= 8
            else min_event.start_date.year - 1
        )
        max_year = (
            max_event.start_date.year
            if max_event.start_date.month >= 8
            else max_event.start_date.year - 1
        )
        for y in range(max_year, min_year - 1, -1):
            academic_years.append(f"{y}-{y+1}")
    academic_years.insert(0, "All")

    # Create current_filters dictionary
    current_filters = {
        "search_title": request.args.get("search_title", "").strip(),
        "event_type": request.args.get("event_type", ""),
        "status": request.args.get("status", ""),
        "start_date": request.args.get("start_date", ""),
        "end_date": request.args.get("end_date", ""),
        "per_page": per_page,
        "academic_year": academic_year,
    }
    current_filters = {
        k: v for k, v in current_filters.items() if v or k == "academic_year"
    }

    # Get sort parameters
    sort_by = request.args.get("sort_by", "start_date")
    sort_direction = request.args.get("sort_direction", "desc")
    current_filters["sort_by"] = sort_by
    current_filters["sort_direction"] = sort_direction

    # Build query
    query = eagerload_event_bundle(Event.query)

    # Apply filters
    if current_filters.get("search_title"):
        search_term = f"%{current_filters['search_title']}%"
        query = query.filter(Event.title.ilike(search_term))
    if current_filters.get("event_type"):
        try:
            event_type = EventType[current_filters["event_type"].upper()]
            query = query.filter(Event.type == event_type)
        except KeyError:
            flash(f"Invalid event type: {current_filters['event_type']}", "warning")
    if current_filters.get("status"):
        status = current_filters["status"]
        query = query.filter(Event.status == status)
    if current_filters.get("start_date"):
        try:
            start_date = datetime.strptime(current_filters["start_date"], "%Y-%m-%d")
            query = query.filter(Event.start_date >= start_date)
        except ValueError:
            flash("Invalid start date format", "warning")
    if current_filters.get("end_date"):
        try:
            end_date = datetime.strptime(current_filters["end_date"], "%Y-%m-%d")
            end_date = end_date + timedelta(days=1) - timedelta(seconds=1)
            query = query.filter(Event.start_date <= end_date)
        except ValueError:
            flash("Invalid end date format", "warning")
    # Academic year filter
    if academic_year and academic_year != "All":
        start_year = int(academic_year.split("-")[0])
        start_dt = datetime(start_year, 8, 1)
        end_dt = datetime(start_year + 1, 7, 31, 23, 59, 59)
        query = query.filter(Event.start_date >= start_dt, Event.start_date <= end_dt)

    # Apply sorting
    sort_column = getattr(Event, sort_by, Event.start_date)
    if sort_direction == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    # Apply pagination
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    event_types = [
        (t.name.lower(), t.name.replace("_", " ").title()) for t in EventType
    ]
    statuses = [
        ("Completed", "Completed"),
        ("Confirmed", "Confirmed"),
        ("Cancelled", "Cancelled"),
        ("Requested", "Requested"),
        ("Draft", "Draft"),
        ("Published", "Published"),
    ]

    return render_template(
        "events/events.html",
        events=pagination.items,
        pagination=pagination,
        current_filters=current_filters,
        event_types=event_types,
        statuses=statuses,
        academic_years=academic_years,
    )


@events_bp.route("/events/add", methods=["GET", "POST"])
@login_required
@global_users_only
def add_event():
    """
    Create a new event

    GET: Display the event creation form
    POST: Process form submission and create new event

    Returns:
        GET: Rendered form template
        POST: Redirect to event view page on success
    """
    form = EventForm()

    if form.validate_on_submit():
        try:
            print("Form validation successful")  # Debug
            print(f"Form data: {form.data}")  # Debug

            # Create new event from form data
            event = Event(
                title=form.title.data,
                type=EventType(form.type.data),
                format=EventFormat(form.format.data),
                start_date=form.start_date.data,
                end_date=form.end_date.data,
                location=form.location.data or "",
                status=EventStatus(form.status.data),
                description=form.description.data or "",
                volunteers_needed=form.volunteers_needed.data or 0,
            )

            print(f"Created event: {event}")  # Debug

            db.session.add(event)
            db.session.commit()

            flash("Event added successfully!", "success")
            return redirect(url_for("events.view_event", id=event.id))

        except Exception as e:
            db.session.rollback()
            print(f"Error adding event: {str(e)}")  # Debug
            flash(f"Error adding event: {str(e)}", "danger")
    else:
        # Print form validation errors for debugging
        if request.method == "POST":
            print(f"Form validation failed: {form.errors}")  # Debug
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{field}: {error}", "danger")

    return render_template("events/add_event.html", form=form)


@events_bp.route("/events/view/<int:id>")
@login_required
@global_users_only
def view_event(id):
    """
    View detailed information about a specific event

    This endpoint displays comprehensive event information including
    participation details, skills, and related data.

    Args:
        id (int): Event ID to view

    Returns:
        Rendered template with event details

    Raises:
        404: If event not found
    """
    event = db.session.get(Event, id)
    if not event:
        abort(404)

    # Fix missing participation records and null delivery hours before displaying
    fix_missing_participation_records(event)

    # Also fix any remaining null delivery hours
    null_hours_participations = EventParticipation.query.filter(
        EventParticipation.event_id == event.id,
        EventParticipation.delivery_hours.is_(None),
    ).all()

    if null_hours_participations:
        # Calculate delivery hours based on event data
        delivery_hours = None
        if event.duration:
            delivery_hours = event.duration / 60  # Convert minutes to hours
        elif event.start_date and event.end_date:
            # Calculate from start/end times
            duration_minutes = (event.end_date - event.start_date).total_seconds() / 60
            delivery_hours = max(1.0, duration_minutes / 60)  # Minimum 1 hour
        else:
            delivery_hours = 0.0  # Default to 0 hours if no timing data

        # Update all null delivery hours for this event
        for participation in null_hours_participations:
            participation.delivery_hours = delivery_hours

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Error fixing delivery hours: {str(e)}")

    # Set default dates if None
    if event.start_date is None:
        event.start_date = datetime.now()
    if event.end_date is None:
        event.end_date = datetime.now()

    # Get volunteer participations with volunteers
    # Use joinedload to efficiently load volunteers
    volunteer_participations = (
        EventParticipation.query.options(joinedload(EventParticipation.volunteer))
        .filter_by(event_id=id)
        .all()
    )

    # Calculate unique volunteer count from participations
    unique_volunteers = set()
    for participation in volunteer_participations:
        if participation.volunteer:
            unique_volunteers.add(participation.volunteer.id)
    volunteer_count = len(unique_volunteers)

    # Get all event teachers with their statuses
    # Use joinedload to efficiently load teachers
    event_teachers = (
        EventTeacher.query.options(
            joinedload(EventTeacher.teacher).joinedload(Teacher.school)
        )
        .filter_by(event_id=id)
        .all()
    )

    # --- NEW: Get student participations with students ---
    # Use joinedload to efficiently load students
    student_participations = (
        EventStudentParticipation.query.options(
            joinedload(EventStudentParticipation.student)
        )
        .filter_by(event_id=id)
        .all()
    )
    # --- End NEW ---

    # Group volunteer participations by status
    participation_stats = {
        "Registered": [],
        "Attended": [],
        "No Show": [],
        "Cancelled": [],
    }

    for participation in volunteer_participations:
        # Ensure volunteer exists before processing
        if participation.volunteer:
            status = participation.status
            if status in participation_stats:
                participation_stats[status].append(
                    {
                        "volunteer": participation.volunteer,
                        "delivery_hours": participation.delivery_hours,
                    }
                )

    # Calculate unique volunteer count for 'Attended' status only
    attended_volunteers = set()
    for p in participation_stats.get("Attended", []):
        if p["volunteer"]:
            attended_volunteers.add(p["volunteer"].id)
    volunteer_count = len(attended_volunteers)

    # Calculate estimated students for virtual events
    estimated_students = 0
    if (
        getattr(event.format, "value", None) == "virtual"
        or getattr(event.type, "value", None) == "virtual_session"
    ):
        for t in event_teachers:
            # Accept both boolean and int for is_simulcast
            is_simulcast = t.is_simulcast == True or t.is_simulcast == 1
            status = (t.status or "").strip().lower()
            if is_simulcast or status in [
                "simulcast",
                "successfully completed",
                "attended",
                "completed",
            ]:
                estimated_students += 1
        estimated_students *= 20
    else:
        estimated_students = event.participant_count or 0

    return render_template(
        "events/view.html",
        event=event,
        volunteer_count=volunteer_count,  # Now matches 'Attended' volunteers
        participation_stats=participation_stats,
        event_teachers=event_teachers,
        student_participations=student_participations,
        estimated_students=estimated_students,
    )


@events_bp.route("/events/edit/<int:id>", methods=["GET", "POST"])
@login_required
def edit_event(id):
    """
    Edit an existing event

    GET: Display the event edit form pre-populated with current data
    POST: Process form submission and update event

    Args:
        id (int): Event ID to edit

    Returns:
        GET: Rendered form template
        POST: Redirect to event view page on success

    Raises:
        404: If event not found
        403: If user doesn't have permission to edit this event

    Phase D-3: District Admin Access (DEC-009)
    - Staff/admin users can edit all events (all fields)
    - Tenant admins can edit their district's events (restricted fields)
    """
    from models.event import CancellationReason

    event = db.session.get(Event, id)
    if not event:
        abort(404)

    # Phase D-3: Check if user can edit this event
    if not can_edit_event(current_user, event):
        flash("You do not have permission to edit this event.", "error")
        abort(403)

    # Get list of editable fields for this user/event
    editable_fields = get_editable_fields(current_user, event)

    form = EventForm()

    if form.validate_on_submit():
        try:
            print("Edit form validation successful")  # Debug
            print(f"Form data: {form.data}")  # Debug

            # Update event from form data
            form.populate_obj(event)

            # Handle enum conversions
            event.type = EventType(form.type.data)
            event.format = EventFormat(form.format.data)
            event.status = EventStatus(form.status.data)

            # Handle cancellation reason (DEC-008)
            if event.status == EventStatus.CANCELLED:
                reason_value = form.cancellation_reason.data
                notes_value = form.cancellation_notes.data

                if reason_value:
                    try:
                        event.set_cancellation_reason(
                            reason=reason_value,
                            notes=notes_value,
                            user_id=current_user.id,
                        )
                    except ValueError as e:
                        flash(str(e), "warning")
            else:
                # Clear cancellation fields if status is not CANCELLED
                event.cancellation_reason = None
                event.cancellation_notes = None
                event.cancellation_set_by = None
                event.cancellation_set_at = None

            print(f"Updated event: {event}")  # Debug

            db.session.commit()
            flash("Event updated successfully!", "success")
            return redirect(url_for("events.view_event", id=event.id))

        except Exception as e:
            db.session.rollback()
            print(f"Error updating event: {str(e)}")  # Debug
            flash(f"Error updating event: {str(e)}", "danger")
    else:
        # Pre-populate form with event data for GET requests
        if request.method == "GET":
            form.title.data = event.title
            form.type.data = event.type.value if event.type else ""
            form.format.data = event.format.value if event.format else ""
            form.start_date.data = event.start_date
            form.end_date.data = event.end_date
            form.location.data = event.location
            form.status.data = event.status.value if event.status else ""
            form.description.data = event.description
            form.volunteers_needed.data = event.volunteers_needed
            # Pre-populate cancellation reason fields
            form.cancellation_reason.data = (
                event.cancellation_reason.value if event.cancellation_reason else ""
            )
            form.cancellation_notes.data = event.cancellation_notes or ""

        # Print form validation errors for debugging
        if request.method == "POST":
            print(f"Form validation failed: {form.errors}")  # Debug
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{field}: {error}", "danger")

    return render_template(
        "events/edit.html",
        event=event,
        form=form,
        # Phase D-3: Field-level restrictions for tenant users
        editable_fields=editable_fields,
        is_tenant_user=is_tenant_user(current_user),
    )


@events_bp.route("/events/purge", methods=["POST"])
@login_required
@admin_required
def purge_events():
    """
    Purge all events and related data (Admin only)

    This endpoint completely removes all events and related data from the database.
    It should only be used for testing or complete data reset scenarios.

    Returns:
        JSON response indicating success or failure
    """
    if not current_user.is_admin:
        return jsonify({"error": "Unauthorized"}), 403

    try:
        # First delete all event participations
        EventParticipation.query.delete()

        # Delete student participations
        EventStudentParticipation.query.delete()

        # Delete event-teacher associations
        EventTeacher.query.delete()

        # Delete event-district associations
        db.session.execute(db.text("DELETE FROM event_districts"))

        # Delete event-skill associations
        db.session.execute(db.text("DELETE FROM event_skills"))

        # Then delete all events
        Event.query.delete()

        db.session.commit()
        log_audit_action(action="purge", resource_type="event")
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)})


@events_bp.route("/events/delete/<int:id>", methods=["DELETE"])
@login_required
@admin_required
def delete_event(id):
    """
    Delete an event (Admin only)

    This endpoint removes an event and all its associated data from the database.

    Args:
        id (int): Event ID to delete

    Returns:
        JSON response indicating success or failure

    Raises:
        404: If event not found
    """
    if not current_user.is_admin:
        return jsonify({"error": "Unauthorized"}), 403

    try:
        event = db.session.get(Event, id)
        if not event:
            abort(404)

        # First delete all participations
        EventParticipation.query.filter_by(event_id=id).delete()

        # Then delete the event
        db.session.delete(event)
        db.session.commit()
        log_audit_action(action="delete", resource_type="event", resource_id=id)

        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@events_bp.route("/api/skills/find-or-create", methods=["POST"])
@login_required
@global_users_only
def find_or_create_skill():
    """
    Find or create a skill by name

    This endpoint provides an API for finding existing skills or creating
    new ones. It's typically used by the frontend for dynamic skill management.

    Returns:
        JSON response with skill information
    """
    try:
        data = request.get_json()
        skill_name = data.get("name").strip()

        # Look for existing skill
        skill = Skill.query.filter(
            func.lower(Skill.name) == func.lower(skill_name)
        ).first()

        # Create new skill if it doesn't exist
        if not skill:
            skill = Skill(name=skill_name)
            db.session.add(skill)
            db.session.commit()

        return jsonify({"success": True, "skill": {"id": skill.id, "name": skill.name}})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@events_bp.route("/events/import-from-salesforce", methods=["POST"])
@login_required
@global_users_only
def import_events_from_salesforce():
    """
    Import events and participants from Salesforce

    This endpoint performs a comprehensive import of both events and
    participant data from Salesforce. It handles authentication, data
    processing, and error reporting.

    Returns:
        JSON response with import results and statistics
    """
    try:
        started_at = datetime.now(timezone.utc)
        print("Fetching data from Salesforce...")
        success_count = 0
        error_count = 0
        errors = []
        skipped_count = 0

        # Connect to Salesforce
        sf = Salesforce(
            username=Config.SF_USERNAME,
            password=Config.SF_PASSWORD,
            security_token=Config.SF_SECURITY_TOKEN,
            domain="login",
        )

        # First query: Get events
        events_query = """
        SELECT Id, Name, Session_Type__c, Format__c, Start_Date_and_Time__c,
            End_Date_and_Time__c, Session_Status__c, Location_Information__c,
            Description__c, Cancellation_Reason__c, Non_Scheduled_Students_Count__c,
            District__c, School__c, Legacy_Skill_Covered_for_the_Session__c,
            Legacy_Skills_Needed__c, Requested_Skills__c, Additional_Information__c,
            Total_Requested_Volunteer_Jobs__c, Available_Slots__c, Parent_Account__c,
            Session_Host__c
        FROM Session__c
        WHERE Session_Status__c != 'Draft' AND Session_Type__c != 'Connector Session'
        ORDER BY Start_Date_and_Time__c DESC
        """

        # Execute events query
        events_result = sf.query_all(events_query)
        events_rows = events_result.get("records", [])
        total_events = len(events_rows)

        print(f"Found {total_events} events in Salesforce")
        print(
            f"Query filter: Session_Status__c != 'Draft' AND Session_Type__c != 'Connector Session'"
        )

        status_counts = {}
        type_counts = {}

        # Process events
        for i, row in enumerate(events_rows):
            # Progress indicator every 500 events
            if i > 0 and i % 500 == 0:
                print(f"Progress: {i}/{total_events} events processed")

            status = row.get("Session_Status__c", "Unknown")
            session_type = row.get("Session_Type__c", "Unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
            type_counts[session_type] = type_counts.get(session_type, 0) + 1

            success_count, error_count, skipped_count = process_event_row(
                row, success_count, error_count, errors, skipped_count
            )

        # Compact status summary
        print(f"\n{'='*60}")
        print(f"EVENTS IMPORT SUMMARY")
        print(f"{'='*60}")
        print(f"Total from Salesforce: {total_events}")
        print(f"Successfully processed: {success_count}")
        print(f"Errors: {error_count}")
        print(f"Skipped (invalid): {skipped_count}")

        if success_count + error_count + skipped_count != total_events:
            print(
                f"⚠️  COUNT MISMATCH: {success_count} + {error_count} + {skipped_count} != {total_events}"
            )
            print(
                f"Missing: {total_events - (success_count + error_count + skipped_count)} events"
            )

        print(f"\nStatus breakdown:")
        for status, count in sorted(status_counts.items()):
            print(f"  {status}: {count}")

        print(f"\nSession type breakdown:")
        for session_type, count in sorted(type_counts.items()):
            print(f"  {session_type}: {count}")

        # Quick validation check
        if (
            type_counts.get("DIA", 0) + type_counts.get("DIA - Classroom Speaker", 0)
            > 0
        ):
            dia_events_in_db = Event.query.filter(
                Event.type.in_([EventType.DIA, EventType.DIA_CLASSROOM_SPEAKER])
            ).count()
            if dia_events_in_db == 0:
                print(
                    f"⚠️  WARNING: {type_counts.get('DIA', 0) + type_counts.get('DIA - Classroom Speaker', 0)} DIA events in Salesforce but 0 in database"
                )

        # Second query: Get participants
        participants_query = """
        SELECT
            Id,
            Name,
            Contact__c,
            Session__c,
            Status__c,
            Delivery_Hours__c,
            Age_Group__c,
            Email__c,
            Title__c
        FROM Session_Participant__c
        WHERE Participant_Type__c = 'Volunteer'
        """

        # Execute participants query
        participants_result = sf.query_all(participants_query)
        participant_rows = participants_result.get("records", [])

        # Process participants
        participant_success = 0
        participant_error = 0
        for row in participant_rows:
            participant_success, participant_error = process_participation_row(
                row, participant_success, participant_error, errors
            )

        # Commit all changes
        db.session.commit()

        # Final summary
        print(f"\n{'='*60}")
        print(f"FINAL IMPORT SUMMARY")
        print(f"{'='*60}")
        print(f"Events: {success_count}/{total_events} processed successfully")
        print(
            f"Participants: {participant_success}/{len(participant_rows)} processed successfully"
        )
        print(f"Total errors: {error_count + participant_error}")
        print(f"{'='*60}")

        # Separate actual errors from expected skips
        actual_errors = [e for e in errors if "Error processing" in e]
        skipped_events = [e for e in errors if "Skipping event" in e]

        if skipped_events:
            print(f"\nSkipped events (missing required data): {len(skipped_events)}")
            for skip in skipped_events[:5]:  # Show first 5 skips
                print(f"  - {skip}")
            if len(skipped_events) > 5:
                print(f"  ... and {len(skipped_events) - 5} more skipped events")

        if actual_errors:
            print(f"\nProcessing errors: {len(actual_errors)}")
            for error in actual_errors[:5]:  # Show first 5 errors
                print(f"  - {error}")
            if len(actual_errors) > 5:
                print(f"  ... and {len(actual_errors) - 5} more errors")

        # Trigger cache refresh if any data was processed
        if success_count > 0 or participant_success > 0:
            print(f"\nData updated. Triggering full cache refresh...")
            try:
                refresh_all_caches()
                print("Cache refresh triggered successfully.")
            except Exception as e:
                print(f"Warning: Cache refresh failed: {e}")

        # Record sync log
        try:
            total_records = total_events + len(participant_rows)
            # Determine status based on errors
            sync_status = SyncStatus.SUCCESS.value
            if error_count + participant_error > 0:
                if success_count + participant_success > 0:
                    sync_status = SyncStatus.PARTIAL.value
                else:
                    sync_status = SyncStatus.FAILED.value

            sync_log = SyncLog(
                sync_type="events_and_participants",
                started_at=started_at,
                completed_at=datetime.now(timezone.utc),
                status=sync_status,
                records_processed=success_count + participant_success,
                records_failed=error_count + participant_error,
                records_skipped=skipped_count,
                error_details=(
                    json.dumps(actual_errors[:100]) if actual_errors else None
                ),
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as e:
            print(f"Warning: Failed to record sync log: {e}")

        return jsonify(
            {
                "success": True,
                "message": f"Import completed successfully",
                "processed_count": success_count + participant_success,
                "error_count": error_count + participant_error,
                "skipped_count": skipped_count,
                "events_processed": success_count,
                "events_skipped": skipped_count,
                "participants_processed": participant_success,
                "total_events_from_salesforce": total_events,
                "total_participants_from_salesforce": len(participant_rows),
                "errors": actual_errors[:50],  # Limit errors in response
                "skipped_events": skipped_events[:50],  # Limit skips in response
            }
        )

    except SalesforceAuthenticationFailed:
        print("Error: Failed to authenticate with Salesforce")
        # Record failure log
        try:
            sync_log = SyncLog(
                sync_type="events_and_participants",
                started_at=(
                    started_at
                    if "started_at" in locals()
                    else datetime.now(timezone.utc)
                ),
                completed_at=datetime.now(timezone.utc),
                status=SyncStatus.FAILED.value,
                error_message="Failed to authenticate with Salesforce",
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as log_e:
            print(f"Warning: Failed to record error sync log: {log_e}")

        return (
            jsonify(
                {"success": False, "message": "Failed to authenticate with Salesforce"}
            ),
            401,
        )
    except Exception as e:
        db.session.rollback()
        print(f"Error: {str(e)}")
        # Record general failure log
        try:
            sync_log = SyncLog(
                sync_type="events_and_participants",
                started_at=(
                    started_at
                    if "started_at" in locals()
                    else datetime.now(timezone.utc)
                ),
                completed_at=datetime.now(timezone.utc),
                status=SyncStatus.FAILED.value,
                error_message=str(e),
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as log_e:
            print(f"Warning: Failed to record general failure sync log: {log_e}")

        return jsonify({"success": False, "error": str(e)}), 500


@events_bp.route("/events/sync-student-participants", methods=["POST"])
@login_required
@global_users_only
def sync_student_participants():
    """
    Sync student participation data from Salesforce

    This endpoint fetches student participation data from Salesforce and
    creates EventStudentParticipation records in the local database.

    Returns:
        JSON response with sync results and statistics
    """
    try:
        started_at = datetime.now(timezone.utc)
        print("Fetching student participation data from Salesforce...")
        success_count = 0
        error_count = 0
        errors = []

        # Connect to Salesforce
        sf = Salesforce(
            username=Config.SF_USERNAME,
            password=Config.SF_PASSWORD,
            security_token=Config.SF_SECURITY_TOKEN,
            domain="login",
        )

        # Query for Student participants
        participants_query = """
        SELECT
            Id,
            Name,
            Contact__c,
            Session__c,
            Status__c,
            Delivery_Hours__c,
            Age_Group__c,
            Email__c,
            Title__c
        FROM Session_Participant__c
        WHERE Participant_Type__c = 'Student'
        ORDER BY Session__c, Name
        """

        # Execute participants query
        participants_result = sf.query_all(participants_query)
        participant_rows = participants_result.get("records", [])

        print(
            f"Found {len(participant_rows)} student participation records in Salesforce."
        )

        # Process participants
        for row in participant_rows:
            success_count, error_count = process_student_participation_row(
                row, success_count, error_count, errors
            )

        # Commit changes (will only commit if association logic is added and adds to session)
        db.session.commit()

        # Print summary and errors
        print(
            f"\nSuccessfully processed {success_count} student participations with {error_count} total errors"
        )
        if errors:
            print("\nErrors encountered:")
            for error in errors[:10]:  # Only print first 10 errors
                print(f"- {error}")
            if len(errors) > 10:
                print(f"... and {len(errors) - 10} more errors")

        # Record sync log
        try:
            sync_status = SyncStatus.SUCCESS.value
            if error_count > 0:
                if success_count > 0:
                    sync_status = SyncStatus.PARTIAL.value
                else:
                    sync_status = SyncStatus.FAILED.value

            sync_log = SyncLog(
                sync_type="student_participants",
                started_at=started_at,
                completed_at=datetime.now(timezone.utc),
                status=sync_status,
                records_processed=success_count,
                records_failed=error_count,
                error_details=json.dumps(errors[:100]) if errors else None,
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as e:
            print(f"Warning: Failed to record sync log: {e}")

        return jsonify(
            {
                "success": True,
                "message": f"Processed {success_count} student participations with {error_count} errors.",
                "processed_count": success_count,
                "error_count": error_count,
                "successCount": success_count,
                "errorCount": error_count,
                "errors": errors,
            }
        )

    except SalesforceAuthenticationFailed:
        print("Error: Failed to authenticate with Salesforce")
        # Record failure log
        try:
            sync_log = SyncLog(
                sync_type="student_participants",
                started_at=(
                    started_at
                    if "started_at" in locals()
                    else datetime.now(timezone.utc)
                ),
                completed_at=datetime.now(timezone.utc),
                status=SyncStatus.FAILED.value,
                error_message="Failed to authenticate with Salesforce",
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as log_e:
            print(f"Warning: Failed to record error sync log: {log_e}")

        return (
            jsonify(
                {"success": False, "message": "Failed to authenticate with Salesforce"}
            ),
            401,
        )
    except Exception as e:
        db.session.rollback()
        print(f"Error in sync_student_participants: {str(e)}")
        # Record failure log
        try:
            sync_log = SyncLog(
                sync_type="student_participants",
                started_at=(
                    started_at
                    if "started_at" in locals()
                    else datetime.now(timezone.utc)
                ),
                completed_at=datetime.now(timezone.utc),
                status=SyncStatus.FAILED.value,
                error_message=str(e),
            )
            db.session.add(sync_log)
            db.session.commit()
        except Exception as log_e:
            print(f"Warning: Failed to record failure sync log: {log_e}")

        return jsonify({"success": False, "error": str(e)}), 500


@events_bp.route("/events/export/<int:id>")
@login_required
@global_users_only
def export_event_excel(id):
    """
    Export event data, including attended volunteer and student participation, as an Excel file.
    """
    event = db.session.get(Event, id)
    if not event:
        abort(404)

    # Get participations
    volunteer_participations = (
        EventParticipation.query.options(joinedload(EventParticipation.volunteer))
        .filter_by(event_id=id)
        .all()
    )
    student_participations = (
        EventStudentParticipation.query.options(
            joinedload(EventStudentParticipation.student)
        )
        .filter_by(event_id=id)
        .all()
    )

    # Only attended volunteers
    attended_volunteers = [
        p for p in volunteer_participations if p.status == "Attended" and p.volunteer
    ]
    # All students (optionally filter by status)

    wb = openpyxl.Workbook()
    ws_event = wb.active
    ws_event.title = "Event Info"

    # Event info
    ws_event.append(["Event Title", event.title])
    ws_event.append(
        [
            "Start Date",
            (
                event.start_date.strftime("%m/%d/%Y %I:%M %p")
                if event.start_date
                else "Not set"
            ),
        ]
    )
    ws_event.append(
        [
            "End Date",
            (
                event.end_date.strftime("%m/%d/%Y %I:%M %p")
                if event.end_date
                else "Not set"
            ),
        ]
    )
    ws_event.append(["Location", event.location or ""])
    ws_event.append(["Format", event.format.value.replace("_", " ").title()])
    ws_event.append(
        [
            "Status",
            event.status.value if hasattr(event.status, "value") else str(event.status),
        ]
    )
    ws_event.append(
        ["Type", event.type.value if hasattr(event.type, "value") else str(event.type)]
    )
    ws_event.append(["Student Count", event.participant_count or 0])
    ws_event.append(["Volunteer Count", len(attended_volunteers)])
    ws_event.append([])

    # Volunteers sheet
    ws_vols = wb.create_sheet("Attended Volunteers")
    ws_vols.append(["First Name", "Last Name", "Email", "Phone", "Delivery Hours"])
    for p in attended_volunteers:
        v = p.volunteer
        ws_vols.append(
            [
                v.first_name,
                v.last_name,
                getattr(v, "primary_email", ""),
                getattr(v, "primary_phone", ""),
                p.delivery_hours or "",
            ]
        )

    # Students sheet
    ws_students = wb.create_sheet("Students")
    ws_students.append(
        ["First Name", "Last Name", "Status", "Delivery Hours", "Age Group"]
    )
    for p in student_participations:
        s = p.student
        ws_students.append(
            [
                s.first_name if s else "",
                s.last_name if s else "",
                p.status or "",
                p.delivery_hours or "",
                p.age_group or "",
            ]
        )

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"event_{event.title.replace(' ', '_')}_{event.id}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@events_bp.route("/events/debug-counts", methods=["GET"])
@login_required
@admin_required
def debug_event_counts():
    """
    Debug route to check event counts and identify missing events
    """
    try:
        from simple_salesforce import Salesforce

        from config import Config

        # Database counts
        db_total = Event.query.count()
        db_by_status = {}
        db_by_type = {}

        for event in Event.query.all():
            status = event.status or "Unknown"
            event_type = event.type.value if event.type else "Unknown"
            db_by_status[status] = db_by_status.get(status, 0) + 1
            db_by_type[event_type] = db_by_type.get(event_type, 0) + 1

        # Salesforce counts (without importing)
        try:
            sf = Salesforce(
                username=Config.SF_USERNAME,
                password=Config.SF_PASSWORD,
                security_token=Config.SF_SECURITY_TOKEN,
                domain="login",
            )

            # Count total events
            total_query = "SELECT COUNT() FROM Session__c"
            total_result = sf.query(total_query)
            sf_total = total_result["totalSize"]

            # Count by status
            status_query = """
            SELECT Session_Status__c, COUNT(Id) total
            FROM Session__c
            GROUP BY Session_Status__c
            """
            status_result = sf.query(status_query)
            sf_by_status = {
                record["Session_Status__c"]: record["total"]
                for record in status_result["records"]
            }

            # Count by session type
            type_query = """
            SELECT Session_Type__c, COUNT(Id) total
            FROM Session__c
            GROUP BY Session_Type__c
            """
            type_result = sf.query(type_query)
            sf_by_type = {
                record["Session_Type__c"]: record["total"]
                for record in type_result["records"]
            }

            # Count filtered events (what we actually import)
            filtered_query = """
            SELECT COUNT()
            FROM Session__c
            WHERE Session_Status__c != 'Draft' AND Session_Type__c != 'Connector Session'
            """
            filtered_result = sf.query(filtered_query)
            sf_filtered = filtered_result["totalSize"]

        except Exception as e:
            sf_total = f"Error: {str(e)}"
            sf_by_status = {}
            sf_by_type = {}
            sf_filtered = "Error"

        return jsonify(
            {
                "database": {
                    "total": db_total,
                    "by_status": db_by_status,
                    "by_type": db_by_type,
                },
                "salesforce": {
                    "total": sf_total,
                    "filtered": sf_filtered,
                    "by_status": sf_by_status,
                    "by_type": sf_by_type,
                },
                "analysis": {
                    "missing_total": (
                        sf_total - db_total if isinstance(sf_total, int) else "Unknown"
                    ),
                    "missing_filtered": (
                        sf_filtered - db_total
                        if isinstance(sf_filtered, int)
                        else "Unknown"
                    ),
                },
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500
