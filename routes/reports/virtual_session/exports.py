"""
Excel export functions for virtual session reports.

Generates Excel files for teacher progress data with
summary, school summary, and detailed teacher sheets.
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_teacher_progress_excel(
    teacher_progress_data, district_name, virtual_year, date_from, date_to
):
    """
    Generate Excel file with teacher progress data including summary and detailed sheets.

    Args:
        teacher_progress_data: Dictionary with school progress data
        district_name: Name of the district
        virtual_year: Virtual year
        date_from: Start date
        date_to: End date

    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Create Summary Sheet
    summary_ws = wb.create_sheet("Summary")

    # Summary headers
    summary_headers = [
        "District",
        "Virtual Year",
        "Date Range",
        "Total Schools",
        "Total Teachers",
        "Goals Achieved",
        "Goals Achieved %",
        "In Progress",
        "In Progress %",
        "Not Started",
        "Not Started %",
    ]

    # Calculate totals
    total_schools = len(teacher_progress_data)
    total_teachers = sum(
        school_data["total_teachers"] for school_data in teacher_progress_data.values()
    )
    total_achieved = sum(
        school_data["goals_achieved"] for school_data in teacher_progress_data.values()
    )
    total_in_progress = sum(
        school_data["goals_in_progress"]
        for school_data in teacher_progress_data.values()
    )
    total_not_started = sum(
        school_data["goals_not_started"]
        for school_data in teacher_progress_data.values()
    )

    # Write summary headers
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write summary data
    summary_data = [
        district_name,
        virtual_year,
        f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}",
        total_schools,
        total_teachers,
        total_achieved,
        (
            f"{(total_achieved / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
        total_in_progress,
        (
            f"{(total_in_progress / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
        total_not_started,
        (
            f"{(total_not_started / total_teachers * 100):.1f}%"
            if total_teachers > 0
            else "0.0%"
        ),
    ]

    for col, value in enumerate(summary_data, 1):
        cell = summary_ws.cell(row=2, column=col, value=value)
        cell.border = border
        if col in [4, 5, 6, 8, 10]:  # Numeric columns
            cell.alignment = center_alignment

    # Auto-adjust column widths for summary
    for col in range(1, len(summary_headers) + 1):
        summary_ws.column_dimensions[get_column_letter(col)].width = 15

    # Create School Summary Sheet
    school_summary_ws = wb.create_sheet("School Summary")

    # School summary headers
    school_headers = [
        "School Name",
        "Total Teachers",
        "Goals Achieved",
        "Goals Achieved %",
        "In Progress",
        "In Progress %",
        "Not Started",
        "Not Started %",
    ]

    # Write school summary headers
    for col, header in enumerate(school_headers, 1):
        cell = school_summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write school summary data
    row = 2
    for school_name, school_data in teacher_progress_data.items():
        school_data_row = [
            school_name,
            school_data["total_teachers"],
            school_data["goals_achieved"],
            (
                f"{(school_data['goals_achieved'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
            school_data["goals_in_progress"],
            (
                f"{(school_data['goals_in_progress'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
            school_data["goals_not_started"],
            (
                f"{(school_data['goals_not_started'] / school_data['total_teachers'] * 100):.1f}%"
                if school_data["total_teachers"] > 0
                else "0.0%"
            ),
        ]

        for col, value in enumerate(school_data_row, 1):
            cell = school_summary_ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [2, 3, 4, 5, 6, 7, 8]:  # Numeric columns
                cell.alignment = center_alignment

        row += 1

    # Auto-adjust column widths for school summary
    for col in range(1, len(school_headers) + 1):
        school_summary_ws.column_dimensions[get_column_letter(col)].width = 18

    # Create Detailed Teacher Sheet
    teacher_detail_ws = wb.create_sheet("Teacher Details")

    # Teacher detail headers
    teacher_headers = [
        "School",
        "Teacher Name",
        "Email",
        "Grade",
        "Target Sessions",
        "Completed Sessions",
        "Planned Sessions",
        "Progress %",
        "Goal Status",
    ]

    # Write teacher detail headers
    for col, header in enumerate(teacher_headers, 1):
        cell = teacher_detail_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Write teacher detail data
    row = 2
    for school_name, school_data in teacher_progress_data.items():
        for teacher in school_data["teachers"]:
            teacher_row = [
                school_name,
                teacher["name"],
                teacher["email"],
                teacher["grade"],
                teacher["target_sessions"],
                teacher["completed_sessions"],
                teacher["planned_sessions"],
                f"{teacher['progress_percentage']:.1f}%",
                teacher["goal_status_text"],
            ]

            for col, value in enumerate(teacher_row, 1):
                cell = teacher_detail_ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 6, 7, 8]:  # Numeric columns
                    cell.alignment = center_alignment

            row += 1

    # Auto-adjust column widths for teacher details
    column_widths = [20, 25, 30, 8, 12, 15, 15, 12, 15]
    for col, width in enumerate(column_widths, 1):
        teacher_detail_ws.column_dimensions[get_column_letter(col)].width = width

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()
