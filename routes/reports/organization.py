from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from datetime import datetime, timedelta
from sqlalchemy import distinct, func, or_, desc, asc
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from models.organization import Organization, VolunteerOrganization
from models.volunteer import Volunteer, EventParticipation
from models.event import Event, EventType, EventFormat
from models.reports import OrganizationReport, OrganizationSummaryCache, OrganizationDetailCache
from models import db
from routes.reports.common import get_current_school_year, get_school_year_date_range

# Create blueprint
organization_bp = Blueprint('organization_reports', __name__)

def load_routes(bp):
    @bp.route('/reports/organization')
    @login_required
    def organization_reports():
        """Display all organizations with their engagement statistics - CACHED VERSION"""
        # Get filter parameters
        school_year = request.args.get('school_year', get_current_school_year())
        sort_by = request.args.get('sort_by', 'total_events')
        sort_order = request.args.get('sort_order', 'desc')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)  # Pagination for large datasets
        
        # Try to get cached data first
        cached_summary = OrganizationSummaryCache.query.filter_by(school_year=school_year).first()
        
        # Check if cache is fresh (less than 24 hours old)
        cache_is_fresh = (cached_summary and 
                         cached_summary.last_updated and 
                         (datetime.utcnow() - cached_summary.last_updated) < timedelta(hours=24))
        
        if cache_is_fresh and cached_summary:
            # Use cached data
            org_data = cached_summary.organizations_data
            last_updated = cached_summary.last_updated
        else:
            # Generate fresh data
            org_data = generate_organization_summary_data(school_year)
            last_updated = datetime.utcnow()
            
            # Cache the data
            cache_organization_summary_data(school_year, org_data)

        # Apply sorting (now in-memory since data is already retrieved)
        org_data = apply_organization_sorting(org_data, sort_by, sort_order)
        
        # Apply pagination
        total_orgs = len(org_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_orgs = org_data[start_idx:end_idx]
        
        # Calculate pagination info
        has_prev = page > 1
        has_next = end_idx < total_orgs
        prev_num = page - 1 if has_prev else None
        next_num = page + 1 if has_next else None
        total_pages = (total_orgs + per_page - 1) // per_page

        # Generate list of school years (from 2020-21 to current+1)
        current_year = int(get_current_school_year()[:2])
        school_years = [f"{y}{y+1}" for y in range(20, current_year + 2)]
        school_years.reverse()  # Most recent first

        return render_template(
            'reports/organization.html',
            organizations=paginated_orgs,
            school_year=school_year,
            school_years=school_years,
            sort_by=sort_by,
            sort_order=sort_order,
            # Pagination variables
            page=page,
            per_page=per_page,
            total_orgs=total_orgs,
            has_prev=has_prev,
            has_next=has_next,
            prev_num=prev_num,
            next_num=next_num,
            total_pages=total_pages,
            # Cache info
            last_updated=last_updated,
            cache_is_fresh=cache_is_fresh,
            now=datetime.now()
        )

    @bp.route('/reports/organization/detail/<int:org_id>')
    @login_required
    def organization_detail(org_id):
        """Display detailed report for a specific organization - CACHED VERSION"""
        # Get filter parameters
        school_year = request.args.get('school_year', get_current_school_year())
        sort_by = request.args.get('sort_by', 'date')
        sort_order = request.args.get('sort_order', 'desc')
        
        # Get organization details
        organization = Organization.query.get_or_404(org_id)
        
        # Try to get cached data first
        cached_detail = OrganizationDetailCache.query.filter_by(
            organization_id=org_id, 
            school_year=school_year
        ).first()
        
        # Check if cache is fresh (less than 24 hours old)
        cache_is_fresh = (cached_detail and 
                         cached_detail.last_updated and 
                         (datetime.utcnow() - cached_detail.last_updated) < timedelta(hours=24))
        
        if cache_is_fresh and cached_detail:
            # Use cached data
            in_person_events = cached_detail.in_person_events or []
            virtual_events = cached_detail.virtual_events or []
            cancelled_events_data = cached_detail.cancelled_events or []
            volunteers_data = cached_detail.volunteers_data or []
            summary = cached_detail.summary_stats or {}
            last_updated = cached_detail.last_updated
        else:
            # Generate fresh data
            detail_data = generate_organization_detail_data(org_id, school_year)
            in_person_events = detail_data['in_person_events']
            virtual_events = detail_data['virtual_events']
            cancelled_events_data = detail_data['cancelled_events']
            volunteers_data = detail_data['volunteers_data']
            summary = detail_data['summary']
            last_updated = datetime.utcnow()
            
            # Cache the data
            cache_organization_detail_data(org_id, school_year, organization.name, detail_data)

        # Apply sorting to cached data (in-memory sorting is fast for detail pages)
        in_person_events = apply_detail_sorting(in_person_events, sort_by, sort_order)
        virtual_events = apply_detail_sorting(virtual_events, sort_by, sort_order)
        cancelled_events_data = apply_detail_sorting(cancelled_events_data, sort_by, sort_order)

        # Generate list of school years (from 2020-21 to current+1)
        current_year = int(get_current_school_year()[:2])
        school_years = [f"{y}{y+1}" for y in range(20, current_year + 2)]
        school_years.reverse()

        return render_template(
            'reports/organization_detail.html',
            organization=organization,
            in_person_events=in_person_events,
            virtual_events=virtual_events,
            cancelled_events=cancelled_events_data,
            volunteers=volunteers_data,
            summary=summary,
            school_year=school_year,
            school_years=school_years,
            sort_by=sort_by,
            sort_order=sort_order,
            # Cache info
            last_updated=last_updated,
            cache_is_fresh=cache_is_fresh,
            now=datetime.now()
        )

    @bp.route('/reports/organization/refresh', methods=['POST'])
    @login_required
    def refresh_organization_reports():
        """Refresh the cached organization report data"""
        school_year = request.args.get('school_year', get_current_school_year())
        
        try:
            # Delete existing cached data for this school year
            OrganizationSummaryCache.query.filter_by(school_year=school_year).delete()
            db.session.commit()
            
            # Generate new data
            org_data = generate_organization_summary_data(school_year)
            
            # Cache the new data
            cache_organization_summary_data(school_year, org_data)
            
            return jsonify({
                'success': True, 
                'message': f'Successfully refreshed organization data for {school_year[:2]}-{school_year[2:]} school year'
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @bp.route('/reports/organization/detail/<int:org_id>/refresh', methods=['POST'])
    @login_required
    def refresh_organization_detail(org_id):
        """Refresh the cached organization detail data"""
        school_year = request.args.get('school_year', get_current_school_year())
        
        try:
            # Get organization details
            organization = Organization.query.get_or_404(org_id)
            
            # Delete existing cached data for this organization and school year
            OrganizationDetailCache.query.filter_by(
                organization_id=org_id, 
                school_year=school_year
            ).delete()
            db.session.commit()
            
            # Generate new data
            detail_data = generate_organization_detail_data(org_id, school_year)
            
            # Cache the new data
            cache_organization_detail_data(org_id, school_year, organization.name, detail_data)
            
            return jsonify({
                'success': True, 
                'message': f'Successfully refreshed detail data for {organization.name}'
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    @bp.route('/reports/organization/detail/<int:org_id>/export')
    @login_required
    def export_organization_detail_excel(org_id):
        """Export organization detail data to Excel"""
        school_year = request.args.get('school_year', get_current_school_year())
        
        # Get organization details
        organization = Organization.query.get_or_404(org_id)
        
        # Try to get cached data first, fallback to generating fresh data
        cached_detail = OrganizationDetailCache.query.filter_by(
            organization_id=org_id, 
            school_year=school_year
        ).first()
        
        if cached_detail:
            # Use cached data
            in_person_events = cached_detail.in_person_events or []
            virtual_events = cached_detail.virtual_events or []
            cancelled_events_data = cached_detail.cancelled_events or []
            volunteers_data = cached_detail.volunteers_data or []
            summary = cached_detail.summary_stats or {}
        else:
            # Generate fresh data
            detail_data = generate_organization_detail_data(org_id, school_year)
            in_person_events = detail_data['in_person_events']
            virtual_events = detail_data['virtual_events']
            cancelled_events_data = detail_data['cancelled_events']
            volunteers_data = detail_data['volunteers_data']
            summary = detail_data['summary']
        
        # Create Excel workbook
        output = io.BytesIO()
        
        try:
            # Create workbook and add sheets
            wb = Workbook()
            
            # Remove default sheet
            if wb.active:
                wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1d3354", end_color="1d3354", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # 1. Summary Sheet
            ws_summary = wb.create_sheet("Summary")
            summary_data = [
                ["Organization", organization.name],
                ["School Year", f"20{school_year[:2]}-20{school_year[2:]}"],
                ["Report Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p")],
                ["", ""],  # Empty row with two values
                ["Total Volunteers Engaged", summary.get('total_volunteers', 0)],
                ["In-Person Events", summary.get('total_in_person_events', 0)],
                ["Students Reached (In-Person)", summary.get('total_in_person_students', 0)],
                ["Virtual Events", summary.get('total_virtual_events', 0)],
                ["Virtual Classrooms Reached", summary.get('total_virtual_classrooms', 0)],
                ["Cancelled Events", summary.get('total_cancelled_events', 0)],
                ["", ""],  # Empty row with two values
                ["Total Events", summary.get('total_in_person_events', 0) + summary.get('total_virtual_events', 0)],
                ["Total Students/Classrooms Reached", summary.get('total_in_person_students', 0) + summary.get('total_virtual_classrooms', 0)]
            ]
            
            for row_idx, row_data in enumerate(summary_data, 1):
                if len(row_data) >= 2:
                    label, value = row_data[0], row_data[1]
                    ws_summary.cell(row=row_idx, column=1, value=label)
                    ws_summary.cell(row=row_idx, column=2, value=value)
                    
                    # Style the first column (labels)
                    if label:  # Skip empty rows
                        ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
            
            # Auto-size columns
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            
            # 2. In-Person Events Sheet
            if in_person_events:
                ws_in_person = wb.create_sheet("In-Person Events")
                
                # Headers
                headers = ["Date", "Volunteer", "Session", "Event Type", "Students Reached"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_in_person.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row_idx, event in enumerate(in_person_events, 2):
                    ws_in_person.cell(row=row_idx, column=1, value=event.get('date', ''))
                    ws_in_person.cell(row=row_idx, column=2, value=event.get('volunteer', ''))
                    ws_in_person.cell(row=row_idx, column=3, value=event.get('session', ''))
                    ws_in_person.cell(row=row_idx, column=4, value=event.get('event_type', ''))
                    ws_in_person.cell(row=row_idx, column=5, value=event.get('students_reached', 0))
                    
                    # Add borders
                    for col_idx in range(1, 6):
                        ws_in_person.cell(row=row_idx, column=col_idx).border = border
                
                # Auto-size columns
                for col in ws_in_person.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_in_person.column_dimensions[column].width = adjusted_width
            
            # 3. Virtual Events Sheet
            if virtual_events:
                ws_virtual = wb.create_sheet("Virtual Events")
                
                # Headers
                headers = ["Date", "Time", "Volunteer", "Session", "Event Type", "Students", "Session ID"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_virtual.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row_idx, event in enumerate(virtual_events, 2):
                    ws_virtual.cell(row=row_idx, column=1, value=event.get('date', ''))
                    ws_virtual.cell(row=row_idx, column=2, value=event.get('time', ''))
                    ws_virtual.cell(row=row_idx, column=3, value=event.get('volunteer', ''))
                    ws_virtual.cell(row=row_idx, column=4, value=event.get('session', ''))
                    ws_virtual.cell(row=row_idx, column=5, value=event.get('event_type', ''))
                    ws_virtual.cell(row=row_idx, column=6, value=event.get('classrooms', 0))
                    ws_virtual.cell(row=row_idx, column=7, value=event.get('session_id', ''))
                    
                    # Add borders
                    for col_idx in range(1, 8):
                        ws_virtual.cell(row=row_idx, column=col_idx).border = border
                
                # Auto-size columns
                for col in ws_virtual.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_virtual.column_dimensions[column].width = adjusted_width
            
            # 4. Volunteer Summary Sheet
            if volunteers_data:
                ws_volunteers = wb.create_sheet("Volunteer Summary")
                
                # Headers
                headers = ["Volunteer Name", "Number of Events", "Total Hours"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_volunteers.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row_idx, volunteer in enumerate(volunteers_data, 2):
                    ws_volunteers.cell(row=row_idx, column=1, value=volunteer.get('name', ''))
                    ws_volunteers.cell(row=row_idx, column=2, value=volunteer.get('events', 0))
                    ws_volunteers.cell(row=row_idx, column=3, value=volunteer.get('hours', 0))
                    
                    # Add borders
                    for col_idx in range(1, 4):
                        ws_volunteers.cell(row=row_idx, column=col_idx).border = border
                
                # Auto-size columns
                ws_volunteers.column_dimensions['A'].width = 30
                ws_volunteers.column_dimensions['B'].width = 18
                ws_volunteers.column_dimensions['C'].width = 15
            
            # 5. Cancelled Events Sheet (if any)
            if cancelled_events_data:
                ws_cancelled = wb.create_sheet("Cancelled Events")
                
                # Headers
                headers = ["Event", "Volunteer", "Cancellation Reason"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_cancelled.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                for row_idx, event in enumerate(cancelled_events_data, 2):
                    ws_cancelled.cell(row=row_idx, column=1, value=event.get('event', ''))
                    ws_cancelled.cell(row=row_idx, column=2, value=event.get('volunteer', ''))
                    ws_cancelled.cell(row=row_idx, column=3, value=event.get('cancellation_reason', ''))
                    
                    # Add borders
                    for col_idx in range(1, 4):
                        ws_cancelled.cell(row=row_idx, column=col_idx).border = border
                
                # Auto-size columns
                for col in ws_cancelled.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_cancelled.column_dimensions[column].width = adjusted_width
            
            # Save workbook to BytesIO
            wb.save(output)
            output.seek(0)
            
            # Generate filename
            safe_org_name = "".join(c for c in organization.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            filename = f"{safe_org_name}_Volunteer_Report_{school_year[:2]}-{school_year[2:]}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            # Return error response
            return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500


# Helper functions for caching and data generation

def generate_organization_summary_data(school_year):
    """Generate organization summary statistics"""
    start_date, end_date = get_school_year_date_range(school_year)
    
    # Query organization engagement statistics (optimized query)
    org_stats = db.session.query(
        Organization,
        func.count(distinct(Event.id)).label('total_events'),
        func.count(distinct(Volunteer.id)).label('unique_volunteers'),
        func.sum(EventParticipation.delivery_hours).label('total_hours')
    ).join(
        VolunteerOrganization, Organization.id == VolunteerOrganization.organization_id
    ).join(
        Volunteer, VolunteerOrganization.volunteer_id == Volunteer.id
    ).join(
        EventParticipation, Volunteer.id == EventParticipation.volunteer_id
    ).join(
        Event, EventParticipation.event_id == Event.id
    ).filter(
        Event.start_date >= start_date,
        Event.start_date <= end_date,
        EventParticipation.status == 'Attended'
    ).group_by(
        Organization.id
    ).all()

    # Get cancelled events count for each organization (separate query for efficiency)
    cancelled_stats = db.session.query(
        Organization.id,
        func.count(distinct(Event.id)).label('cancelled_events')
    ).join(
        VolunteerOrganization, Organization.id == VolunteerOrganization.organization_id
    ).join(
        Volunteer, VolunteerOrganization.volunteer_id == Volunteer.id
    ).join(
        EventParticipation, Volunteer.id == EventParticipation.volunteer_id
    ).join(
        Event, EventParticipation.event_id == Event.id
    ).filter(
        Event.start_date >= start_date,
        Event.start_date <= end_date,
        Event.status == 'Cancelled'
    ).group_by(
        Organization.id
    ).all()

    # Create a dictionary for quick lookup of cancelled events
    cancelled_dict = {org_id: count for org_id, count in cancelled_stats}

    # Format the data for caching
    org_data = [{
        'id': org.id,
        'name': org.name,
        'total_events': total_events or 0,
        'unique_volunteers': unique_volunteers or 0,
        'total_hours': round(float(total_hours or 0), 2),
        'cancelled_events': cancelled_dict.get(org.id, 0)
    } for org, total_events, unique_volunteers, total_hours in org_stats]
    
    return org_data


def cache_organization_summary_data(school_year, org_data):
    """Cache organization summary data"""
    # Delete existing cache entry
    OrganizationSummaryCache.query.filter_by(school_year=school_year).delete()
    
    # Create new cache entry
    cache_entry = OrganizationSummaryCache(
        school_year=school_year,
        organizations_data=org_data,
        last_updated=datetime.utcnow()
    )
    
    db.session.add(cache_entry)
    db.session.commit()


def apply_organization_sorting(org_data, sort_by, sort_order):
    """Apply sorting to organization data"""
    reverse = sort_order == 'desc'
    
    if sort_by == 'name':
        return sorted(org_data, key=lambda x: x.get('name', ''), reverse=reverse)
    elif sort_by == 'total_events':
        return sorted(org_data, key=lambda x: x.get('total_events', 0), reverse=reverse)
    elif sort_by == 'unique_volunteers':
        return sorted(org_data, key=lambda x: x.get('unique_volunteers', 0), reverse=reverse)
    elif sort_by == 'total_hours':
        return sorted(org_data, key=lambda x: x.get('total_hours', 0), reverse=reverse)
    elif sort_by == 'cancelled_events':
        return sorted(org_data, key=lambda x: x.get('cancelled_events', 0), reverse=reverse)
    else:
        # Default sort by total_events desc
        return sorted(org_data, key=lambda x: x.get('total_events', 0), reverse=True)


def generate_organization_detail_data(org_id, school_year):
    """Generate detailed organization data for caching"""
    from datetime import datetime
    
    # Get date range for the school year
    start_date, end_date = get_school_year_date_range(school_year)
    
    # Get attended events for this organization
    attended_events_query = db.session.query(
        Event,
        EventParticipation.delivery_hours,
        Volunteer
    ).join(
        EventParticipation, Event.id == EventParticipation.event_id
    ).join(
        Volunteer, EventParticipation.volunteer_id == Volunteer.id
    ).join(
        VolunteerOrganization, Volunteer.id == VolunteerOrganization.volunteer_id
    ).filter(
        VolunteerOrganization.organization_id == org_id,
        Event.start_date >= start_date,
        Event.start_date <= end_date,
        db.or_(
            EventParticipation.status == 'Attended',
            EventParticipation.status == 'Completed', 
            EventParticipation.status == 'Successfully Completed',
            EventParticipation.status == 'Confirmed'
        )
    ).order_by(desc(Event.start_date)).all()

    # Get cancelled events for this organization
    cancelled_events = db.session.query(
        Event,
        Volunteer
    ).join(
        EventParticipation, Event.id == EventParticipation.event_id
    ).join(
        Volunteer, EventParticipation.volunteer_id == Volunteer.id
    ).join(
        VolunteerOrganization, Volunteer.id == VolunteerOrganization.volunteer_id
    ).filter(
        VolunteerOrganization.organization_id == org_id,
        Event.start_date >= start_date,
        Event.start_date <= end_date,
        Event.status == 'Cancelled'
    ).order_by(Event.start_date.desc()).all()

    # Get organization volunteers with their engagement stats
    volunteer_stats = db.session.query(
        Volunteer,
        func.count(distinct(Event.id)).label('event_count'),
        func.sum(EventParticipation.delivery_hours).label('total_hours')
    ).join(
        VolunteerOrganization, Volunteer.id == VolunteerOrganization.volunteer_id
    ).join(
        EventParticipation, Volunteer.id == EventParticipation.volunteer_id
    ).join(
        Event, EventParticipation.event_id == Event.id
    ).filter(
        VolunteerOrganization.organization_id == org_id,
        Event.start_date >= start_date,
        Event.start_date <= end_date,
        db.or_(
            EventParticipation.status == 'Attended',
            EventParticipation.status == 'Completed', 
            EventParticipation.status == 'Successfully Completed',
            EventParticipation.status == 'Confirmed'
        )
    ).group_by(
        Volunteer.id
    ).order_by(
        db.desc('total_hours')
    ).all()

    # Separate in-person and virtual events from attended events
    in_person_events = []
    virtual_events = []
    
    # Define which event types are considered in-person vs virtual
    in_person_types = {
        EventType.IN_PERSON,
        EventType.CAREER_JUMPING,
        EventType.CAREER_SPEAKER,
        EventType.EMPLOYABILITY_SKILLS,
        EventType.IGNITE,
        EventType.CAREER_FAIR,
        EventType.PATHWAY_CAMPUS_VISITS,
        EventType.WORKPLACE_VISIT,
        EventType.PATHWAY_WORKPLACE_VISITS,
        EventType.COLLEGE_OPTIONS,
        EventType.DIA_CLASSROOM_SPEAKER,
        EventType.DIA,
        EventType.CAMPUS_VISIT,
        EventType.ADVISORY_SESSIONS,
        EventType.MENTORING,
        EventType.FINANCIAL_LITERACY,
        EventType.MATH_RELAYS,
        EventType.CLASSROOM_SPEAKER,
        EventType.INTERNSHIP,
        EventType.COLLEGE_APPLICATION_FAIR,
        EventType.FAFSA,
        EventType.CLASSROOM_ACTIVITY,
        EventType.VOLUNTEER_ORIENTATION,
        EventType.VOLUNTEER_ENGAGEMENT,
        EventType.CLIENT_CONNECTED_PROJECT
    }
    
    virtual_types = {
        EventType.VIRTUAL_SESSION,
        EventType.CONNECTOR_SESSION,
        EventType.DATA_VIZ,  # Data visualization sessions are often virtual
        EventType.P2GD,      # Prep to Goal Directed 
        EventType.SLA,       # Student Learning Activities
        EventType.HEALTHSTART,
        EventType.P2T,       # Prep to Trade
        EventType.BFI,       # Business and Finance Institute
        EventType.HISTORICAL
    }
    
    for event, hours, volunteer in attended_events_query:
        event_data = {
            'date': event.start_date.strftime('%m/%d/%y') if event.start_date else '',
            'volunteer': f"{volunteer.first_name} {volunteer.last_name}",
            'volunteer_id': volunteer.id,
            'session': event.title,
            'event_id': event.id,
            'hours': round(float(hours or 0), 2),
            'students_reached': event.participant_count or 0,
            'event_type': event.type.value if event.type else 'Unknown'
        }
        
        # Check both event type and format to determine if virtual
        is_virtual = (event.type in virtual_types or 
                     event.format == EventFormat.VIRTUAL)
        
        if is_virtual:
            event_data['time'] = event.start_date.strftime('%I:%M %p') if event.start_date else ''
            # For virtual events, show estimated student participation or registered count
            event_data['classrooms'] = event.participant_count or event.registered_count or 0
            # Add additional virtual session details
            event_data['session_id'] = event.session_id or ''
            event_data['series'] = event.series or ''
            virtual_events.append(event_data)
        elif event.type in in_person_types or event.format == EventFormat.IN_PERSON:
            in_person_events.append(event_data)
        else:
            # For unknown types, default to in-person but log the type
            event_data['note'] = f"Unknown event type: {event.type}"
            in_person_events.append(event_data)

    # Process cancelled events
    cancelled_events_data = []
    for event, volunteer in cancelled_events:
        cancelled_events_data.append({
            'event': event.title,
            'volunteer': f"{volunteer.first_name} {volunteer.last_name}",
            'volunteer_id': volunteer.id,
            'event_id': event.id,
            'date': event.start_date.strftime('%m/%d/%y') if event.start_date else '',
            'cancellation_reason': event.cancellation_reason.value if event.cancellation_reason else 'Unknown'
        })

    # Format volunteer data
    volunteers_data = [{
        'name': f"{v.first_name} {v.last_name}",
        'events': events,
        'hours': round(float(hours or 0), 2)
    } for v, events, hours in volunteer_stats]

    # Calculate summary statistics
    total_in_person_students = sum(item['students_reached'] for item in in_person_events)
    total_virtual_classrooms = sum(item.get('classrooms', 0) for item in virtual_events)
    total_volunteers = len(volunteers_data)
    total_in_person_events = len(in_person_events)
    total_virtual_events = len(virtual_events)

    summary = {
        'total_volunteers': total_volunteers,
        'total_in_person_events': total_in_person_events,
        'total_virtual_events': total_virtual_events,
        'total_in_person_students': total_in_person_students,
        'total_virtual_classrooms': total_virtual_classrooms,
        'total_cancelled_events': len(cancelled_events_data)
    }

    return {
        'in_person_events': in_person_events,
        'virtual_events': virtual_events,
        'cancelled_events': cancelled_events_data,
        'volunteers_data': volunteers_data,
        'summary': summary
    }


def cache_organization_detail_data(org_id, school_year, org_name, detail_data):
    """Cache organization detail data"""
    try:
        # Delete existing cache for this organization and school year
        OrganizationDetailCache.query.filter_by(
            organization_id=org_id, 
            school_year=school_year
        ).delete()
        
        # Create new cache entry
        cache_entry = OrganizationDetailCache(
            organization_id=org_id,
            school_year=school_year,
            organization_name=org_name,
            in_person_events=detail_data['in_person_events'],
            virtual_events=detail_data['virtual_events'],
            cancelled_events=detail_data['cancelled_events'],
            volunteers_data=detail_data['volunteers_data'],
            summary_stats=detail_data['summary'],
            last_updated=datetime.utcnow()
        )
        
        db.session.add(cache_entry)
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"Error caching organization detail data: {e}")


def apply_detail_sorting(events_data, sort_by, sort_order):
    """Apply sorting to detail page event data"""
    if not events_data:
        return events_data
        
    reverse = sort_order == 'desc'
    
    if sort_by == 'date':
        return sorted(events_data, key=lambda x: x.get('date', ''), reverse=reverse)
    elif sort_by == 'volunteer':
        return sorted(events_data, key=lambda x: x.get('volunteer', ''), reverse=reverse)
    elif sort_by == 'students':
        # For in-person events use 'students_reached', for virtual use 'classrooms'
        return sorted(events_data, key=lambda x: x.get('students_reached', x.get('classrooms', 0)), reverse=reverse)
    elif sort_by == 'event_type':
        return sorted(events_data, key=lambda x: x.get('event_type', ''), reverse=reverse)
    else:
        # Default sort by date desc
        return sorted(events_data, key=lambda x: x.get('date', ''), reverse=True)
